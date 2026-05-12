#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
kanban_csv2excel_novo_layout.py — Migração CSV → Kanbans_Producao_NOVO.xlsx
Metalogalva — Irmãos Silvas, S.A.

Suporta dois formatos de CSV gerados pelo sistema OCR:

── FORMATO A  (=== marcadores) ─────────────────────────────────────────────
Ficheiros: *_kanban.csv  (ex.: JulioLima_2026.04.02_kanban.csv)
Gerado por: kanban_ocr.py / kanban_claude_vision.py (versão anterior)

    linha 1  : FICHEIRO;INICIO;FIM;DURACAO_SEG   ← metadados de execução
    === CABEÇALHO ===
    OPERADOR;SETOR/MÁQUINA;DATA;NÚMERO;TURNO
    <valores>
    === TABELA DE PRODUÇÃO ===
    PRI;CLIENTE;OV;OF;MODELO;QTD COLUNAS;COMP (mm);LARG (mm);LOTE;CONI;ESP;LBASE;LTOPO
    <linhas…>
    === RODAPÉ ===
    COLUNAS PRODUZIDAS;HORAS TRABALHADAS
    <valores>

── FORMATO B  (### marcadores) ─────────────────────────────────────────────
Ficheiros: Operador_AAAA.MM.DD.csv  (ex.: AugustoMonteiro_2026.04.09.csv)
Gerado por: skill kanban-ocr (versão actual)

    ### CABEÇALHO
    FICHEIRO;DATA;OPERADOR;N_OPERADOR;SETOR_MAQUINA
    <valores>
    ### TABELA DE PRODUÇÃO
    FICHEIRO;DATA;OPERADOR;PRI;CLIENTE;OV;OF;MODELO;QTD;COMP_MM;LARG_MM;LOTE;CONI;ESP;LBASE;LTOPO
    <linhas…>
    ### RODAPÉ
    FICHEIRO;DATA;OPERADOR;COLUNAS_PRODUZIDAS;HORAS_TRABALHADAS
    <valores>

O formato é detectado automaticamente. Qualquer .csv na pasta csv/ é processado.

Regras de validação:
  1. StockSAP  — confirma que o LOTE existe em stock e que LARG (±10 mm) e
                 ESP (exacta) batem certo com o OCR.
  2. Plano     — valida cruzada com plan_colunas_cpis.xlsx:
                 OF, OV, Modelo, Cliente, Espessura, Largura e Comprimento.

Comportamento em caso de falha: regista a linha com indicador visual
colorido na coluna "Validação" e notas em "Notas". Não bloqueia.

Execução (Windows):
    cd "06_Kanban_OCR"
    python kanban_csv2excel_novo_layout.py

Dependências:
    pip install openpyxl

Versão: 1.0  |  2026-04-20  (suporte dual-formato: === e ###)
"""

import os
import csv
import sys
import shutil
import logging
import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("ERRO: openpyxl não instalado. Execute: pip install openpyxl")

# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO DE CAMINHOS
# ═══════════════════════════════════════════════════════════════════════════

# O script está em 06_Kanban_OCR/; a raiz do projecto é a pasta-pai.
RAIZ       = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIR_DADOS  = os.path.join(RAIZ, "02_Dados_Extraidos")
DIR_CSV    = os.path.join(DIR_DADOS, "csv")         # CSVs gerados pelo OCR
DIR_IMPORT = os.path.join(DIR_CSV,   "imported")    # CSVs já processados
DIR_DOC    = os.path.join(RAIZ, "04_Documentacao")

EXCEL_DEST = os.path.join(DIR_DADOS, "Kanbans_Producao_NOVO.xlsx")
STOCK_SAP  = os.path.join(DIR_DOC,   "StockSAP.xlsx")
PLAN_COL   = os.path.join(DIR_DOC,   "plan_colunas_cpis.xlsx")
LOG_FILE   = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "kanban_csv2excel_novo_layout.log")

# ─── Tolerâncias de validação ─────────────────────────────────────────────
TOL_LARG_MM = 10    # ±10 mm para Largura (StockSAP e plano)
TOL_COMP_MM = 50    # ±50 mm para Comprimento (plano)
# Espessura: comparação exacta (diferença > 0.01 mm já é assinalada)

# ─── Cores ───────────────────────────────────────────────────────────────
COR_HEADER_FILL = "2F5597"
COR_HEADER_FONT = "FFFFFF"
COR_OK          = "C6EFCE"   # verde  — tudo confere
COR_AVISO       = "FFEB9C"   # amarelo — sem match ou aviso
COR_DIVERGENCIA = "FFC7CE"   # vermelho — divergência confirmada
COR_ALT         = "DCE6F1"   # azul claro — linhas alternadas sem problema
COR_BRANCO      = "FFFFFF"

# ═══════════════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════
# UTILITÁRIOS DE ESTILO EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def _fill(rgb):
    return PatternFill("solid", fgColor=rgb)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _estilo(cell, bg_rgb, bold=False, wrap=False):
    cell.fill      = _fill(bg_rgb)
    cell.font      = Font(name="Arial", size=9, bold=bold)
    cell.border    = _border()
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)

# ═══════════════════════════════════════════════════════════════════════════
# CARREGAR FICHEIROS DE REFERÊNCIA
# ═══════════════════════════════════════════════════════════════════════════

def carregar_stocksap():
    """
    Lê StockSAP.xlsx e devolve um dicionário:
        { lote_normalizado_upper: {esp, larg, qtd, desc} }
    """
    if not os.path.exists(STOCK_SAP):
        log.warning("StockSAP.xlsx não encontrado — validação de lotes desactivada")
        return {}

    log.info("A carregar StockSAP.xlsx...")
    wb = openpyxl.load_workbook(STOCK_SAP, read_only=True, data_only=True)
    ws = wb.active

    # Mapear cabeçalhos (case-insensitive)
    hdrs = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for j, v in enumerate(row):
            if v is not None:
                hdrs[str(v).strip().upper()] = j

    stock = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        lote = row[hdrs.get("LOTE", 0)]
        if not lote:
            continue
        # Suporta colunas "ESPESSURA" ou "ESP", "LARGURA" ou "LARG"
        idx_esp  = hdrs.get("ESPESSURA",  hdrs.get("ESP",  2))
        idx_larg = hdrs.get("LARGURA",    hdrs.get("LARG", 3))
        idx_qtd  = hdrs.get("QTD",  hdrs.get("QUANTIDADE", 1))
        idx_desc = hdrs.get("DESCRIÇÃO ADICIONAL", hdrs.get("DESCRICAO", hdrs.get("DESC", 4)))
        stock[str(lote).strip().upper()] = {
            "esp":  row[idx_esp],
            "larg": row[idx_larg],
            "qtd":  row[idx_qtd],
            "desc": row[idx_desc],
        }

    wb.close()
    log.info("  StockSAP: %d lotes carregados", len(stock))
    return stock


def carregar_plancolunas():
    """
    Lê plan_colunas_cpis.xlsx e devolve um dicionário:
        { of_int: [ {cliente, ov, designacao, esp, lbase, ltopo, ltotal, comp, ...} ] }
    A chave é o OF como inteiro; o valor é uma lista (pode haver várias
    entradas por OF em OFs multi-parte).
    """
    if not os.path.exists(PLAN_COL):
        log.warning("plan_colunas_cpis.xlsx não encontrado — validação de plano desactivada")
        return {}

    log.info("A carregar plan_colunas_cpis.xlsx...")
    wb = openpyxl.load_workbook(PLAN_COL, read_only=True, data_only=True)
    ws = wb.active

    hdrs = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for j, v in enumerate(row):
            if v is not None:
                hdrs[str(v).strip().lower()] = j

    plano = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        of_raw = row[hdrs["of"]] if "of" in hdrs else None
        try:
            of_int = int(float(str(of_raw)))
        except (ValueError, TypeError):
            continue

        def _v(k, alt=None):
            return row[hdrs[k]] if k in hdrs else alt

        plano.setdefault(of_int, []).append({
            "cliente":    str(_v("cliente") or "").strip().upper(),
            "ov":         str(_v("ov") or "").strip(),
            "designacao": str(_v("designacao") or "").strip(),
            "esp":        _v("esp"),
            "lbase":      _v("lbase"),
            "ltopo":      _v("ltopo"),
            "ltotal":     _v("ltotal"),
            "comp":       _v("comp"),
            "quanttrp":   _v("quanttrp"),
            "bf":         _v("bf"),
        })

    wb.close()
    log.info("  plan_colunas_cpis: %d OFs carregados", len(plano))
    return plano

# ═══════════════════════════════════════════════════════════════════════════
# LÓGICA DE VALIDAÇÃO
# ═══════════════════════════════════════════════════════════════════════════

def _num(v):
    """Converte para float; devolve None se não for possível."""
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def validar_stocksap(row, stock):
    """
    Regra 1 — Confirma que o LOTE existe em stock e que LARG (±10 mm)
    e ESP (exacta) batem certo com os valores do OCR no CSV.

    Devolve: (status_str, lista_de_notas)
    """
    lote = str(row.get("lote", "") or "").strip().upper()

    if not lote or lote in ("?", ""):
        return "SEM LOTE", ["LOTE: não preenchido pelo OCR"]

    if lote not in stock:
        return "LOTE NÃO ENCONTRADO", [f"LOTE {lote}: não existe em StockSAP"]

    entrada = stock[lote]
    notas   = []
    status  = "OK"

    # — Largura: OCR vs StockSAP, tolerância ±TOL_LARG_MM
    larg_ocr = _num(row.get("largura_mm"))
    larg_sap = _num(entrada.get("larg"))
    if larg_ocr is not None and larg_sap is not None:
        diff = abs(larg_ocr - larg_sap)
        if diff > TOL_LARG_MM:
            status = "DIVERGÊNCIA"
            notas.append(
                f"LARG: OCR={larg_ocr:.0f} SAP={larg_sap:.0f} "
                f"(dif={diff:.0f}mm > ±{TOL_LARG_MM}mm)"
            )
    elif larg_sap is not None:
        notas.append(f"LARG: OCR não legível; SAP={larg_sap:.0f}mm")

    # — Espessura: exacta
    esp_ocr = _num(row.get("espessura_mm"))
    esp_sap = _num(entrada.get("esp"))
    if esp_ocr is not None and esp_sap is not None:
        if abs(esp_ocr - esp_sap) > 0.01:
            status = "DIVERGÊNCIA"
            notas.append(f"ESP: OCR={esp_ocr} SAP={esp_sap} (não coincide)")

    return status, notas


def _melhor_entrada(entradas, modelo_ocr, comp_ocr):
    """
    Escolhe a entrada do plano mais adequada para este OCR:
      1. Tenta correspondência pelo código do modelo na designação.
      2. Se não encontrar, escolhe a mais próxima em comprimento.
      3. Senão, a primeira.
    """
    if not entradas:
        return None

    if modelo_ocr:
        for e in entradas:
            if modelo_ocr.upper() in e["designacao"].upper():
                return e

    if comp_ocr is not None:
        candidatos = [e for e in entradas if _num(e.get("comp")) is not None]
        if candidatos:
            return min(candidatos,
                       key=lambda e: abs((_num(e["comp"]) or 0) - comp_ocr))

    return entradas[0]


def validar_plano(row, plano):
    """
    Regra 2 — Valida cruzada com plan_colunas_cpis:
      OF, OV, Modelo (no campo designacao), Espessura, Largura (lbase), Comprimento.

    Devolve: (status_str, lista_de_notas, entrada_plano_ou_None)
    """
    of_raw = str(row.get("of", "") or "").strip()
    try:
        of_int = int(float(of_raw))
    except (ValueError, TypeError):
        return "OF INVÁLIDO", [f"OF '{of_raw}': valor não numérico"], None

    entradas = plano.get(of_int, [])
    if not entradas:
        return "OF SEM MATCH", [f"OF {of_int}: não encontrado no plano"], None

    modelo_ocr = str(row.get("modelo", "") or "").strip().upper()
    comp_ocr   = _num(row.get("comprimento_mm"))
    entrada    = _melhor_entrada(entradas, modelo_ocr, comp_ocr)

    notas  = []
    status = "OK"

    # — OV: compara OCR com plano; regista diferença como nota (não bloqueia)
    ov_ocr  = str(row.get("ov", "") or "").strip()
    ov_plan = str(entrada.get("ov") or "").strip()
    if ov_ocr and ov_plan and ov_ocr not in ("?",):
        if ov_ocr != ov_plan:
            # Não é divergência grave — OCR pode ter lido mal; o plano é a fonte
            if status == "OK":
                status = "AVISO"
            notas.append(f"OV ajustado: OCR={ov_ocr} → Plano={ov_plan}")

    # — Modelo: compara OCR com designação do plano; regista diferença como nota
    if modelo_ocr and modelo_ocr not in ("?",):
        if modelo_ocr not in entrada["designacao"].upper():
            if status == "OK":
                status = "AVISO"
            notas.append(
                f"Modelo ajustado: OCR={modelo_ocr} → Plano={entrada['designacao'][:40]}"
            )

    # — Espessura (exacta)
    esp_ocr  = _num(row.get("espessura_mm"))
    esp_plan = _num(entrada.get("esp"))
    if esp_ocr is not None and esp_plan is not None:
        if abs(esp_ocr - esp_plan) > 0.01:
            status = "DIVERGÊNCIA"
            notas.append(f"ESP: OCR={esp_ocr} Plano={esp_plan}")

    # — LBASE: OCR.lbase (largura da base da coluna, do kanban) vs plan.lbase
    # NOTA (patch 2026-05-03): comparava antes row["largura_mm"] (largura
    # da bobine, ~950mm) contra plan.lbase (largura da base da coluna, ~600mm).
    # São dimensões físicas DIFERENTES — a OCR já captura LBASE separada-
    # mente no campo `coni_lbase`. Ver reports/sap_validation/_summary.md.
    lbase_ocr = _num(row.get("coni_lbase"))
    lbase_pl  = _num(entrada.get("lbase"))
    if lbase_ocr is not None and lbase_pl is not None:
        diff = abs(lbase_ocr - lbase_pl)
        if diff > TOL_LARG_MM:
            if status == "OK":
                status = "AVISO"
            notas.append(
                f"LBASE: OCR={lbase_ocr:.0f} Plano.lbase={lbase_pl:.0f} "
                f"(dif={diff:.0f}mm > ±{TOL_LARG_MM}mm)"
            )

    # — Comprimento: ±TOL_COMP_MM
    comp_plan = _num(entrada.get("comp"))
    if comp_ocr is not None and comp_plan is not None:
        diff = abs(comp_ocr - comp_plan)
        if diff > TOL_COMP_MM:
            status = "DIVERGÊNCIA"
            notas.append(
                f"COMP: OCR={comp_ocr:.0f} Plano={comp_plan:.0f} "
                f"(dif={diff:.0f}mm > ±{TOL_COMP_MM}mm)"
            )

    return status, notas, entrada

# ═══════════════════════════════════════════════════════════════════════════
# ESCREVER LINHA NO EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def prox_linha(ws, inicio=3):
    """Devolve o índice da próxima linha vazia a partir de `inicio`."""
    last = inicio - 1
    for row in ws.iter_rows(min_row=inicio):
        if any(c.value is not None for c in row):
            last = row[0].row
    return last + 1


def _val(v, como_int=False):
    """
    Prepara um valor para célula Excel:
      - strings vazias, '?' e None → None
      - se como_int=True, tenta converter para inteiro
    """
    if v is None or str(v).strip() in ("", "?"):
        return None
    if como_int:
        n = _num(v)
        if n is not None:
            return int(n) if n == int(n) else n
    return v


def escrever_linha(ws, row_dict, row_idx, validacao, notas_extra, is_alt,
                   entrada_plano=None):
    """
    Escreve uma linha do CSV na folha Registos do Excel (34 colunas A–AH).
    Aplica cor de fundo conforme o resultado da validação.

    Se `entrada_plano` estiver disponível, os campos OF, OV e Modelo
    exportados para o Excel são os do PLANO (fonte mais fiável), não os
    valores OCR brutos. As diferenças já ficam registadas em notas_extra.
    """
    hoje = datetime.date.today().strftime("%d/%m/%Y")

    # Cor de fundo conforme validação
    if validacao in ("DIVERGÊNCIA", "LOTE NÃO ENCONTRADO", "OF INVÁLIDO"):
        bg = COR_DIVERGENCIA
    elif validacao in ("SEM MATCH", "SEM LOTE", "OF SEM MATCH", "AVISO"):
        bg = COR_AVISO
    elif validacao == "OK":
        bg = COR_OK if not is_alt else COR_OK  # sempre verde quando OK
    else:
        bg = COR_ALT if is_alt else COR_BRANCO

    # Campo Notas: combina notas OCR (do CSV) com notas de validação
    notas_ocr = str(row_dict.get("ocr_notas", "") or "").strip()
    notas_val = " | ".join(notas_extra) if notas_extra else ""
    notas_full = " | ".join(filter(None, [notas_ocr, notas_val])) or None

    # ── Valores do plano prevalecem sobre OCR para OF, OV e Modelo ──────────
    # Razão: o plano é a fonte de verdade; o OCR pode ter lidos parciais.
    # As diferenças já foram registadas em notas_extra por validar_plano().
    if entrada_plano:
        of_excel      = _val(str(entrada_plano.get("of_int") or
                                 row_dict.get("of", "")), como_int=True)
        ov_excel      = _val(str(entrada_plano.get("ov") or
                                 row_dict.get("ov", "")), como_int=True)
        modelo_excel  = _val(entrada_plano.get("designacao") or
                             row_dict.get("modelo", ""))
        cliente_excel = _val(entrada_plano.get("cliente") or
                             row_dict.get("cliente", ""))
    else:
        of_excel      = _val(row_dict.get("of"), como_int=True)
        ov_excel      = _val(row_dict.get("ov"), como_int=True)
        modelo_excel  = _val(row_dict.get("modelo"))
        cliente_excel = _val(row_dict.get("cliente"))

    linha = [
        _val(row_dict.get("data")),                          # A  Data
        _val(row_dict.get("cod_funcionario")),               # B  Cód. Funcionário
        _val(row_dict.get("nome_funcionario")),              # C  Nome Funcionário
        _val(row_dict.get("turno")),                         # D  Turno
        _val(row_dict.get("setor_maquina_desc")),            # E  Setor / Máquina Desc.
        _val(row_dict.get("cod_maquina")),                   # F  Cód. Máquina
        _val(row_dict.get("num_kanban"), como_int=True),     # G  N.º Kanban
        of_excel,                                            # H  OF  ← do plano
        ov_excel,                                            # I  OV  ← do plano
        cliente_excel,                                       # J  Cliente ← do plano
        modelo_excel,                                        # K  Modelo ← do plano
        _val(row_dict.get("qtd"), como_int=True),            # L  QTD
        None,                                                # M  QTD Metros (manual)
        _val(row_dict.get("comprimento_mm"), como_int=True), # N  Comprimento (mm)
        _val(row_dict.get("largura_mm"), como_int=True),     # O  Largura (mm)
        _val(row_dict.get("espessura_mm")),                  # P  Espessura (mm)
        _val(row_dict.get("lote")),                          # Q  LOTE
        _val(row_dict.get("coni_lbase")),                    # R  CONI
        _val(row_dict.get("nesting")),                       # S  Nesting (alfanum: C6, G10, etc.)
        None,                                                # T  Cesta (manual)
        _val(row_dict.get("duracao")),                       # U  Duração
        None,                                                # V  Peso (kg)       ← PENDENTE
        None,                                                # W  Desperdício (kg)← PENDENTE
        None,                                                # X  % Desperdício   ← PENDENTE
        None,                                                # Y  Sucata
        None,                                                # Z  Fim Bobine
        None,                                                # AA Importar p/Consumo
        None,                                                # AB Enviado
        None,                                                # AC Kanban Exportado
        None,                                                # AD Stock
        _val(row_dict.get("horas_trabalhadas")),             # AE Total Horas
        validacao,                                           # AF Validação
        notas_full,                                          # AG Notas
        hoje,                                               # AH Data Registo
    ]

    for ci, valor in enumerate(linha, start=1):
        cell = ws.cell(row=row_idx, column=ci, value=valor)
        bold = (ci == 3)      # Nome Funcionário a bold
        wrap = (ci == 33)     # Notas com wrap_text
        _estilo(cell, bg, bold=bold, wrap=wrap)

    ws.row_dimensions[row_idx].height = 18

# ═══════════════════════════════════════════════════════════════════════════
# RESOLUÇÃO DE DITTOS ( " = igual à célula acima )
# ═══════════════════════════════════════════════════════════════════════════

#: Caracteres/sequências reconhecidos como "ditto" (igual à linha de cima).
_DITTO_VALORES = {'"', '""', "''", "idem", "id.", "id"}


def _e_ditto(valor):
    """Devolve True se o valor representar um ditto (repetição da célula de cima)."""
    if valor is None:
        return False
    return str(valor).strip().lower() in _DITTO_VALORES


def resolver_ditto(linhas):
    """
    Percorre a lista de dicionários lida do CSV e substitui qualquer valor
    reconhecido como ditto pelo valor válido mais próximo na coluna acima.

    Regras:
      • Se uma célula contém '"' (ou equivalente), herda o valor da linha
        imediatamente anterior que não seja também ditto.
      • Se houver várias linhas consecutivas com '"', todas herdam do mesmo
        valor superior válido (procura retroactiva até encontrar valor real).
      • Colunas sem valor acima válido ficam com string vazia (sem alteração).

    Altera as linhas *in-place* e devolve a mesma lista (para encadeamento).
    """
    if not linhas:
        return linhas

    # Tabela auxiliar: { campo: último valor válido visto }
    ultimo_valido = {campo: "" for campo in linhas[0].keys()}

    for i, linha in enumerate(linhas):
        for campo, valor in linha.items():
            if _e_ditto(valor):
                # Substituir pelo último valor válido registado para esta coluna
                linhas[i][campo] = ultimo_valido.get(campo, "")
                log.debug(
                    '  Ditto resolvido linha %d campo "%s": "%s" → "%s"',
                    i + 1, campo, valor, linhas[i][campo],
                )
            else:
                # Valor real — actualiza o registo para futuras linhas ditto
                if valor is not None and str(valor).strip() not in ("", "?"):
                    ultimo_valido[campo] = valor

    return linhas


# ═══════════════════════════════════════════════════════════════════════════
# DETECÇÃO AUTOMÁTICA DE FORMATO
# ═══════════════════════════════════════════════════════════════════════════

def _detectar_formato(caminho):
    """
    Detecta o formato do CSV lendo as primeiras linhas do ficheiro:
      • 'novo_layout' — secções delimitadas por === ... ===
      • 'flat'        — secções delimitadas por ### ...
      • None          — formato não reconhecido (ficheiro ignorado)
    """
    try:
        with open(caminho, encoding="utf-8-sig", newline="") as f:
            for linha in f:
                txt = linha.strip()
                if not txt:
                    continue
                if txt.startswith("===") and txt.endswith("==="):
                    return "novo_layout"
                if txt.startswith("###"):
                    return "flat"
    except Exception as exc:
        log.warning("Não foi possível detectar formato de %s: %s",
                    os.path.basename(caminho), exc)
    return None


# ═══════════════════════════════════════════════════════════════════════════
# PARSER — FORMATO A  (=== marcadores, *_kanban.csv)
# ═══════════════════════════════════════════════════════════════════════════

# Mapeamento: coluna do novo layout → campo interno padrão
_MAP_TABELA = {
    "PRI":            "nesting",
    "CLIENTE":        "cliente",
    "OV":             "ov",
    "OF":             "of",
    "MODELO":         "modelo",
    "QTD COLUNAS":    "qtd",
    "COMP (MM)":      "comprimento_mm",
    "LARG (MM)":      "largura_mm",
    "LOTE":           "lote",
    "CONI":           "_coni",       # coluna extra — vai para ocr_notas
    "ESP":            "espessura_mm",
    "LBASE":          "coni_lbase",
    "LTOPO":          "coni_ltopo",
}

_CODIGOS_MAQUINAS = {
    "BOBINE-FORMATO":  "M032",
    "BOB.FORMATO":     "M032",
    "BOBINE FORMATO":  "M032",
    "GUILHOTINA 9M":   "M020",
    "GUILH. 9M":       "M020",
    "GUILHOTINA 10M":  "M091",
    "GUILH. 10M":      "M091",
    "ACABAMENTO MTG2": "M001",
    "ACAB. MTG2":      "M001",
    "MTG2":            "M001",
}


def _derivar_cod_maquina(setor_desc):
    """Converte descrição da máquina em código (tabela interna)."""
    s = str(setor_desc or "").strip().upper()
    # Correspondência exacta primeiro
    if s in _CODIGOS_MAQUINAS:
        return _CODIGOS_MAQUINAS[s]
    # Correspondência parcial
    for chave, cod in _CODIGOS_MAQUINAS.items():
        if chave in s:
            return cod
    return ""


def _limpar_setor(valor_bruto):
    """
    Remove o prefixo 'Setor/Máquina ' que o novo CSV inclui no campo.
    Ex.: 'Setor/Máquina BOBINE-FORMATO' → 'BOBINE-FORMATO'
    """
    v = str(valor_bruto or "").strip()
    for prefixo in ("Setor/Máquina ", "Setor/Maquina ", "SETOR/MÁQUINA ",
                    "SETOR/MAQUINA "):
        if v.upper().startswith(prefixo.upper()):
            v = v[len(prefixo):].strip()
            break
    return v.upper()


def ler_csv_novo_layout(caminho):
    """
    Lê um CSV no formato A (=== secções) e devolve uma lista de dicionários
    com campos no formato interno padrão.

    Estrutura esperada do ficheiro:
      1ª linha  : FICHEIRO;INICIO;FIM;DURACAO_SEG  (metadados — usado para ocr_notas)
      secção    : === CABEÇALHO ===
      secção    : === TABELA DE PRODUÇÃO ===
      secção    : === RODAPÉ ===
    """
    data_reg = datetime.date.today().strftime("%d/%m/%Y")

    with open(caminho, encoding="utf-8-sig", newline="") as f:
        leitor = csv.reader(f, delimiter=";")
        linhas_raw = [r for r in leitor]

    # ── 1. Metadados da 1ª linha (FICHEIRO;INICIO;FIM;DURACAO_SEG) ────────────
    meta_timing = ""
    if linhas_raw and len(linhas_raw[0]) >= 4:
        hdrs_meta = [c.strip().upper() for c in linhas_raw[0]]
        if "DURACAO_SEG" in hdrs_meta:
            try:
                dados_meta = linhas_raw[1]
                idx_dur = hdrs_meta.index("DURACAO_SEG")
                dur = dados_meta[idx_dur].strip() if idx_dur < len(dados_meta) else ""
                if dur:
                    meta_timing = f"OCR duração: {dur}s"
            except (IndexError, ValueError):
                pass

    # ── 2. Identificar secções pelo marcador === ... === ──────────────────────
    seccoes = {}   # { "CABEÇALHO": [linhas], "TABELA DE PRODUÇÃO": [linhas], ... }
    seccao_atual = None
    for linha in linhas_raw:
        if not any(c.strip() for c in linha):   # linha vazia
            continue
        txt = ";".join(linha).strip()
        if txt.startswith("===") and txt.endswith("==="):
            nome = txt.strip("= ").strip().upper()
            seccao_atual = nome
            seccoes[seccao_atual] = []
        elif seccao_atual is not None:
            seccoes[seccao_atual].append(linha)

    # ── 3. Cabeçalho ──────────────────────────────────────────────────────────
    cab = {}
    sec_cab = seccoes.get("CABEÇALHO", seccoes.get("CABECALHO", []))
    if len(sec_cab) >= 2:
        hdrs_cab = [c.strip().upper() for c in sec_cab[0]]
        vals_cab = sec_cab[1]
        for i, h in enumerate(hdrs_cab):
            cab[h] = vals_cab[i].strip() if i < len(vals_cab) else ""

    setor_bruto  = cab.get("SETOR/MÁQUINA", cab.get("SETOR/MAQUINA", ""))
    setor_limpo  = _limpar_setor(setor_bruto)
    cod_maquina  = _derivar_cod_maquina(setor_limpo)
    num_kanban   = cab.get("NÚMERO", cab.get("NUMERO", "")).lstrip("0") or \
                   cab.get("NÚMERO", cab.get("NUMERO", ""))
    if num_kanban.startswith("0") and len(num_kanban) > 1:
        num_kanban_clean = num_kanban.lstrip("0") or "0"
    else:
        num_kanban_clean = num_kanban

    dados_cab = {
        "nome_funcionario": cab.get("OPERADOR", "").upper(),
        "cod_funcionario":  num_kanban_clean,
        "num_kanban":       num_kanban_clean,
        "setor_maquina_desc": setor_limpo,
        "cod_maquina":      cod_maquina,
        "data":             cab.get("DATA", "").replace(".", "/").replace("-", "/"),
        "turno":            cab.get("TURNO", ""),
    }

    # ── 4. Rodapé ─────────────────────────────────────────────────────────────
    rod = {}
    sec_rod = seccoes.get("RODAPÉ", seccoes.get("RODAPE", []))
    if len(sec_rod) >= 2:
        hdrs_rod = [c.strip().upper() for c in sec_rod[0]]
        vals_rod = sec_rod[1]
        for i, h in enumerate(hdrs_rod):
            rod[h] = vals_rod[i].strip() if i < len(vals_rod) else ""

    colunas_prod  = rod.get("COLUNAS PRODUZIDAS", "")
    horas_trab    = rod.get("HORAS TRABALHADAS", "")

    # ── 5. Tabela de produção ─────────────────────────────────────────────────
    sec_tab = seccoes.get("TABELA DE PRODUÇÃO",
              seccoes.get("TABELA DE PRODUCAO",
              seccoes.get("TABELA",  [])))
    if not sec_tab:
        log.warning("  Secção 'TABELA DE PRODUÇÃO' não encontrada em %s",
                    os.path.basename(caminho))
        return []

    hdrs_tab = [c.strip().upper() for c in sec_tab[0]]
    linhas_prod = []

    for linha_raw in sec_tab[1:]:
        if not any(c.strip() for c in linha_raw):
            continue   # linha vazia dentro da tabela

        row_orig = {}
        for i, h in enumerate(hdrs_tab):
            row_orig[h] = linha_raw[i].strip() if i < len(linha_raw) else ""

        row = {}
        coni_val = ""
        for col_novo, campo in _MAP_TABELA.items():
            valor = row_orig.get(col_novo, "")
            if campo == "_coni":
                coni_val = valor
            else:
                row[campo] = valor

        notas_partes = []
        if coni_val:
            notas_partes.append(f"CONI={coni_val}")
        if meta_timing:
            notas_partes.append(meta_timing)
        row["ocr_notas"] = " | ".join(notas_partes)

        row.update(dados_cab)
        row["colunas_produzidas"] = colunas_prod
        row["horas_trabalhadas"]  = horas_trab
        row["observacoes"]        = ""
        row["duracao"]            = ""
        row["qtd_metros"]         = ""
        row["cesta"]              = ""
        row["data_registo"]       = data_reg

        linhas_prod.append(row)

    log.info(
        "  Formato A (===): %d linha(s) | operador=%s | data=%s | kanban=%s",
        len(linhas_prod),
        dados_cab.get("nome_funcionario", "?"),
        dados_cab.get("data", "?"),
        dados_cab.get("num_kanban", "?"),
    )
    return linhas_prod


# ═══════════════════════════════════════════════════════════════════════════
# PARSER — FORMATO B  (### marcadores, Operador_AAAA.MM.DD.csv)
# ═══════════════════════════════════════════════════════════════════════════

# Mapeamento: coluna do formato flat → campo interno padrão
_MAP_TABELA_FLAT = {
    "PRI":      "nesting",
    "CLIENTE":  "cliente",
    "OV":       "ov",
    "OF":       "of",
    "MODELO":   "modelo",
    "QTD":      "qtd",
    "COMP_MM":  "comprimento_mm",
    "LARG_MM":  "largura_mm",
    "LOTE":     "lote",
    "CONI":     "_coni",       # coluna extra — vai para ocr_notas
    "ESP":      "espessura_mm",
    "LBASE":    "coni_lbase",
    "LTOPO":    "coni_ltopo",
}


def ler_csv_formato_flat(caminho):
    """
    Lê um CSV no formato B (### secções) e devolve uma lista de dicionários
    com campos no formato interno padrão.

    Estrutura esperada:
      ### CABEÇALHO
      FICHEIRO;DATA;OPERADOR;N_OPERADOR;SETOR_MAQUINA
      <valores — cada linha repete FICHEIRO, DATA, OPERADOR>

      ### TABELA DE PRODUÇÃO
      FICHEIRO;DATA;OPERADOR;PRI;CLIENTE;OV;OF;MODELO;QTD;COMP_MM;LARG_MM;LOTE;CONI;ESP;LBASE;LTOPO
      <linhas de produção>

      ### RODAPÉ
      FICHEIRO;DATA;OPERADOR;COLUNAS_PRODUZIDAS;HORAS_TRABALHADAS
      <valores>

    Diferenças face ao formato A:
      • Marcadores ### em vez de === ... ===
      • Nomes de colunas: N_OPERADOR, SETOR_MAQUINA, QTD, COMP_MM, LARG_MM,
        COLUNAS_PRODUZIDAS, HORAS_TRABALHADAS
      • Cada linha de dados inclui FICHEIRO, DATA e OPERADOR como primeiras colunas
      • ESP pode usar vírgula como separador decimal (ex.: 2,6) — normalizado aqui
    """
    data_reg = datetime.date.today().strftime("%d/%m/%Y")

    with open(caminho, encoding="utf-8-sig", newline="") as f:
        leitor = csv.reader(f, delimiter=";")
        linhas_raw = [r for r in leitor]

    # ── 1. Identificar secções pelo marcador ### ──────────────────────────────
    seccoes = {}
    seccao_atual = None
    for linha in linhas_raw:
        txt = ";".join(linha).strip()
        if txt.startswith("###"):
            nome = txt.lstrip("#").strip().upper()
            seccao_atual = nome
            seccoes[seccao_atual] = []
        elif seccao_atual is not None:
            if any(c.strip() for c in linha):   # ignora linhas vazias
                seccoes[seccao_atual].append(linha)

    # ── 2. Cabeçalho ─────────────────────────────────────────────────────────
    cab = {}
    sec_cab = seccoes.get("CABEÇALHO", seccoes.get("CABECALHO", []))
    if len(sec_cab) >= 2:
        hdrs_cab = [c.strip().upper() for c in sec_cab[0]]
        vals_cab = sec_cab[1]
        for i, h in enumerate(hdrs_cab):
            cab[h] = vals_cab[i].strip() if i < len(vals_cab) else ""

    setor_bruto = cab.get("SETOR_MAQUINA", cab.get("SETOR/MAQUINA", ""))
    setor_limpo = _limpar_setor(setor_bruto) if setor_bruto else setor_bruto.upper()
    cod_maquina = _derivar_cod_maquina(setor_limpo)

    num_kanban = cab.get("N_OPERADOR", "")
    if num_kanban.startswith("0") and len(num_kanban) > 1:
        num_kanban_clean = num_kanban.lstrip("0") or "0"
    else:
        num_kanban_clean = num_kanban

    # Normalizar separadores de data
    data_ocr = cab.get("DATA", "").replace(".", "/").replace("-", "/")

    dados_cab = {
        "nome_funcionario":   cab.get("OPERADOR", "").upper(),
        "cod_funcionario":    num_kanban_clean,
        "num_kanban":         num_kanban_clean,
        "setor_maquina_desc": setor_limpo,
        "cod_maquina":        cod_maquina,
        "data":               data_ocr,
        "turno":              "",
    }

    # ── 3. Rodapé ─────────────────────────────────────────────────────────────
    rod = {}
    sec_rod = seccoes.get("RODAPÉ", seccoes.get("RODAPE", []))
    if len(sec_rod) >= 2:
        hdrs_rod = [c.strip().upper() for c in sec_rod[0]]
        vals_rod = sec_rod[1]
        for i, h in enumerate(hdrs_rod):
            rod[h] = vals_rod[i].strip() if i < len(vals_rod) else ""

    colunas_prod = rod.get("COLUNAS_PRODUZIDAS", "")
    horas_trab   = rod.get("HORAS_TRABALHADAS", "")

    # ── 4. Tabela de produção ─────────────────────────────────────────────────
    sec_tab = seccoes.get("TABELA DE PRODUÇÃO",
              seccoes.get("TABELA DE PRODUCAO",
              seccoes.get("TABELA", [])))
    if not sec_tab:
        log.warning("  Secção 'TABELA DE PRODUÇÃO' não encontrada em %s",
                    os.path.basename(caminho))
        return []

    hdrs_tab = [c.strip().upper() for c in sec_tab[0]]
    linhas_prod = []

    for linha_raw in sec_tab[1:]:
        if not any(c.strip() for c in linha_raw):
            continue

        # Dicionário com nomes originais da coluna
        row_orig = {}
        for i, h in enumerate(hdrs_tab):
            row_orig[h] = linha_raw[i].strip() if i < len(linha_raw) else ""

        # Mapear para campos internos padrão
        row = {}
        coni_val = ""
        for col_flat, campo in _MAP_TABELA_FLAT.items():
            valor = row_orig.get(col_flat, "")
            if campo == "_coni":
                coni_val = valor
            else:
                row[campo] = valor

        # Normalizar separador decimal: vírgula → ponto nos campos numéricos
        for campo_num in ("espessura_mm", "comprimento_mm", "largura_mm"):
            if campo_num in row:
                row[campo_num] = row[campo_num].replace(",", ".")

        # Construir ocr_notas com CONI
        notas_partes = []
        if coni_val:
            notas_partes.append(f"CONI={coni_val}")
        row["ocr_notas"] = " | ".join(notas_partes)

        # Campos do cabeçalho e rodapé (repetem em todas as linhas)
        row.update(dados_cab)
        row["colunas_produzidas"] = colunas_prod
        row["horas_trabalhadas"]  = horas_trab
        row["observacoes"]        = ""
        row["duracao"]            = ""
        row["qtd_metros"]         = ""
        row["cesta"]              = ""
        row["data_registo"]       = data_reg

        linhas_prod.append(row)

    log.info(
        "  Formato B (###): %d linha(s) | operador=%s | data=%s | kanban=%s",
        len(linhas_prod),
        dados_cab.get("nome_funcionario", "?"),
        dados_cab.get("data", "?"),
        dados_cab.get("num_kanban", "?"),
    )
    return linhas_prod


# ═══════════════════════════════════════════════════════════════════════════
# PROCESSAMENTO DE UM CSV
# ═══════════════════════════════════════════════════════════════════════════

def processar_csv(caminho, stock, plano, ws):
    """
    Lê um CSV (formato A ou B) e escreve as linhas no Excel.
    Detecta automaticamente o formato do ficheiro.
    Devolve (n_ok, n_avisos, n_divergencias).
    """
    log.info("A processar: %s", os.path.basename(caminho))

    # ── Detecção automática de formato ────────────────────────────────────
    formato = _detectar_formato(caminho)

    if formato == "novo_layout":
        linhas = ler_csv_novo_layout(caminho)
    elif formato == "flat":
        linhas = ler_csv_formato_flat(caminho)
    else:
        log.warning("  Formato não reconhecido, ficheiro ignorado: %s",
                    os.path.basename(caminho))
        return 0, 0, 0

    if not linhas:
        log.warning("  CSV sem linhas de produção, ignorado: %s",
                    os.path.basename(caminho))
        return 0, 0, 0

    # Resolver dittos: substituir '"' pelo valor válido da linha acima
    resolver_ditto(linhas)

    n_ok = n_avisos = n_div = 0
    row_idx = prox_linha(ws)
    is_alt  = (row_idx % 2 == 0)

    for linha in linhas:
        notas_all = []

        # ── Validação 1: StockSAP ─────────────────────────────────────────
        st_sap, notas_sap = validar_stocksap(linha, stock)
        notas_all.extend(notas_sap)

        # ── Validação 2: Plano ────────────────────────────────────────────
        st_plan, notas_plan, entrada_plano = validar_plano(linha, plano)
        notas_all.extend(notas_plan)

        # ── Enriquecer entrada_plano com of_int para uso em escrever_linha ──
        if entrada_plano is not None:
            of_raw = str(linha.get("of", "") or "").strip()
            try:
                entrada_plano["of_int"] = int(float(of_raw))
            except (ValueError, TypeError):
                entrada_plano["of_int"] = None

        # ── Status consolidado ────────────────────────────────────────────
        todos = (st_sap, st_plan)
        if "DIVERGÊNCIA" in todos or "LOTE NÃO ENCONTRADO" in todos or "OF INVÁLIDO" in todos:
            status_final = "DIVERGÊNCIA"
            n_div += 1
        elif any(s in todos for s in ("SEM MATCH", "OF SEM MATCH", "SEM LOTE")):
            status_final = "SEM MATCH"
            n_avisos += 1
        elif "AVISO" in todos:
            status_final = "AVISO"
            n_avisos += 1
        else:
            status_final = "OK"
            n_ok += 1

        escrever_linha(ws, linha, row_idx, status_final, notas_all, is_alt,
                       entrada_plano=entrada_plano)
        row_idx += 1
        is_alt   = not is_alt

    log.info(
        "  %d linha(s): %d OK | %d avisos | %d divergências",
        len(linhas), n_ok, n_avisos, n_div
    )
    return n_ok, n_avisos, n_div

# ═══════════════════════════════════════════════════════════════════════════
# GESTÃO DE FICHEIROS
# ═══════════════════════════════════════════════════════════════════════════

def encontrar_csvs():
    """
    Lista todos os CSVs pendentes em 02_Dados_Extraidos/csv/.

    Suporta dois padrões de nomes:
      • Formato A (=== marcadores): *_kanban.csv
        Ex.: JulioLima_2026.04.02_kanban.csv
      • Formato B (### marcadores): Operador_AAAA.MM.DD.csv
        Ex.: AugustoMonteiro_2026.04.09.csv

    Qualquer ficheiro .csv presente na pasta (excluindo subpastas) é
    considerado. O formato é detectado automaticamente em processar_csv().
    Ficheiros com formato não reconhecido são registados no log e ignorados.
    """
    csvs = []
    if not os.path.isdir(DIR_CSV):
        log.warning("Pasta CSV não encontrada: %s", DIR_CSV)
        return csvs
    for nome in sorted(os.listdir(DIR_CSV)):
        caminho = os.path.join(DIR_CSV, nome)
        if os.path.isfile(caminho) and nome.lower().endswith(".csv"):
            csvs.append(caminho)
    return csvs


def marcar_importado(caminho):
    """Move o CSV para a subpasta imported/ após processamento com sucesso."""
    os.makedirs(DIR_IMPORT, exist_ok=True)
    nome = os.path.basename(caminho)
    dest = os.path.join(DIR_IMPORT, nome)
    if os.path.exists(dest):
        base, ext = os.path.splitext(nome)
        ts   = int(datetime.datetime.now().timestamp())
        dest = os.path.join(DIR_IMPORT, f"{base}_{ts}{ext}")
    shutil.move(caminho, dest)
    log.info("  CSV arquivado: imported/%s", os.path.basename(dest))


def guardar_xlsx(wb, dest):
    """
    Guarda o Excel com backup automático.
    Estratégia robusta:
      1. Guarda numa cópia temporária primeiro.
      2. Se OK, substitui o ficheiro principal.
      3. O .bak é mantido como histórico (não apagado, para não depender
         de permissões de eliminação no sistema do utilizador).
    """
    import tempfile

    dest_dir  = os.path.dirname(dest)
    dest_base = os.path.basename(dest)

    # Guardar numa cópia temporária no mesmo directório
    try:
        fd, tmp_path = tempfile.mkstemp(suffix=".tmp", dir=dest_dir)
        os.close(fd)
        wb.save(tmp_path)
    except Exception as exc:
        log.error("  Erro ao escrever ficheiro temporário: %s", exc)
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise

    # Criar backup do ficheiro actual (ignorar erros — pode já existir ou ser read-only)
    bak = dest + ".bak"
    if os.path.exists(dest):
        try:
            shutil.copy2(dest, bak)
            log.info("  Backup actualizado: %s", os.path.basename(bak))
        except Exception as bak_exc:
            log.warning("  Não foi possível criar backup: %s (prosseguindo)", bak_exc)

    # Substituir ficheiro principal pela versão temporária
    try:
        shutil.move(tmp_path, dest)
        log.info("  Excel guardado: %s", dest_base)
    except Exception as exc:
        log.error("  Erro ao substituir Excel: %s", exc)
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise

# ═══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

def main():
    log.info("=" * 60)
    log.info("kanban_csv2excel_novo_layout.py — %s",
             datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    log.info("Raiz do projecto: %s", RAIZ)
    log.info("=" * 60)

    # 1. Carregar ficheiros de referência
    stock = carregar_stocksap()
    plano = carregar_plancolunas()

    # 2. Encontrar CSVs não importados (ambos os formatos)
    csvs = encontrar_csvs()
    if not csvs:
        log.info("Nenhum CSV pendente em: %s", DIR_DADOS)
        return

    log.info("CSVs a processar (%d):", len(csvs))
    for c in csvs:
        log.info("  %s", os.path.basename(c))

    # 3. Abrir Excel de destino
    if not os.path.exists(EXCEL_DEST):
        log.error("Excel de destino não encontrado: %s", EXCEL_DEST)
        log.error("Crie primeiro o ficheiro Excel ou verifique o caminho.")
        sys.exit(1)

    log.info("A abrir Excel: %s", os.path.basename(EXCEL_DEST))
    wb = openpyxl.load_workbook(EXCEL_DEST)

    if "Registos" not in wb.sheetnames:
        log.error("Folha 'Registos' não encontrada no Excel.")
        sys.exit(1)

    ws = wb["Registos"]

    # 4. Processar cada CSV
    total_ok = total_avisos = total_div = 0

    for caminho in csvs:
        try:
            n_ok, n_av, n_div = processar_csv(caminho, stock, plano, ws)
            total_ok     += n_ok
            total_avisos += n_av
            total_div    += n_div

            guardar_xlsx(wb, EXCEL_DEST)
            marcar_importado(caminho)

        except Exception as exc:
            log.error("ERRO ao processar %s: %s", os.path.basename(caminho), exc)
            log.error("O CSV NÃO foi movido — corrigir e executar novamente.")

    # 5. Sumário final
    log.info("─" * 60)
    log.info(
        "CONCLUÍDO — %d OK  |  %d avisos  |  %d divergências",
        total_ok, total_avisos, total_div
    )
    if total_div > 0:
        log.warning(
            "  %d linha(s) com DIVERGÊNCIA — verificar coluna 'Validação' no Excel.",
            total_div
        )
    if total_avisos > 0:
        log.info(
            "  %d linha(s) com AVISO/SEM MATCH — sem dados no plano ou lote.",
            total_avisos
        )
    log.info("Excel: %s", EXCEL_DEST)


if __name__ == "__main__":
    main()
