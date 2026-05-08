"""End-to-end factory validation: R27 outputs → CSVs → patched validator → OK count.

Sets up a clone of the factory directory structure under /tmp/_factory_test/,
generates CSVs from our R27 *_fixed.json outputs in the ### marker format
the validator expects, then runs the patched validator (LBASE fix from
Round 27) and reads back the Excel Validação column.

Compare actual OK count with our predicted 64.
"""
from __future__ import annotations

import json
import shutil
import subprocess
import sys
from collections import Counter
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "reports" / "extractions_v_qwen35_round27"
TEMP_FACTORY = Path(r"C:\Users\User\AppData\Local\Temp\_factory_test")
SOURCE_REFS = Path(r"C:\Users\User\AppData\Local\Temp\ai_check\AI")
PATCHED_VALIDATOR = REPO / "inputs" / "_factory_test" / "kanban_csv2excel_novo_layout_PATCHED.py"

ROW_FIELDS = ["pri", "cliente", "ov", "of", "modelo", "qtd",
              "comp_mm", "larg_mm", "lote", "coni", "esp", "lbase", "ltopo"]


def setup_dir_structure():
    """Create clone of factory directory structure."""
    if TEMP_FACTORY.exists():
        shutil.rmtree(TEMP_FACTORY)
    (TEMP_FACTORY / "02_Dados_Extraidos" / "csv").mkdir(parents=True)
    (TEMP_FACTORY / "04_Documentacao").mkdir(parents=True)
    (TEMP_FACTORY / "06_Kanban_OCR").mkdir(parents=True)
    print(f"Created: {TEMP_FACTORY}")


def copy_refs():
    """Copy StockSAP, plan_colunas, and patched validator."""
    shutil.copy2(SOURCE_REFS / "nifruka/04_Documentacao/StockSAP.xlsx",
                 TEMP_FACTORY / "04_Documentacao/StockSAP.xlsx")
    shutil.copy2(SOURCE_REFS / "nifruka/04_Documentacao/plan_colunas_cpis.xlsx",
                 TEMP_FACTORY / "04_Documentacao/plan_colunas_cpis.xlsx")
    shutil.copy2(PATCHED_VALIDATOR,
                 TEMP_FACTORY / "06_Kanban_OCR/kanban_csv2excel_novo_layout.py")
    print("Copied refs and patched validator")


def create_empty_kanbans_excel():
    """Clone Kanbans_Producao_NOVO.xlsx structure, clear data rows."""
    src = SOURCE_REFS / "nifruka/02_Dados_Extraidos/Kanbans_Producao_NOVO.xlsx"
    dst = TEMP_FACTORY / "02_Dados_Extraidos/Kanbans_Producao_NOVO.xlsx"
    wb = openpyxl.load_workbook(src)
    ws = wb["Registos"]
    # Delete data rows (keep title+header rows 1-2)
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    wb.save(dst)
    print(f"Created empty Kanbans_Producao_NOVO at {dst}")


def fmt_date(d: str) -> str:
    """Convert DD-MM-YYYY → DD-MM-YYYY (keep), or DD/MM/YYYY → DD-MM-YYYY."""
    if not d:
        return ""
    return d.replace("/", "-").replace(".", "-")


def json_to_csv(json_path: Path, csv_path: Path):
    """Convert R27 *_fixed.json to factory ### marker format."""
    data = json.loads(json_path.read_text(encoding="utf-8"))
    h = data.get("header", {})
    rows = data.get("rows", [])
    f = data.get("footer", {})

    sheet_name = json_path.stem.replace("_fixed", "")
    image_name = f"{sheet_name}.jpeg"
    data_str = fmt_date(h.get("data", ""))
    operador = h.get("operador", "")
    n_operador = h.get("n_operador", "")
    setor = h.get("setor_maquina", "")

    lines = []
    lines.append("### CABEÇALHO")
    lines.append("FICHEIRO;DATA;OPERADOR;N_OPERADOR;SETOR_MAQUINA")
    lines.append(f"{image_name};{data_str};{operador};{n_operador};{setor}")
    lines.append("")
    lines.append("### TABELA DE PRODUÇÃO")
    lines.append("FICHEIRO;DATA;OPERADOR;PRI;CLIENTE;OV;OF;MODELO;QTD;COMP_MM;LARG_MM;LOTE;CONI;ESP;LBASE;LTOPO")
    for row in rows:
        cells = [
            image_name, data_str, operador,
            row.get("pri", ""), row.get("cliente", ""), row.get("ov", ""),
            row.get("of", ""), row.get("modelo", ""), row.get("qtd", ""),
            row.get("comp_mm", ""), row.get("larg_mm", ""), row.get("lote", ""),
            row.get("coni", ""), row.get("esp", ""), row.get("lbase", ""), row.get("ltopo", ""),
        ]
        lines.append(";".join(str(c) for c in cells))
    lines.append("")
    lines.append("### RODAPÉ")
    lines.append("FICHEIRO;DATA;OPERADOR;COLUNAS_PRODUZIDAS;HORAS_TRABALHADAS")
    lines.append(f"{image_name};{data_str};{operador};{f.get('colunas_produzidas','')};{f.get('horas_trabalhadas','')}")

    # Factory CSVs use UTF-8 BOM
    csv_path.write_text("\n".join(lines), encoding="utf-8-sig")


def generate_csvs():
    """Generate one CSV per R27 sheet."""
    csv_dir = TEMP_FACTORY / "02_Dados_Extraidos/csv"
    fixeds = sorted(SRC.glob("*_fixed.json"))
    for fp in fixeds:
        name = fp.stem.replace("_fixed", "") + ".csv"
        json_to_csv(fp, csv_dir / name)
    print(f"Generated {len(fixeds)} CSVs in {csv_dir}")


def run_validator():
    """Execute the patched validator."""
    script = TEMP_FACTORY / "06_Kanban_OCR/kanban_csv2excel_novo_layout.py"
    print(f"\nRunning validator: {script}")
    print("=" * 70)
    result = subprocess.run(
        [sys.executable, str(script)],
        cwd=str(script.parent),
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    print(result.stdout)
    if result.stderr:
        print("STDERR:", result.stderr)
    if result.returncode != 0:
        print(f"Validator failed with exit code {result.returncode}")
        sys.exit(1)


def read_results():
    """Read Validação column from output Excel and tally."""
    excel = TEMP_FACTORY / "02_Dados_Extraidos/Kanbans_Producao_NOVO.xlsx"
    wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
    ws = wb["Registos"]
    # Header is row 2; Validação is column 32 (AF, 1-indexed)
    validacao_col = 32
    counts = Counter()
    rows_data = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        if row[0] is None:
            continue
        v = row[validacao_col - 1] if validacao_col - 1 < len(row) else None
        if v:
            counts[str(v)] += 1
            rows_data.append((row[0], row[7], row[10], v, row[32] if 32 < len(row) else ""))
    wb.close()
    return counts, rows_data


def main():
    print("=" * 70)
    print("End-to-end factory validation: R27 outputs → patched validator")
    print("=" * 70)

    setup_dir_structure()
    copy_refs()
    create_empty_kanbans_excel()
    generate_csvs()
    run_validator()

    print("\n" + "=" * 70)
    print("ACTUAL RESULTS (from Kanbans_Producao_NOVO.xlsx Validação column)")
    print("=" * 70)
    counts, rows_data = read_results()
    total = sum(counts.values())
    for status in ("OK", "AVISO", "SEM MATCH", "DIVERGÊNCIA"):
        n = counts.get(status, 0)
        pct = 100.0 * n / total if total else 0
        print(f"  {status:18s} {n:4d}  ({pct:5.1f} %)")
    print(f"  {'TOTAL ROWS':18s} {total:4d}")

    print("\n" + "=" * 70)
    print("COMPARISON vs PREDICTED (Round 27d)")
    print("=" * 70)
    predicted = {"OK": 64, "AVISO": 17, "SEM MATCH": 2, "DIVERGÊNCIA": 38}
    for status in ("OK", "AVISO", "SEM MATCH", "DIVERGÊNCIA"):
        a = counts.get(status, 0)
        p = predicted.get(status, 0)
        delta = a - p
        mark = "✓" if delta == 0 else f"Δ{delta:+d}"
        print(f"  {status:18s} actual={a:4d}  predicted={p:4d}  {mark}")


if __name__ == "__main__":
    main()
