"""R87 — Merge NEW plan_colunas_cpis (1).xlsx with CURRENT preserving
historical OFs.

Strategy:
- For OFs in BOTH files: prefer NEW (more recent data)
- For OFs only in CURRENT (4638 historical): keep CURRENT entries
- For OFs only in NEW (76 new): add them

Output: merged plan_colunas_cpis.xlsx with ~9922 OFs total.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

NEW = Path(r"C:\Users\User\Downloads\plan_colunas_cpis (1).xlsx")
CUR_BACKUP = Path(r"C:\kanban\nifruka\04_Documentacao\_plan_cur_tmp.xlsx")
OUT = Path(r"C:\kanban\nifruka\04_Documentacao\plan_colunas_cpis.xlsx")
MIRROR = Path(r"C:\Users\User\ocr\kanban_refs\04_Documentacao\plan_colunas_cpis.xlsx")


def load_rows(path: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def extract_of(row: tuple, hdrs: dict[str, int]) -> str | None:
    try:
        return str(int(float(str(row[hdrs["of"]]))))
    except (ValueError, TypeError, KeyError):
        return None


def main() -> int:
    new_rows = load_rows(NEW)
    cur_rows = load_rows(CUR_BACKUP)

    new_header = new_rows[0]
    cur_header = cur_rows[0]
    if new_header != cur_header:
        print(f"!! Header mismatch:")
        print(f"   NEW: {new_header}")
        print(f"   CUR: {cur_header}")
        return 1
    hdrs = {str(h).strip().lower(): i for i, h in enumerate(new_header) if h}

    # Group by OF
    new_by_of: dict[str, list[tuple]] = {}
    for row in new_rows[1:]:
        if row[0] is None:
            continue
        of = extract_of(row, hdrs)
        if not of:
            continue
        new_by_of.setdefault(of, []).append(row)

    cur_by_of: dict[str, list[tuple]] = {}
    for row in cur_rows[1:]:
        if row[0] is None:
            continue
        of = extract_of(row, hdrs)
        if not of:
            continue
        cur_by_of.setdefault(of, []).append(row)

    ofs_new = set(new_by_of.keys())
    ofs_cur = set(cur_by_of.keys())

    # 1. All NEW rows (preferred for overlapping OFs)
    merged_rows: list[tuple] = []
    for of, rows in new_by_of.items():
        merged_rows.extend(rows)
    # 2. CURRENT-only OFs (historicals not in NEW)
    historical_only = ofs_cur - ofs_new
    for of in historical_only:
        merged_rows.extend(cur_by_of[of])

    print(f"NEW OFs: {len(ofs_new)}")
    print(f"CUR OFs: {len(ofs_cur)}")
    print(f"Common OFs: {len(ofs_new & ofs_cur)}")
    print(f"NEW-only: {len(ofs_new - ofs_cur)}")
    print(f"CUR-only (preserved as historical): {len(historical_only)}")
    print(f"Merged total OFs: {len(ofs_new | ofs_cur)}")
    print(f"Merged total rows: {len(merged_rows)}")

    # Write output
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "plan_colunas_cpis"
    out_ws.append(list(new_header))
    for row in merged_rows:
        out_ws.append(list(row))
    out_wb.save(OUT)
    print(f"Wrote {OUT}  ({OUT.stat().st_size:,} bytes)")

    shutil.copy(OUT, MIRROR)
    print(f"Mirrored to {MIRROR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
