"""R87b — Merge NEW StockSAP_Dinamico with CURRENT (backup) preserving
historical lotes that the new "deduplicated" export omitted.

Strategy:
- NEW Dinamico (1469 unique lotes) is the authoritative deduplicated view
- CURRENT pre-R67 backup (~3000 lotes) has historical lotes M25B/M26B
  that the NEW view is missing
- Output: merge — NEW wins on overlap, CURRENT-only lotes are added
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

NEW = Path(r"C:\Users\User\Downloads\StockSAP_Dinamico.xlsx")
BACKUP = Path(r"C:\kanban\nifruka\_backup_pre_R67\StockSAP_pre_R67.xlsx")
OUT = Path(r"C:\kanban\nifruka\04_Documentacao\StockSAP.xlsx")
MIRROR = Path(r"C:\Users\User\ocr\kanban_refs\04_Documentacao\StockSAP.xlsx")


def load_rows(path: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def main() -> int:
    print(f"Loading NEW: {NEW.name}")
    new_rows = load_rows(NEW)
    print(f"  {len(new_rows)-1} rows")
    print(f"Loading BACKUP: {BACKUP.name}")
    bk_rows = load_rows(BACKUP)
    print(f"  {len(bk_rows)-1} rows")

    new_header = new_rows[0]
    bk_header = bk_rows[0]
    if new_header != bk_header:
        print(f"!! Header mismatch:")
        print(f"   NEW: {new_header}")
        print(f"   BK:  {bk_header}")
        return 1

    # Index by lote — NEW wins on overlap
    # Each lote may have multiple rows (different qtd/material entries);
    # keep all rows of NEW for NEW lotes, all rows of BACKUP for backup-only lotes.
    new_lotes: dict[str, list[tuple]] = {}
    for row in new_rows[1:]:
        if row[0] is None:
            continue
        lote = str(row[0]).strip().upper()
        new_lotes.setdefault(lote, []).append(row)

    bk_lotes: dict[str, list[tuple]] = {}
    for row in bk_rows[1:]:
        if row[0] is None:
            continue
        lote = str(row[0]).strip().upper()
        bk_lotes.setdefault(lote, []).append(row)

    # Merge
    merged: list[tuple] = []
    for lote, rows in new_lotes.items():
        merged.extend(rows)
    backup_only = set(bk_lotes.keys()) - set(new_lotes.keys())
    for lote in backup_only:
        merged.extend(bk_lotes[lote])

    print(f"NEW lotes: {len(new_lotes)}")
    print(f"BACKUP lotes: {len(bk_lotes)}")
    print(f"Common: {len(set(new_lotes) & set(bk_lotes))}")
    print(f"BACKUP-only (added as historical): {len(backup_only)}")
    print(f"Merged total lotes: {len(set(new_lotes) | set(bk_lotes))}")
    print(f"Merged total rows: {len(merged)}")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Folha1"
    out_ws.append(list(new_header))
    for row in merged:
        out_ws.append(list(row))
    out_wb.save(OUT)
    print(f"Wrote {OUT} ({OUT.stat().st_size:,} bytes)")

    shutil.copy(OUT, MIRROR)
    print(f"Mirrored to {MIRROR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
