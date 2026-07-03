"""R241 — Fit do CANAL DE VISÃO (matriz de confusão de caracteres) + medição
do CANAL HUMANO (quant4), a partir do app.db da fábrica.

O cross é um descodificador: P(escrito | verdadeiro) por carácter é o modelo
do canal de visão. Fitted por alinhamento Needleman-Wunsch carácter-a-carácter
sobre pares (raw, verdade):
  - pares FORTES: edits source='human' (old≠new) em of/ov/lote/pri;
  - pares LARGOS: linhas validadas raw vs final com distância de edição <= 2
    e |Δlen| <= 1 (misreads plausíveis; ~5-10% de contaminação por rewrites
    do motor — tolerável com suavização, documentado).
Prior de pseudo-contagens nos pares de glifos sabidamente confusos (0/O, 1/I,
3/8, 5/S, ...) — Bayes, não regra: os dados dominam onde existem.

quant4 (canal humano): das OFs escritas VÁLIDAS no plano mas diferentes da
final, que fração aponta para a MESMA família (cliente ou 1º token da
designação partilhados)? Calibra o relaxamento do veto R236.

Output: lexicons/cross_params.json (TRACKED — parâmetros são dados
versionados, como o plano), com contagens, data e proveniência.

Uso:
    uv run python scripts/diag/fit_char_confusion.py \
        --db ~/Downloads/auditoria_humana/app.db \
        --plan "~/Downloads/plan_colunas_cpis (8).xlsx"
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sqlite3
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

_REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_REPO / "backend"))

_ID_FIELDS = ("of", "ov", "lote", "pri")
_EDIT_RE = re.compile(r"^rows\[(\d+)\]\.(of|ov|lote|pri)$")

# Prior de glifos confusos (pseudo-contagens simétricas): conhecimento de
# domínio como PRIOR bayesiano — os dados dominam onde há contagens.
_GLYPH_PRIOR_PAIRS = [
    ("0", "O", 6), ("1", "I", 6), ("1", "L", 4), ("1", "7", 4), ("3", "8", 4),
    ("5", "S", 5), ("8", "B", 4), ("6", "G", 3), ("2", "Z", 4), ("7", "T", 3),
    ("4", "A", 2), ("9", "G", 2), ("0", "D", 2), ("6", "5", 2), ("9", "4", 2),
    ("3", "5", 2), ("2", "7", 2), ("1", "4", 2), ("M", "H", 3), ("E", "F", 2),
]
_PRIOR_ANY = 0.25          # suavização para qualquer par nunca visto
_PRIOR_MATCH = 50.0        # pseudo-contagem de matches por carácter


def _norm(s: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(s or "").upper())


def _lev(a: str, b: str) -> int:
    if a == b:
        return 0
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1,
                           prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def _nw_align(a: str, b: str) -> list[tuple[str, str]]:
    """Alinhamento global de custo unitário; devolve pares (char_a, char_b),
    com '-' para inserção/apagamento."""
    n, m = len(a), len(b)
    dp = [[0] * (m + 1) for _ in range(n + 1)]
    for i in range(n + 1):
        dp[i][0] = i
    for j in range(m + 1):
        dp[0][j] = j
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            dp[i][j] = min(dp[i - 1][j] + 1, dp[i][j - 1] + 1,
                           dp[i - 1][j - 1] + (a[i - 1] != b[j - 1]))
    out: list[tuple[str, str]] = []
    i, j = n, m
    while i > 0 or j > 0:
        if i > 0 and j > 0 and dp[i][j] == dp[i - 1][j - 1] + (a[i - 1] != b[j - 1]):
            out.append((a[i - 1], b[j - 1]))
            i, j = i - 1, j - 1
        elif i > 0 and dp[i][j] == dp[i - 1][j] + 1:
            out.append((a[i - 1], "-"))
            i -= 1
        else:
            out.append(("-", b[j - 1]))
            j -= 1
    return out[::-1]


def collect_pairs(db: Path) -> tuple[list[tuple[str, str, str]], dict]:
    con = sqlite3.connect(f"file:{db}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    pairs: list[tuple[str, str, str]] = []  # (truth, written, fonte)

    raw_cache: dict[int, dict] = {}

    def raw_rows(sheet_id: int) -> list[dict]:
        if sheet_id not in raw_cache:
            r = con.execute(
                "SELECT raw_extraction FROM sheets WHERE id=?", (sheet_id,)
            ).fetchone()
            try:
                raw_cache[sheet_id] = json.loads(r["raw_extraction"] or "{}")
            except (TypeError, ValueError):
                raw_cache[sheet_id] = {}
        return raw_cache[sheet_id].get("rows") or []

    # Pares FORTES: edições humanas.
    n_strong = 0
    for r in con.execute(
        "SELECT sheet_id, field_path, new_value FROM edits "
        "WHERE source='human' AND old_value <> new_value"
    ):
        m = _EDIT_RE.match(r["field_path"])
        if not m:
            continue
        i, field = int(m.group(1)), m.group(2)
        rows = raw_rows(r["sheet_id"])
        if i >= len(rows):
            continue
        w, t = _norm((rows[i] or {}).get(field)), _norm(r["new_value"])
        if w and t and w != t and abs(len(w) - len(t)) <= 2 and _lev(w, t) <= 3:
            pairs.append((t, w, "human"))
            n_strong += 1

    # Pares LARGOS: validadas, raw vs final, misreads plausíveis.
    n_broad = 0
    for s in con.execute(
        "SELECT sheet_data, raw_extraction FROM sheets "
        "WHERE status='validated' AND raw_extraction IS NOT NULL"
    ):
        try:
            raw = json.loads(s["raw_extraction"])
            fin = json.loads(s["sheet_data"])
        except (TypeError, ValueError):
            continue
        rr, rf = raw.get("rows") or [], fin.get("rows") or []
        for i in range(min(len(rr), len(rf))):
            for field in _ID_FIELDS:
                w = _norm((rr[i] or {}).get(field))
                t = _norm((rf[i] or {}).get(field))
                if (w and t and w != t and abs(len(w) - len(t)) <= 1
                        and _lev(w, t) <= 2):
                    pairs.append((t, w, "validated"))
                    n_broad += 1
    return pairs, {"strong_pairs": n_strong, "broad_pairs": n_broad}


def fit_matrix(pairs: list[tuple[str, str, str]]) -> dict:
    sub = Counter()
    ins = del_ = match = 0
    for truth, written, _src in pairs:
        for ct, cw in _nw_align(truth, written):
            if ct == "-":
                ins += 1
            elif cw == "-":
                del_ += 1
            elif ct == cw:
                match += 1
            else:
                sub[(ct, cw)] += 1
    # priors
    prior = defaultdict(float)
    for a, b, k in _GLYPH_PRIOR_PAIRS:
        prior[(a, b)] += k
        prior[(b, a)] += k
    total_pos = match + sum(sub.values()) + ins + del_ + _PRIOR_MATCH
    p_match = (match + _PRIOR_MATCH) / max(total_pos, 1)
    # custo em bits de cada substituição: log2(p_match / p_sub(a→b))
    keys = set(sub) | set(prior)
    costs = {}
    for k in keys:
        p_sub = (sub.get(k, 0) + prior.get(k, _PRIOR_ANY)) / max(total_pos, 1)
        costs[f"{k[0]}>{k[1]}"] = round(
            max(3.0, min(10.0, math.log2(p_match / p_sub))), 2)
    p_indel = (ins + del_ + 1) / max(total_pos, 1)
    return {
        "p_match": round(p_match, 4),
        "cost_default_bits": 10.0,       # substituição nunca vista
        "cost_indel_bits": round(max(3.0, min(10.0, math.log2(p_match / p_indel))), 2),
        "sub_costs_bits": dict(sorted(costs.items())),
        "counts": {"match": match, "sub": sum(sub.values()),
                   "ins": ins, "del": del_,
                   "top_subs": {f"{a}>{b}": c for (a, b), c in sub.most_common(25)}},
    }


def quant4_human_channel(db: Path, plan_xlsx: Path) -> dict:
    """Das OFs escritas VÁLIDAS mas ≠ da final: fração mesma-família."""
    import openpyxl

    wb = openpyxl.load_workbook(plan_xlsx, read_only=True, data_only=True)
    ws = wb["plan_colunas_cpis"] if "plan_colunas_cpis" in wb.sheetnames else wb.active
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").lower() for h in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    fam: dict[str, set[str]] = defaultdict(set)  # of -> {cliente_c | ft}
    for r in it:
        of = re.sub(r"\D", "", str(r[ix["of"]] or ""))
        if not of:
            continue
        of = of.zfill(6) if len(of) <= 6 else of
        cli = _norm(r[ix["cliente"]])
        des = str(r[ix["designacao"]] or "").strip().upper()
        ft = _norm(des.split()[0]) if des else ""
        if cli:
            fam[of].add("C:" + cli)
        if ft:
            fam[of].add("F:" + ft)

    con = sqlite3.connect(f"file:{db}?mode=ro", uri=True)
    n = same = 0
    for s in con.execute(
        "SELECT sheet_data, raw_extraction FROM sheets "
        "WHERE status='validated' AND raw_extraction IS NOT NULL"
    ):
        try:
            raw, fin = json.loads(s[1]), json.loads(s[0])
        except (TypeError, ValueError):
            continue
        rr, rf = raw.get("rows") or [], fin.get("rows") or []
        for i in range(min(len(rr), len(rf))):
            w = re.sub(r"\D", "", str((rr[i] or {}).get("of") or ""))
            t = re.sub(r"\D", "", str((rf[i] or {}).get("of") or ""))
            w = w.zfill(6) if w and len(w) <= 6 else w
            t = t.zfill(6) if t and len(t) <= 6 else t
            if w and t and w != t and w in fam and t in fam:
                n += 1
                if fam[w] & fam[t]:
                    same += 1
    frac = (same + 1) / (n + 2)  # Laplace
    return {
        "n_valid_but_wrong": n, "same_family": same,
        "p_same_family": round(frac, 3),
        # veto relaxado = -log2((1-m_of_valid)*p_same... pragmático:
        # veto_full=-3.3 (medido R236); relaxado proporcional à evidência de
        # família: -3.3 * (1 - p_same_family) — família comum torna o erro
        # humano PROVÁVEL, o veto encolhe.
        "veto_relaxed_bits": round(-3.3 * (1 - frac), 2),
    }


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--db", required=True, type=Path)
    ap.add_argument("--plan", required=True, type=Path)
    ap.add_argument("--out", type=Path,
                    default=_REPO / "lexicons" / "cross_params.json")
    args = ap.parse_args()

    pairs, meta = collect_pairs(args.db.expanduser())
    matrix = fit_matrix(pairs)
    q4 = quant4_human_channel(args.db.expanduser(), args.plan.expanduser())

    db_hash = hashlib.sha256(args.db.expanduser().read_bytes()).hexdigest()[:16]
    out = {
        "fitted_at": datetime.now(timezone.utc).isoformat(),
        "source_db_sha256_16": db_hash,
        "pair_counts": meta,
        "char_channel": matrix,
        "human_channel": q4,
    }
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(out, indent=2, ensure_ascii=False) + "\n",
                        encoding="utf-8")
    print(json.dumps({"pairs": meta, "p_match": matrix["p_match"],
                      "n_sub_costs": len(matrix["sub_costs_bits"]),
                      "top_subs": matrix["counts"]["top_subs"],
                      "human_channel": q4, "out": str(args.out)},
                     indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
