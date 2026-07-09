"""R242/R250/R252 — Medições dos priors de CONTEXTO (quant5-quant8), do app.db.

quant5 · Coerência de FOLHA: numa folha multi-linha, com que força as linhas
partilham cliente/OF? (calibra o passe 2 — a folha é uma conversa).

quant6 · Prior de PRODUÇÃO: a linha certa costuma ser uma OF com atividade
recente? w = log2(P(ativa | OF verdadeira) / P(ativa | OF aleatória do
plano)) — cap ±2 bits no motor (o prior nunca decide sozinho).

quant7 · OOD por IDADE do plano (R252): P(OF verdadeira fora do plano | a
folha é N dias mais antiga que o snapshot do plano). Alimenta o prior π_H0
da hipótese nula do posterior (antes era só descritivo — nada o lia).

quant8 · IDENTIDADE CONJUNTA (R250): m̂_A = P(todos os campos de A escritos
EXATOS | entry certa) para A ⊆ {of, ov, cliente} — os misreads são
correlacionados entre campos (operador cuidadoso escreve os 3 bem), logo
m̂_A > Π m_f. E ufloor_f = colisão exata média de um valor verdadeiro no
plano (análogo identidade do _FS_DIM_U_FLOOR).

Uso:
    uv run python scripts/diag/quant_context_priors.py \
        --db ~/Downloads/auditoria_humana/app.db \
        --plan "~/Downloads/plan_colunas_cpis (8).xlsx" \
        [--window-days 14] [--plan-day YYYY-MM-DD] [--merge]
        # --merge grava em lexicons/cross_params.json
"""
from __future__ import annotations

import argparse
import json
import math
import re
import sqlite3
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

_REPO = Path(__file__).resolve().parents[2]


def _norm_of(v: object) -> str:
    s = re.sub(r"\D", "", str(v or ""))
    return (s.zfill(6) if len(s) <= 6 else s) if s else ""


def _iso(v: object) -> date | None:
    s = str(v or "")[:10]
    try:
        d = date.fromisoformat(s)
    except ValueError:
        return None
    # lixo pontual (ex.: 2096) fora
    return d if date(2025, 1, 1) <= d <= date(2027, 12, 31) else None


# Normalizações-espelho do motor (locais de propósito — o script mede a
# régua do motor sem o importar; ver o mesmo padrão em backtest_winner).
def _ident_compact(v: object, *, pad6: bool = False) -> str:
    s = re.sub(r"[^A-Z0-9]", "", str(v or "").upper()).replace("O", "0")
    s = re.sub(r"\D", "", s)
    if pad6 and s and len(s) < 6:
        s = s.zfill(6)
    return s


def _cli_compact(v: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(v or "").upper())


_OOD_BUCKETS = (("0-3", 0, 3), ("4-7", 4, 7), ("8-14", 8, 14),
                ("15-30", 15, 30), (">30", 31, 10_000))


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--db", required=True, type=Path)
    ap.add_argument("--plan", required=True, type=Path)
    ap.add_argument("--window-days", type=int, default=14)
    ap.add_argument("--plan-day", type=str, default=None,
                    help="data do snapshot do plano (default: mtime do xlsx)")
    ap.add_argument("--merge", action="store_true")
    args = ap.parse_args()

    con = sqlite3.connect(f"file:{args.db.expanduser()}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row

    # ---------------- quant5: coerência de folha ----------------
    # A estatística CERTA é o LIFT vs pares aleatórios (não a taxa absoluta —
    # as folhas são mistas: só 18% têm cliente único; mas partilhar OF/cliente
    # com a linha ADJACENTE é 21x/8x mais provável do que ao acaso).
    from collections import Counter

    adj_cli = same_cli = adj_of_n = same_of = 0
    all_cli: Counter = Counter()
    all_of: Counter = Counter()
    for s in con.execute(
        "SELECT sheet_data FROM sheets WHERE status='validated'"
    ):
        try:
            fin = json.loads(s["sheet_data"] or "{}")
        except (TypeError, ValueError):
            continue
        rows = [r for r in (fin.get("rows") or [])
                if any(str(v or "").strip() for v in (r or {}).values())]
        clis = [str((r or {}).get("cliente") or "").strip().upper() for r in rows]
        ofs = [_norm_of((r or {}).get("of")) for r in rows]
        for c in clis:
            if c:
                all_cli[c] += 1
        for o in ofs:
            if o:
                all_of[o] += 1
        for a, b in zip(clis, clis[1:]):
            if a and b:
                adj_cli += 1
                same_cli += a == b
        for a, b in zip(ofs, ofs[1:]):
            if a and b:
                adj_of_n += 1
                same_of += a == b

    def _pcol(c: Counter) -> float:
        t = sum(c.values())
        return sum((n / t) ** 2 for n in c.values()) if t else 1.0

    p_adj_cli = (same_cli + 1) / (adj_cli + 2)
    p_adj_of = (same_of + 1) / (adj_of_n + 2)
    quant5 = {
        "pares_adjacentes_cliente": adj_cli, "mesmo_cliente": same_cli,
        "p_mesmo_cliente_adjacente": round(p_adj_cli, 3),
        "p_mesmo_cliente_aleatorio": round(_pcol(all_cli), 3),
        "pares_adjacentes_of": adj_of_n, "mesma_of": same_of,
        "p_mesma_of_adjacente": round(p_adj_of, 3),
        "p_mesma_of_aleatorio": round(_pcol(all_of), 4),
        # bits de coerência = log2(lift), cap ±2 aplicado no MOTOR
        "coherence_cliente_bits": round(
            math.log2(p_adj_cli / max(_pcol(all_cli), 1e-6)), 2),
        "coherence_of_bits": round(
            math.log2(p_adj_of / max(_pcol(all_of), 1e-6)), 2),
    }

    # ---------------- quant6: prior de produção ----------------
    # Atividade: OF com produção VALIDADA no intervalo [D-K, D-1].
    act_by_of: dict[str, list[date]] = defaultdict(list)
    for r in con.execute(
        "SELECT of, sheet_iso_date FROM production_rows "
        "WHERE sheet_status='validated' AND of IS NOT NULL"
    ):
        o, d = _norm_of(r["of"]), _iso(r["sheet_iso_date"])
        if o and d:
            act_by_of[o].append(d)
    for o in act_by_of:
        act_by_of[o].sort()

    def active(of_key: str, d: date) -> bool:
        lo = d - timedelta(days=args.window_days)
        return any(lo <= x < d for x in act_by_of.get(of_key, ()))

    import openpyxl

    wb = openpyxl.load_workbook(args.plan.expanduser(), read_only=True,
                                data_only=True)
    ws = wb["plan_colunas_cpis"] if "plan_colunas_cpis" in wb.sheetnames else wb.active
    it = ws.iter_rows(values_only=True)
    hdr = [str(h or "").lower() for h in next(it)]
    iof = hdr.index("of")
    plan_ofs = sorted({_norm_of(r[iof]) for r in it if _norm_of(r[iof])})

    n = true_act = 0
    rand_act = rand_n = 0
    for r in con.execute(
        "SELECT sheet_id, of, sheet_iso_date FROM production_rows "
        "WHERE sheet_status='validated' AND of IS NOT NULL ORDER BY id"
    ):
        o, d = _norm_of(r["of"]), _iso(r["sheet_iso_date"])
        if not o or not d:
            continue
        n += 1
        true_act += active(o, d)
        # contraste: 3 OFs "aleatórias" determinísticas do plano
        base = (r["sheet_id"] * 31 + n) % max(len(plan_ofs) - 3, 1)
        for k in range(3):
            rand_n += 1
            rand_act += active(plan_ofs[(base + k * 977) % len(plan_ofs)], d)
    p_true = (true_act + 1) / (n + 2)
    p_rand = (rand_act + 1) / (rand_n + 2)
    quant6 = {
        "window_days": args.window_days,
        "n_linhas": n, "ativa_quando_verdadeira": true_act,
        "p_ativa_true": round(p_true, 3),
        "n_controlo": rand_n, "ativa_controlo": rand_act,
        "p_ativa_random": round(p_rand, 3),
        "production_prior_bits": round(
            max(-2.0, min(2.0, math.log2(p_true / p_rand))), 2),
        "production_prior_inactive_bits": round(
            max(-2.0, min(2.0, math.log2((1 - p_true) / (1 - p_rand)))), 2),
    }

    # ---------------- quant7: OOD por idade do plano (R252) ----------------
    # OOD ⇔ a OF FINAL (verdade humana) não existe no snapshot do plano.
    # Idade = plan_day − data_da_folha (folhas mais antigas que o snapshot
    # referem encomendas que entretanto saíram do plano; em produção a
    # relação inverte-se — plano envelhece — e assumimos staleness simétrica).
    if args.plan_day:
        plan_day = date.fromisoformat(args.plan_day)
    else:
        from datetime import datetime as _dt
        plan_day = _dt.fromtimestamp(args.plan.expanduser().stat().st_mtime).date()
    plan_of_set = set(plan_ofs)
    b_stats = {name: [0, 0] for name, _, _ in _OOD_BUCKETS}  # [n, ood]
    for r in con.execute(
        "SELECT of, sheet_iso_date FROM production_rows "
        "WHERE sheet_status='validated' AND of IS NOT NULL"
    ):
        o, d = _norm_of(r["of"]), _iso(r["sheet_iso_date"])
        if not o or not d:
            continue
        age = max((plan_day - d).days, 0)
        for name, lo, hi in _OOD_BUCKETS:
            if lo <= age <= hi:
                b_stats[name][0] += 1
                b_stats[name][1] += o not in plan_of_set
                break
    quant7 = {
        "plan_day": plan_day.isoformat(),
        "buckets": {
            name: {
                "n": n, "ood": k,
                "p_ood": round((k + 1) / (n + 2), 3) if n else None,
            }
            for name, (n, k) in b_stats.items()
        },
    }
    # R253.2 — isotonic (PAVA, pesado por n) sobre a série por idade: os
    # buckets crus não são monótonos (8-14d: 31% > 15-30d: 19,5% com n
    # díspares, 305 vs 1391) mas o prior "plano mais velho => mais OOD" é —
    # o PAVA só funde violações adjacentes (média pesada), nunca inventa
    # sinal. Escreve p_ood_isotonic A PAR do cru (auditoria); _pi_h0 lê o
    # isotonic com fallback para o cru.
    ordered = [name for name, _, _ in _OOD_BUCKETS]
    blocks: list[list[float]] = []  # [soma_p*n, n, count_buckets]
    for name in ordered:
        n, k = b_stats[name]
        if not n:
            blocks.append([0.0, 0.0, 1])
            continue
        blocks.append([((k + 1) / (n + 2)) * n, float(n), 1])
        while (len(blocks) >= 2 and blocks[-2][1] > 0 and blocks[-1][1] > 0
               and blocks[-2][0] / blocks[-2][1]
               > blocks[-1][0] / blocks[-1][1]):
            b = blocks.pop()
            blocks[-1] = [blocks[-1][0] + b[0], blocks[-1][1] + b[1],
                          blocks[-1][2] + b[2]]
    iso: list[float | None] = []
    for wsum, wn, cnt in blocks:
        iso.extend([wsum / wn if wn else None] * int(cnt))
    for name, p in zip(ordered, iso):
        quant7["buckets"][name]["p_ood_isotonic"] = (
            round(p, 4) if p is not None else None)

    # ---------------- quant8: identidade conjunta (R250) ----------------
    # m̂_A sobre linhas validadas: raw vs FINAL (verdade humana) por campo,
    # exato segundo a régua do motor (compact + O→0; OF com zfill6).
    subsets = (("of",), ("ov",), ("cliente",),
               ("of", "ov"), ("of", "cliente"), ("ov", "cliente"),
               ("of", "ov", "cliente"))
    sub_stats = {"+".join(a): [0, 0] for a in subsets}  # [n_todos_escritos, todos_exatos]
    # ufloor: colisão exata média de um valor VERDADEIRO no plano.
    plan_freq_of: Counter = Counter()
    plan_freq_ov: Counter = Counter()
    plan_freq_cli: Counter = Counter()
    wb2 = openpyxl.load_workbook(args.plan.expanduser(), read_only=True,
                                 data_only=True)
    ws2 = (wb2["plan_colunas_cpis"]
           if "plan_colunas_cpis" in wb2.sheetnames else wb2.active)
    it2 = ws2.iter_rows(values_only=True)
    hdr2 = [str(h or "").lower() for h in next(it2)]
    i_of2, i_ov2, i_cli2 = hdr2.index("of"), hdr2.index("ov"), hdr2.index("cliente")
    n_plan = 0
    for r in it2:
        o = _norm_of(r[i_of2])
        if not o:
            continue
        n_plan += 1
        plan_freq_of[o] += 1
        ov_c = _ident_compact(r[i_ov2])
        if ov_c:
            plan_freq_ov[ov_c] += 1
        cli_c = _cli_compact(r[i_cli2])
        if cli_c:
            plan_freq_cli[cli_c] += 1
    ufloor_num = {"of": 0.0, "ov": 0.0, "cliente": 0.0}
    ufloor_den = {"of": 0, "ov": 0, "cliente": 0}

    def _row_vals(r: dict) -> dict:
        return {
            "of": _ident_compact((r or {}).get("of"), pad6=True),
            "ov": _ident_compact((r or {}).get("ov")),
            "cliente": _cli_compact((r or {}).get("cliente")),
        }

    for s in con.execute(
        "SELECT raw_extraction, sheet_data FROM sheets "
        "WHERE status='validated' AND raw_extraction IS NOT NULL"
    ):
        try:
            raw = json.loads(s["raw_extraction"] or "{}")
            fin = json.loads(s["sheet_data"] or "{}")
        except (TypeError, ValueError):
            continue
        rr, rf = raw.get("rows") or [], fin.get("rows") or []
        for i in range(min(len(rr), len(rf))):
            rv, fv = _row_vals(rr[i]), _row_vals(rf[i])
            written = {f for f in ("of", "ov", "cliente") if rv[f] and fv[f]}
            exact = {f for f in written if rv[f] == fv[f]}
            for a in subsets:
                if set(a) <= written:
                    key = "+".join(a)
                    sub_stats[key][0] += 1
                    sub_stats[key][1] += set(a) <= exact
            # ufloor com o valor VERDADEIRO (fv)
            freq = {"of": plan_freq_of, "ov": plan_freq_ov,
                    "cliente": plan_freq_cli}
            for f in ("of", "ov", "cliente"):
                if fv[f] and freq[f].get(fv[f]):
                    ufloor_num[f] += freq[f][fv[f]] / max(n_plan, 1)
                    ufloor_den[f] += 1
    quant8 = {
        "m_joint": {
            k: {"n": n, "exatos": e,
                "m": round((e + 1) / (n + 2), 3) if n else None}
            for k, (n, e) in sub_stats.items()
        },
        "u_floor": {
            f: round(ufloor_num[f] / ufloor_den[f], 5) if ufloor_den[f] else None
            for f in ("of", "ov", "cliente")
        },
        "n_plan": n_plan,
    }

    out = {"quant5_sheet_coherence": quant5, "quant6_production_prior": quant6,
           "quant7_ood_by_age": quant7, "quant8_identity_joint": quant8}
    print(json.dumps(out, indent=2, ensure_ascii=False))
    if args.merge:
        path = _REPO / "lexicons" / "cross_params.json"
        params = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}
        params.update(out)
        path.write_text(json.dumps(params, indent=2, ensure_ascii=False) + "\n",
                        encoding="utf-8")
        print(f"merged -> {path}")


if __name__ == "__main__":
    main()
