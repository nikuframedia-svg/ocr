"""FastAPI MVP — capture → review → dashboard.

Run:
    cd <repo-root>
    .venv/Scripts/python.exe -m uvicorn backend.app.web.main:app \\
        --reload --host 0.0.0.0 --port 8080
"""
from __future__ import annotations

import asyncio
import datetime as dt
import json
import os
import secrets
import sys
import threading
import traceback
from pathlib import Path

import httpx  # R120 — endpoint /admin/qwen-tools-test

from fastapi import FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

_REPO = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(_REPO))

# Ensure backend app importable
sys.path.insert(0, str(_REPO / "backend"))

from app import kernel  # noqa: E402  — R110.E event log (R117: hoisted for hot-path emits)
from app.web import attractors, db, export, kpis, llm_assistant, ocr_queue, ocr_runner  # noqa: E402
from app.cross_check import (  # noqa: E402
    cross_check_sheet,
    get_watcher,
    load_summary,
    load_to_analisar,
    store_cross_check,
)
from app.dq.operador_snap import snap_operador  # noqa: E402
from app.learning import (  # noqa: E402
    materialize as learning_materialize,
    metrics as learning_metrics,
    scheduler as learning_scheduler,
    store as learning_store,
)

# ----- App + paths -----
_TEMPLATES_DIR = Path(__file__).parent / "templates"
_STATIC_DIR = Path(__file__).parent / "static"
_DATA_DIR = _REPO / "data"
_IMAGES_DIR = _DATA_DIR / "images"
_DATA_DIR.mkdir(parents=True, exist_ok=True)
_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

# Operadores conhecidos (dropdown na review). Keep in sync com ground_truth.
OPERADORES = ("AUGUSTO MONTEIRO", "JÚLIO LIMA", "VITOR CARVALHO")

ROW_FIELDS = (
    "pri", "cliente", "ov", "of", "modelo", "qtd",
    "comp_mm", "larg_mm", "lote", "coni", "esp", "lbase", "ltopo",
)
HEADER_FIELDS = ("operador", "n_operador", "setor_maquina", "cod_maquina", "data")
FOOTER_FIELDS = ("colunas_produzidas", "horas_trabalhadas")


# Round 54 — per-sheet template context. Looks up the TemplateSpec for
# a sheet so the UI iterates the right row_fields / footer_fields.
# Legacy sheets (no template_name) fall back to bobine_formato via
# db.get_sheet_template_name's inference rules.
def _template_ctx_for_sheet(sheet: dict | None) -> dict:
    """Build template-aware context vars for rendering a sheet.

    Always returns keys: ``template`` (TemplateSpec or None),
    ``template_name``, ``row_fields``, ``footer_fields``, ``header_fields``.
    Backward-compat: if sheet is None, returns bobine_formato fields.
    """
    from app.templates_registry import DEFAULT_TEMPLATE, get_template
    if sheet is None:
        tpl = DEFAULT_TEMPLATE
    else:
        tname = db.get_sheet_template_name(sheet)
        tpl = get_template(tname)
    return {
        "template": tpl,
        "template_name": tpl.name,
        "row_fields": tpl.row_fields,
        "footer_fields": tpl.footer_fields,
        "header_fields": tpl.header_fields,
    }

app = FastAPI(title="Metalogalva OCR — MVP")
app.mount("/static", StaticFiles(directory=_STATIC_DIR), name="static")
templates = Jinja2Templates(directory=str(_TEMPLATES_DIR))

# R69 — production sectors available as a Jinja global so the export
# modal can iterate them without importing kpis in templates.
from app.web.kpis import PRODUCTION_SECTORS  # noqa: E402
templates.env.globals["production_sectors"] = PRODUCTION_SECTORS  # type: ignore[assignment]


def _process_sheet_ocr(sheet_id: int) -> None:
    """R71 — worker callback. Runs OCR + DQ + cross-check + CSV deposit
    for a single sheet. Identical pipeline to the old inline /upload
    logic but invoked from the background thread instead of the request
    handler.

    Idempotent: skips if sheet is no longer 'pending' (e.g. concurrent
    edit or already processed). Catches every Exception and persists
    error to the DB so the worker loop never crashes.
    """
    try:
        sheet = db.get_sheet(sheet_id)
        if sheet is None:
            return
        if sheet.get("status") != "pending":
            return  # idempotency — already processed by another invocation
        img_path = _DATA_DIR / sheet["image_path"]
        if not img_path.exists():
            db.update_error(sheet_id, "image file missing")
            return
        result = ocr_runner.run_pipeline(img_path)
        db.update_extraction(
            sheet_id=sheet_id,
            raw_extraction=result["raw"],
            dq_audit=result["dq"],
            sheet_data=result["current"],
        )
        try:
            _run_and_store_cross_check(sheet_id)
        except Exception as cc_err:  # noqa: BLE001
            print(f"[worker cross-check] sheet {sheet_id}: {cc_err}", file=sys.stderr)
            traceback.print_exc()
        try:
            _deposit_csv_to_factory(sheet_id)
        except Exception as dep_err:  # noqa: BLE001
            print(f"[worker deposit] sheet {sheet_id}: {dep_err}", file=sys.stderr)
    except Exception as e:  # noqa: BLE001
        try:
            db.update_error(sheet_id, f"{type(e).__name__}: {e}")
        except Exception:  # noqa: BLE001
            pass
        traceback.print_exc()


@app.on_event("startup")
def _startup() -> None:
    db.init_db()
    # R71 — boot background OCR worker + recover any sheets stuck in
    # status='pending' from a previous process. The 10s window skips
    # sheets that are about to be enqueued by /upload in flight right now.
    ocr_queue.start_worker(_process_sheet_ocr)
    n_recovered = ocr_queue.recover_pending(
        older_than_seconds=10,
        list_pending_fn=db.list_stuck_pending,
    )
    if n_recovered:
        print(f"[R71 startup] re-enqueued {n_recovered} pending sheet(s)", file=sys.stderr)


_MOBILE_UA_PATTERNS = ("mobile", "iphone", "android", "ipad", "ipod")


def _is_mobile_request(request: Request) -> bool:
    """Detect mobile clients by User-Agent. Used to:
    - hide sidebar (mobile sees full-screen capture only)
    - drive flow decisions (mobile = capture-only, no review/edit)
    """
    ua = (request.headers.get("user-agent") or "").lower()
    return any(p in ua for p in _MOBILE_UA_PATTERNS)


@app.middleware("http")
async def _attach_watermark(request: Request, call_next):
    """Inject a monotonic sheet watermark into every HTML response so
    the client banner ("X new sheets") can detect new uploads across
    HTMX polls. Adds a request-scoped attribute too so templates can
    embed it server-side on first paint. Also attaches a `mobile` flag
    used by templates to render mobile-only / desktop-only sections.
    """
    request.state.watermark = kpis.latest_sheet_id()
    request.state.mobile = _is_mobile_request(request)
    response = await call_next(request)
    # Header is always present; client uses it to detect increments.
    response.headers["X-Sheet-Watermark"] = str(request.state.watermark)
    return response




# ----- Pages -----

@app.get("/", response_class=HTMLResponse)
def home(request: Request) -> RedirectResponse:
    # R66 — mobile lands straight on the camera; desktop on the queue.
    target = "/capture" if _is_mobile_request(request) else "/queue"
    return RedirectResponse(target, status_code=302)


@app.get("/capture", response_class=HTMLResponse)
def capture_page(request: Request) -> Response:
    # R114 — operadores para dropdown de "Quem está a validar?" em
    # folhas com cesta (Expedição).
    return templates.TemplateResponse(
        request, "capture.html",
        {"operadores": OPERADORES},
    )


_MAX_UPLOAD_BYTES = 25 * 1024 * 1024  # 25 MB — phone photos are typically 3-10 MB


@app.post("/upload")
async def upload(
    image: UploadFile = File(...),
    return_mode: str | None = Query(default=None, alias="return"),
) -> Response:
    if not image.filename:
        raise HTTPException(400, "no filename")
    suffix = Path(image.filename).suffix.lower() or ".jpeg"
    if suffix not in (".jpeg", ".jpg", ".png", ".bmp", ".tiff", ".webp"):
        raise HTTPException(400, f"unsupported extension {suffix}")

    # Size check before reading the whole file. Starlette exposes size via
    # Content-Length when available (mobile uploads usually have it).
    if image.size is not None and image.size > _MAX_UPLOAD_BYTES:
        raise HTTPException(
            413, f"image too large ({image.size} bytes > {_MAX_UPLOAD_BYTES})"
        )

    token = secrets.token_hex(8)
    target = _IMAGES_DIR / f"{dt.datetime.now():%Y%m%d_%H%M%S}_{token}{suffix}"
    bytes_written = 0
    with target.open("wb") as f:
        while chunk := await image.read(1024 * 1024):
            bytes_written += len(chunk)
            if bytes_written > _MAX_UPLOAD_BYTES:
                f.close()
                target.unlink(missing_ok=True)
                raise HTTPException(413, "image exceeded size limit")
            f.write(chunk)

    rel_path = str(target.relative_to(_DATA_DIR))
    sheet_id = db.insert_sheet(rel_path)

    # R117 — kernel event: folha aceite pelo servidor (pré-OCR).
    try:
        kernel.emit_event("sheet_uploaded", {"sheet_id": sheet_id, "image_path": rel_path})
    except Exception:  # noqa: BLE001
        pass

    # Round 46 — auto-crop kanban paper from photo (background removed,
    # perspective corrected). Saved as <stem>_cropped.jpg next to original.
    # Silent no-op if detection fails (image route falls back to original).
    try:
        from .image_crop import auto_crop
        auto_crop(target)
    except Exception as crop_err:  # noqa: BLE001
        print(f"[auto-crop] sheet upload {target.name}: {crop_err}", file=sys.stderr)

    # R71 — enqueue for background OCR processing. Returns near-instantly
    # (single SQLite INSERT + queue.put, ~50ms total) instead of blocking
    # ~22s for OCR. Worker thread (in ocr_queue) drains FIFO serially,
    # matching Ollama's single-inference GPU constraint while letting
    # /upload accept N requests in parallel without timeouts.
    queue_pos = ocr_queue.enqueue(sheet_id)

    # Mobile flow uses `?return=json` to drive the polling UI from JS.
    # Returns sheet_id + queue_pos for immediate display; status updates
    # come via subsequent /sheet/{id}/status polls.
    if return_mode == "json":
        return JSONResponse(
            {
                "sheet_id": sheet_id,
                "status": "pending",
                "queue_pos": queue_pos,
                "error": None,
            },
            status_code=200,
        )
    return RedirectResponse(f"/sheet/{sheet_id}", status_code=303)


# --- Cross-check helper (Round 33: pure verification) ---
# R124: política de substituição vive em `_apply_auto_overwrites` e
# `_maybe_apply_snap` — cada célula decide pelo seu `engine_status` e
# `source`. As constantes R61/R66 anteriores (campos hard-coded) saíram
# por dead code; R109 substituiu-as pela flag `snapped` por célula.

# Campos dimensionais — `very_different` aqui fica para revisão humana
# (não auto-aplicado), porque o plan/SAP pode estar desactualizado e o
# operador viu a peça real. `snapped` (delta pequena) é sempre aplicado.
_DIM_FIELDS = ("comp_mm", "larg_mm", "lbase", "ltopo", "esp")


def _maybe_apply_snap(sheet_id: int, field_path: str, cell: dict) -> bool:
    """R124 — aplica o canonical proposto pelo motor para uma célula.

    Política:
      - `snapped` (motor confiante, delta suave) → aplica sempre.
      - `very_different` (motor confiante mas delta grande) → aplica
        APENAS para campos não-dimensionais (modelo, cliente, of, ov,
        operador, cod_maquina, etc.) e só quando o motor TEM proposta
        concreta vinda de uma ref (`source` != "ocr_raw"). Dimensões
        físicas ficam para revisão humana porque o plan/SAP pode estar
        desactualizado e o operador viu a peça real.
      - Outros estados (`confirmed`, `NA`) → no-op.

    Retorna True quando aplicou um edit.
    """
    # R125 — quando a linha do plan está totalmente concluída na etapa
    # actual, o motor marca cada cell com source="obra_concluida". Não
    # auto-substituir: o operador tem de investigar (kanban a tentar
    # registar produção numa obra fechada).
    if cell.get("source") == "obra_concluida":
        return False
    engine_status = cell.get("engine_status")
    if engine_status == "snapped":
        pass
    elif engine_status == "very_different":
        field_name = field_path.rsplit(".", 1)[-1]
        if field_name in _DIM_FIELDS:
            return False
        if cell.get("source") in (None, "ocr_raw"):
            return False  # fallback do motor sem proposta concreta — no-op
    else:
        return False
    canonical = (cell.get("value") or "").strip()
    if not canonical:
        return False
    try:
        db.apply_edit(sheet_id, field_path, canonical, source="system")
        return True
    except Exception:  # noqa: BLE001
        return False


def _apply_auto_overwrites(sheet_id: int, result: dict) -> int:
    """R109/R124 — aplica snaps em rows + header + footer.

    R109 introduziu o ciclo sobre `result["rows"]` e `snapped=True`.
    R124 estende:
      - cobre também `result["header"]` e `result["footer"]` (motor já
        produz cells nessas secções desde R123 Fase 4 B9);
      - aplica `very_different` em campos não-dimensionais quando o
        motor tem proposta concreta (modelo/cliente/of/ov, cod_maquina,
        etc. que antes ficavam vermelhos sem auto-correcção).

    Ver `_maybe_apply_snap` para a regra por célula.
    """
    n_applied = 0
    for row_r in result.get("rows", []):
        i = row_r.get("row_index")
        if i is None:
            continue
        for fn, cell in row_r.get("fields", {}).items():
            if _maybe_apply_snap(sheet_id, f"rows[{i}].{fn}", cell):
                n_applied += 1
    for section in ("header", "footer"):
        for fn, cell in (result.get(section) or {}).items():
            if _maybe_apply_snap(sheet_id, f"{section}.{fn}", cell):
                n_applied += 1
    return n_applied


def _apply_operador_snap(sheet_id: int, sheet: dict, refs: dict) -> int:
    """R70 — resolve operator identity against ListaColaboradores.

    Reads ``header.operador`` + ``header.n_operador`` from sheet_data,
    runs ``snap_operador`` against ``refs["colaboradores"]``, and persists
    canonical values via ``db.apply_edit`` when:
      - ``snapped_name`` differs from current operador, OR
      - ``snapped_cod`` differs from current n_operador, OR
      - ``pernr`` is set and differs from current header.pernr

    Returns count of fields edited (0-3). Suspended cells (yellow flag)
    do not trigger edits; engine handles the visual flag separately.
    """
    colabs = refs.get("colaboradores") or {}
    if not colabs:
        return 0
    header = (sheet.get("sheet_data") or {}).get("header") or {}
    raw_name = (header.get("operador") or "").strip()
    raw_cod = (header.get("n_operador") or "").strip()
    cur_pernr = (header.get("pernr") or "").strip()

    if not raw_name and not raw_cod:
        return 0

    # R91: pass aliases lexicon for memorized OCR-corrupt → canonical
    # mappings. Empty dict if no aliases registered yet.
    aliases = refs.get("operador_aliases") or {}
    sr = snap_operador(raw_name, raw_cod, colabs, aliases=aliases)

    n_applied = 0

    # Persist pernr whenever we have a confident match (HIGH levels A/B/C),
    # even if no-op on name/cod (Condition A still has pernr to record).
    if sr.pernr and sr.pernr != cur_pernr:
        try:
            db.apply_edit(sheet_id, "header.pernr", sr.pernr, source="system")
            n_applied += 1
        except Exception:  # noqa: BLE001
            pass

    # Snap name when changed
    if sr.applied and sr.snapped_name and sr.snapped_name != raw_name:
        try:
            db.apply_edit(sheet_id, "header.operador", sr.snapped_name, source="system")
            n_applied += 1
        except Exception:  # noqa: BLE001
            pass

    # Snap cod when changed (Condition C — Lev-1)
    if sr.applied and sr.snapped_cod and sr.snapped_cod != raw_cod:
        try:
            db.apply_edit(sheet_id, "header.n_operador", sr.snapped_cod, source="system")
            n_applied += 1
        except Exception:  # noqa: BLE001
            pass

    return n_applied


def _apply_codmaq_fill(sheet_id: int, sheet: dict, refs: dict) -> int:
    """R85/R124 — fill OR correct header.cod_maquina from setor_maquina.

    Looks up ``header.setor_maquina`` (e.g. "HPE32", "GUIFIL", "LASER")
    in ``refs["maquinas_by_kanban"]`` and writes the canonical
    ``codmaq`` (M024 / M067 / M030) to ``header.cod_maquina``.

    R124: substitui também quando o operador escreveu um cod errado —
    antes só preenchia se estivesse vazio. O setor é a chave fiável (vem
    do template da folha); o cod é derivável e portanto sobrescrevível.
    Skipped quando o setor não mapeia unambiguamente (ex: "GUILHOTINA"
    sem largura — registry tem GUILHOTINA 3M/6M/9M/10M).

    Returns 1 if applied, 0 otherwise.
    """
    header = (sheet.get("sheet_data") or {}).get("header") or {}
    setor = (header.get("setor_maquina") or "").strip().upper()
    if not setor:
        return 0
    maq = (refs.get("maquinas_by_kanban") or {}).get(setor)
    if not maq or not maq.get("codmaq"):
        return 0
    canonical = maq["codmaq"]
    current = (header.get("cod_maquina") or "").strip()
    if current == canonical:
        return 0  # já igual — nada a fazer
    try:
        db.apply_edit(sheet_id, "header.cod_maquina", canonical, source="system")
        return 1
    except Exception:  # noqa: BLE001
        return 0


def _run_and_store_cross_check(sheet_id: int) -> dict | None:
    """Round 33 — invisible verification inline in /upload pipeline.

    R109/R123 — para QUALQUER célula que o motor marque como ``snapped``
    (escolheu ou preencheu um valor diferente do OCR), ``_apply_auto_overwrites``
    escreve o valor canónico do plano na sheet_data. Células ``very_different``
    (vermelho) ficam para revisão humana — nunca são aplicadas.

    Steps:
      1. Run cross_check_sheet → per-cell status against refs
      2. apply_auto_overwrites (todas as células snapped) + operador snap
         + cod_maquina fill
      3. If any edits were applied, re-run cross_check_sheet on the updated
         sheet_data so the persisted JSON reflects the final state
      4. Persist JSON to ``C:\\kanban\\nifruka\\03_Cross_Check\\``
    """
    sheet = db.get_sheet(sheet_id)
    if sheet is None or not sheet.get("sheet_data"):
        return None
    refs = get_watcher().get_refs()
    if not refs.get("available"):
        return None

    result = cross_check_sheet(sheet["sheet_data"], sheet.get("dq_audit"), refs)

    # R61 — auto-overwrite modelo/cliente when MATCH but value diverges
    n_overwritten = _apply_auto_overwrites(sheet_id, result)
    # R70 — operator snap against ListaColaboradores (SAP employee list).
    # Resolves OCR name/cod against canonical sname/cod/pernr and applies
    # auto-substitution when there's strong identity signal (cod + token
    # overlap). See backend/app/dq/operador_snap.py for the 5 rules.
    n_op_snapped = _apply_operador_snap(sheet_id, sheet, refs)
    # R85 — auto-fill cod_maquina from setor_maquina via maquinas.xlsx
    # lookup. Fills empty cod_maquina when setor maps to a known machine.
    n_codmaq_filled = _apply_codmaq_fill(sheet_id, sheet, refs)
    if n_overwritten > 0 or n_op_snapped > 0 or n_codmaq_filled > 0:
        # Re-fetch sheet (sheet_data was modified by apply_edit) and
        # re-run cross-check to refresh statuses against new values.
        # R80 note: the `*` indicator is derived in _build_snapped_map_from_raw
        # by comparing current sheet_data vs raw_extraction, so it survives
        # this re-run automatically (no need to preserve `snapped` flags
        # on the cross-check JSON).
        refreshed = db.get_sheet(sheet_id)
        if refreshed is not None and refreshed.get("sheet_data"):
            sheet = refreshed
            result = cross_check_sheet(sheet["sheet_data"], sheet.get("dq_audit"), refs)

    if sheet is None or not sheet.get("sheet_data"):
        return result  # defensive — should not happen
    header = sheet["sheet_data"].get("header", {}) or {}
    operador = header.get("operador") or sheet.get("operador") or "?"
    date_pt = (header.get("data") or "").strip()
    date_iso = date_pt
    if len(date_pt) == 10 and date_pt[2] == "-":
        date_iso = f"{date_pt[6:10]}-{date_pt[3:5]}-{date_pt[0:2]}"
    store_cross_check(
        sheet_id=sheet_id,
        image_path=sheet["image_path"],
        operador=operador,
        date_iso=date_iso,
        sheet_status=sheet["status"],
        cross_check_result=result,
    )
    # R108 — shadow scoring engine corre em background, escreve em coluna
    # própria. Não bloqueia, não interfere com `result`. Try/except wrap
    # garante que qualquer falha no shadow não toca em produção.
    _spawn_shadow_scoring(sheet_id, sheet["sheet_data"], sheet.get("dq_audit"), refs)
    return result


def _spawn_shadow_scoring(
    sheet_id: int,
    sheet_data: dict,
    dq_audit: dict | None,
    refs: dict,
) -> None:
    """R108 — dispara `scoring_engine.shadow_score` em thread daemon.

    Devolve imediatamente. Erros silenciados (logged) — shadow nunca
    deve afectar o output de produção.
    """
    def _run() -> None:
        try:
            from app.pipeline.scoring_engine import shadow_score
            run_id = db.start_shadow_run(sheet_id)
            try:
                scoring, total, snapped, confirmed, na, dur_ms = shadow_score(
                    sheet_data, dq_audit, refs
                )
                db.finish_shadow_run(
                    run_id, sheet_id, scoring,
                    total, snapped, confirmed, na, dur_ms,
                )
                # R117 — kernel event: shadow scoring concluído (thread daemon,
                # kernel.emit_event é thread-safe via _LOCK).
                try:
                    kernel.emit_event("shadow_run_completed", {
                        "sheet_id": sheet_id,
                        "total": total,
                        "snapped": snapped,
                        "confirmed": confirmed,
                        "na": na,
                        "duration_ms": dur_ms,
                    })
                except Exception:  # noqa: BLE001
                    pass
            except Exception as exc:
                db.fail_shadow_run(run_id, f"{type(exc).__name__}: {exc}")
                traceback.print_exc()
        except Exception:
            # Falha a abrir o run sequer — não devia acontecer, mas guarda
            traceback.print_exc()

    threading.Thread(
        target=_run, daemon=True, name=f"shadow-score-{sheet_id}"
    ).start()


@app.get("/sheet/{sheet_id}", response_class=HTMLResponse)
def sheet_page(
    request: Request,
    sheet_id: int,
    view: str | None = None,
    back: str | None = None,
) -> Response:
    """Round 41d: ``view`` query param toggles between 'final' (default,
    post-snap + manual edits) and 'raw' (original OCR before any auto-fix).
    Lets supervisor see what the OCR actually extracted vs what the system
    auto-corrected via DQ snap + cross-check.

    R88: ``back`` is an optional URL where the "← Voltar à lista" button
    should point (typically /queue?... or /kanbans?... with active
    filters). Only relative paths are accepted (open-redirect guard).
    """
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")

    # R88 — validate back URL: relative path only, no scheme/host
    back_url: str | None = None
    if back:
        candidate = back.strip()
        if candidate.startswith("/") and not candidate.startswith("//"):
            back_url = candidate

    cells_by_path: dict[str, dict] = {}
    if sheet.get("dq_audit"):
        cells_by_path = sheet["dq_audit"].get("cells", {})

    view_mode = "raw" if view == "raw" else "final"
    if view_mode == "raw" and sheet.get("raw_extraction"):
        # Show original OCR — disable cross-check colors (they validate
        # the post-snap values, not raw)
        src = sheet["raw_extraction"]
        (cc_status_by_path, cc_ref_by_path, cc_suspended_by_path,
         cc_snapped_by_path, cc_obra_concluida_by_path) = (
            {}, {}, {}, {}, {},
        )
    else:
        src = sheet.get("sheet_data") or {}
        (cc_status_by_path, cc_ref_by_path, cc_suspended_by_path,
         cc_snapped_by_path, cc_obra_concluida_by_path) = (
            _build_cc_maps(sheet_id)
        )

    rows = src.get("rows", []) or []
    header = src.get("header", {}) or {}
    footer = src.get("footer", {}) or {}

    flagged = sum(1 for c in cells_by_path.values() if c.get("requires_review"))

    tpl_ctx = _template_ctx_for_sheet(sheet)

    # R94 — pre-fill ISO date for the validation form's <input type="date">.
    # header.data is DD-MM-YYYY (and variations); normalize defensively.
    from app.web.db import _normalize_data_pt_to_iso  # local import to avoid cycle
    data_iso_for_validate = _normalize_data_pt_to_iso(header.get("data"))

    # R111 — flag para a UI saber se a imagem servida em /image/<id> é cropped
    # (paper detectado) ou raw (fallback silencioso). Quando False, sheet.html
    # mostra badge + botão "Tentar recortar agora".
    from .image_crop import has_cropped as _has_cropped
    sheet_has_cropped = _has_cropped(_DATA_DIR / sheet["image_path"])

    return templates.TemplateResponse(
        request,
        "sheet.html",
        {
            "sheet": sheet,
            "header": header,
            "rows": rows,
            "footer": footer,
            "cells_by_path": cells_by_path,
            "cc_status_by_path": cc_status_by_path,
            "cc_ref_by_path": cc_ref_by_path,
            "cc_suspended_by_path": cc_suspended_by_path,
            "cc_snapped_by_path": cc_snapped_by_path,
            "cc_obra_concluida_by_path": cc_obra_concluida_by_path,
            "operadores": OPERADORES,
            "flagged_count": flagged,
            "view_mode": view_mode,
            "back_url": back_url,
            "data_iso_for_validate": data_iso_for_validate,
            "has_cropped": sheet_has_cropped,
            **tpl_ctx,  # template, template_name, row/footer/header_fields
        },
    )


def _build_snapped_map_from_raw(sheet: dict) -> dict[str, bool]:
    """R80 — compute {field_path: True} for cells whose current value
    differs from the raw OCR extraction. Captures every cell that was
    modified after upload (auto-correction via DQ snap, cross-check
    overwrite, or operator manual edit). Used to render the `*` indicator
    on cells that aren't the OCR original.

    Compares ``sheet_data`` (current) against ``raw_extraction`` (snapshot
    at upload time). Fields covered: rows[*].*, header.*, footer.*.

    Returns {} if no raw_extraction available.
    """
    raw = sheet.get("raw_extraction") or {}
    cur = sheet.get("sheet_data") or {}
    if not raw or not cur:
        return {}

    def _norm(v: object) -> str:
        return str(v).strip() if v is not None else ""

    out: dict[str, bool] = {}

    raw_rows = raw.get("rows") or []
    cur_rows = cur.get("rows") or []
    for i, cur_r in enumerate(cur_rows):
        if i >= len(raw_rows):
            break
        raw_r = raw_rows[i] or {}
        for fn in (cur_r or {}).keys():
            if _norm(cur_r.get(fn)) != _norm(raw_r.get(fn)):
                out[f"rows[{i}].{fn}"] = True

    for section in ("header", "footer"):
        raw_sec = raw.get(section) or {}
        cur_sec = cur.get(section) or {}
        for fn in cur_sec.keys():
            if _norm(cur_sec.get(fn)) != _norm(raw_sec.get(fn)):
                out[f"{section}.{fn}"] = True

    return out


def _build_cc_maps(sheet_id: int) -> tuple[
    dict[str, str], dict[str, str], dict[str, bool], dict[str, bool],
    dict[str, bool],
]:
    """Round 33: load cross-check JSON for sheet, build {field_path: status}
    + {field_path: ref} maps for template rendering of green/red cell colors.

    R52 F4: also returns {field_path: suspended_by_stub} for distinguishing
    NA from stub-accept (amarelo soft) vs NA from no-ref (cinza).

    R80: also returns {field_path: snapped} derived from comparing current
    ``sheet_data`` against ``raw_extraction``. Captures every cell that was
    modified after upload (auto-correction or manual edit). Used for the
    `*` indicator showing operator which values aren't the OCR original.

    Returns ({}, {}, {}, {}) if no cross-check data available."""
    from app.cross_check.storage import load_sheet_cross_check
    from app.pipeline.scoring_engine import ENGINE_VERSION
    cc = load_sheet_cross_check(sheet_id)
    # R123 (D1) — fallback on-demand. A folha nunca teve cross-check
    # (processada antes do R118) ou o JSON é de um motor anterior ao R123:
    # regenera-o agora e relê, para nenhuma folha abrir toda cinza nem com
    # cores de um motor antigo.
    if not cc or cc.get("engine_version") != ENGINE_VERSION:
        try:
            _run_and_store_cross_check(sheet_id)
            cc = load_sheet_cross_check(sheet_id)
        except Exception:  # noqa: BLE001
            pass
    snapped_map = _build_snapped_map_from_raw(db.get_sheet(sheet_id) or {})
    if not cc:
        return {}, {}, {}, snapped_map, {}
    status_map: dict[str, str] = {}
    ref_map: dict[str, str] = {}
    suspended_map: dict[str, bool] = {}
    # R125 — paths cujas cells foram marcadas pelo motor com
    # source="obra_concluida" (todas as linhas do plan para esta OF
    # estão fechadas na etapa actual). Usado pelo template para mostrar
    # tooltip "obra concluída — verificar".
    obra_concluida_map: dict[str, bool] = {}
    for r in cc.get("rows", []):
        i = r.get("row_index")
        for f, info in (r.get("fields") or {}).items():
            path = f"rows[{i}].{f}"
            status_map[path] = info.get("status", "NA")
            ref = info.get("ref")
            if ref is not None:
                ref_map[path] = str(ref)
            if info.get("suspended_by_stub"):
                suspended_map[path] = True
            if info.get("source") == "obra_concluida":
                obra_concluida_map[path] = True
    # R123 (B9) — header/footer também coloridos (operador, data, máquina,
    # colunas_produzidas, ...). Cross-checks gravados antes do R123 não os
    # têm — `cc.get(section)` devolve {} e o cabeçalho fica neutro.
    for section in ("header", "footer"):
        for f, info in (cc.get(section) or {}).items():
            path = f"{section}.{f}"
            status_map[path] = info.get("status", "NA")
            ref = info.get("ref")
            if ref is not None:
                ref_map[path] = str(ref)
    return status_map, ref_map, suspended_map, snapped_map, obra_concluida_map


def _maybe_record_operador_alias(sheet_id: int) -> None:
    """R91 — Persist OCR-corrupt → canonical mapping in operator aliases.

    Triggered after the operator manually edits header.operador or
    header.n_operador in the UI. If the new (name, cod) combination
    resolves to a known colaborador AND the original OCR-captured name
    was different, save it so future OCRs of the same corrupt string
    resolve directly via the lexicon (skip confusion-map).

    Idempotent + safe — silently bails when:
      - sheet has no raw_extraction OR no header
      - edit doesn't yield a numeric cod
      - cod not in colaboradores
      - canonical sname doesn't match the operator-typed name
    """
    import datetime as _dt
    sheet = db.get_sheet(sheet_id)
    if not sheet or not sheet.get("sheet_data"):
        return
    header = (sheet.get("sheet_data") or {}).get("header") or {}
    raw_header = (sheet.get("raw_extraction") or {}).get("header") or {}
    raw_ocr_name = str(raw_header.get("operador") or "").strip()
    new_name = str(header.get("operador") or "").strip()
    new_cod_str = str(header.get("n_operador") or "").strip()
    if not raw_ocr_name or not new_name:
        return
    if raw_ocr_name.upper() == new_name.upper():
        return  # nothing to memorize — operator's edit didn't change the name
    try:
        cod_int = int(new_cod_str.lstrip("0") or "0")
    except ValueError:
        return
    if cod_int <= 0:
        return
    refs = get_watcher().get_refs()
    colabs = refs.get("colaboradores") or {}
    entry = colabs.get(cod_int)
    if not entry or entry.get("sname", "").upper() != new_name.upper():
        return  # edit doesn't resolve to a known colaborador
    # Persist alias — keyed by normalized OCR name (UPPER + ASCII-stripped)
    import unicodedata
    nfd = unicodedata.normalize("NFD", raw_ocr_name)
    key = " ".join("".join(ch for ch in nfd if not unicodedata.combining(ch)).upper().split())
    if not key:
        return
    aliases_path = _REPO / "lexicons" / "operador_aliases.json"
    aliases: dict = {}
    if aliases_path.exists():
        try:
            aliases = json.loads(aliases_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            aliases = {}
    aliases[key] = {
        "cod": cod_int,
        "pernr": entry.get("pernr", ""),
        "sname": entry.get("sname", ""),
        "source": "manual",
        "added_at": _dt.datetime.now(tz=_dt.timezone.utc).isoformat(timespec="seconds"),
    }
    aliases_path.parent.mkdir(parents=True, exist_ok=True)
    aliases_path.write_text(
        json.dumps(aliases, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


@app.post("/sheet/{sheet_id}/edit", response_class=HTMLResponse)
async def sheet_edit(
    request: Request,
    sheet_id: int,
    field_path: str = Form(...),
    new_value: str = Form(""),
) -> Response:
    """HTMX endpoint — applies an edit + returns the cell HTML fragment."""
    # Round 34 — mobile cannot do full edits (only via /mobile/qtds-batch
    # which has whitelist of qty-related fields)
    if _is_mobile_request(request):
        raise HTTPException(403, "Edição só pode ser feita em desktop")
    # Round 50 — folha validada é read-only; sem edits posteriores.
    sheet_pre = db.get_sheet(sheet_id)
    if sheet_pre is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    if sheet_pre.get("status") == "validated":
        raise HTTPException(409, "Folha já validada — edits bloqueados")
    try:
        old, new = db.apply_edit(sheet_id, field_path, new_value)
    except ValueError as e:
        raise HTTPException(400, str(e))
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    # R91 — after editing header.operador or header.n_operador, try to
    # memorize the OCR→canonical mapping in the aliases lexicon so future
    # OCRs of the same corrupt name resolve directly.
    if field_path in ("header.operador", "header.n_operador"):
        try:
            _maybe_record_operador_alias(sheet_id)
        except Exception as e:  # noqa: BLE001
            print(f"[alias] sheet {sheet_id}: {e}", file=sys.stderr)
    # Re-run cross-check after edit (auto-fill / status may shift)
    try:
        _run_and_store_cross_check(sheet_id)
    except Exception as cc_err:  # noqa: BLE001
        print(f"[cross-check] sheet {sheet_id} edit: {cc_err}", file=sys.stderr)
    # R123 — re-ler a folha DEPOIS do cross-check: _apply_auto_overwrites /
    # _apply_operador_snap / _apply_codmaq_fill podem ter reescrito sheet_data
    # com o valor canónico. Devolver `new` (valor submetido) fazia a célula
    # mostrar uma coisa e a DB guardar outra → o valor "mudava sozinho" no
    # próximo reload. Devolvemos o valor REAL persistido.
    sheet = db.get_sheet(sheet_id) or sheet
    try:
        real_value = db._get_by_path(sheet.get("sheet_data") or {}, field_path)
    except Exception:  # noqa: BLE001
        real_value = new
    if real_value is None:
        real_value = new
    cells_by_path = (sheet.get("dq_audit") or {}).get("cells", {})
    (cc_status_by_path, cc_ref_by_path, cc_suspended_by_path,
     cc_snapped_by_path, cc_obra_concluida_by_path) = (
        _build_cc_maps(sheet_id)
    )
    return templates.TemplateResponse(
        request,
        "_cell.html",
        {
            "sheet_id": sheet_id,
            "field_path": field_path,
            "value": real_value,
            "audit": cells_by_path.get(field_path, {}),
            "edited": old != real_value,
            "cc_status_by_path": cc_status_by_path,
            "cc_ref_by_path": cc_ref_by_path,
            "cc_suspended_by_path": cc_suspended_by_path,
            "cc_snapped_by_path": cc_snapped_by_path,
            "cc_obra_concluida_by_path": cc_obra_concluida_by_path,
            "sheet_status": sheet.get("status"),
        },
    )


# Factory deposit: CSVs go here automatically when a sheet is validated.
# Defaults to the local factory clone in C:\kanban\nifruka\... (set up
# by the user in this workspace). Set FACTORY_CSV_DIR env var to override
# or to "" to disable auto-deposit.
# R118 — usar resolve_kanban_path para cair em repo-local quando o disco
# C:\kanban\ não existe (laptop dev sem .env).
if os.environ.get("FACTORY_CSV_DIR", "_DEFAULT_") != "":
    from app.config import resolve_kanban_path
    _FACTORY_CSV_DIR = resolve_kanban_path(
        "FACTORY_CSV_DIR",
        r"C:\kanban\nifruka\02_Dados_Extraidos\csv",
        "kanban_refs/02_Dados_Extraidos/csv",
    )
else:
    _FACTORY_CSV_DIR = None


def _factory_csv_filename(sheet: dict) -> str:
    """Return the canonical filename to use when depositing a sheet's CSV
    in the factory dir. Prefer the operador+date encoded in sheet_data
    (e.g. ``JulioLima_2026.04.15-1.csv``) so the validator log lines stay
    human-readable; fall back to the raw image stem.
    """
    data = sheet.get("sheet_data") or {}
    h = data.get("header", {}) or {}
    operador = (h.get("operador") or "").strip().title().replace(" ", "")
    data_str = (h.get("data") or "").strip()
    if operador and data_str:
        # 15-04-2026 → 2026.04.15
        parts = data_str.split("-")
        if len(parts) == 3:
            iso = f"{parts[2]}.{parts[1]}.{parts[0]}"
            return f"{operador}_{iso}.csv"
    return f"{Path(sheet['image_path']).stem}.csv"


def _deposit_csv_to_factory(sheet_id: int) -> Path | None:
    """Write the sheet's 3-block CSV to FACTORY_CSV_DIR. Returns the path
    written, or None if the deposit dir isn't configured / doesn't exist /
    the sheet has no data. Idempotent: overwrites existing file."""
    if _FACTORY_CSV_DIR is None or not _FACTORY_CSV_DIR.exists():
        return None
    sheet = db.get_sheet(sheet_id)
    if sheet is None or not sheet.get("sheet_data"):
        return None
    filename = _factory_csv_filename(sheet)
    target = _FACTORY_CSV_DIR / filename
    csv_text = _to_3block_csv(Path(sheet["image_path"]).name, sheet["sheet_data"])
    target.write_text(csv_text, encoding="utf-8")
    return target


@app.post("/sheet/{sheet_id}/validate")
async def sheet_validate(
    sheet_id: int,
    request: Request,
    operador: str = Form(...),
    data: str = Form(...),
    n_operador: str = Form(...),
) -> RedirectResponse:
    # Round 34 — mobile cannot validate (server-side enforcement)
    if _is_mobile_request(request):
        raise HTTPException(403, "Validação só pode ser feita em desktop")
    # R122 — o operador vem do cabeçalho já cruzado contra o
    # ListaColaboradores (_apply_operador_snap). Já não há dropdown de
    # 3 nomes hardcoded; só rejeitamos se vier mesmo vazio.
    if not operador.strip():
        raise HTTPException(400, "operador em falta — corrige o cabeçalho antes de validar")
    # Round 50 — re-validate bloqueada; folha validada é final.
    sheet_pre = db.get_sheet(sheet_id)
    if sheet_pre is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    if sheet_pre.get("status") == "validated":
        raise HTTPException(409, "Folha já validada — não é possível re-validar")

    # R94 — confirm date + n_operador before locking validation.
    data_iso = data.strip()
    if not _ISO_DATE_RE.match(data_iso):
        raise HTTPException(400, f"data must be YYYY-MM-DD, got {data!r}")
    n_op_clean = n_operador.strip()
    if not n_op_clean.isdigit() or len(n_op_clean) > 5:
        raise HTTPException(400, f"n_operador must be 1-5 digits, got {n_operador!r}")
    # Convert ISO → DD-MM-YYYY for storage compatibility with existing format
    data_pt = f"{data_iso[8:10]}-{data_iso[5:7]}-{data_iso[0:4]}"
    cur_header = (sheet_pre.get("sheet_data") or {}).get("header") or {}
    # Apply edits before validation lock — uses standard apply_edit path so
    # production_rows + cross-check stay in sync.
    if (cur_header.get("data") or "").strip() != data_pt:
        try:
            db.apply_edit(sheet_id, "header.data", data_pt)
        except Exception:  # noqa: BLE001
            pass
    if (cur_header.get("n_operador") or "").strip() != n_op_clean:
        try:
            db.apply_edit(sheet_id, "header.n_operador", n_op_clean)
        except Exception:  # noqa: BLE001
            pass

    # R126 — edição de cesta_n foi removida do sheet.html (validate desktop).
    # A cesta entra exclusivamente pelo fluxo mobile (capture.html → /mobile/qtds-batch).

    db.validate_sheet(sheet_id, operador)
    # R117 — kernel event: folha validada (lock confirmado pelo operador).
    try:
        kernel.emit_event("sheet_validated", {"sheet_id": sheet_id, "operador": operador})
    except Exception:  # noqa: BLE001
        pass
    # R113 — folha acabada de validar entra no cálculo de consumption.
    # Invalida cache para o /of-lookup seguinte ver os números actualizados.
    # R115 — também invalida o agregado /obras (qtd produzida muda).
    try:
        from app.pipeline.of_consumption import invalidate_cache
        invalidate_cache()
        from app.pipeline.obras_status import invalidate_cache as obras_inv
        obras_inv()
    except Exception:  # noqa: BLE001
        pass
    # Closed loop: drop CSV in the factory CSV dir so the next run of
    # ``kanban_csv2excel_novo_layout.py`` picks it up. Failure is silent —
    # the user can still pull the CSV via the /sheet/{id}/csv endpoint.
    try:
        _deposit_csv_to_factory(sheet_id)
    except Exception as e:  # noqa: BLE001
        traceback.print_exc()
        print(f"[factory deposit] sheet {sheet_id}: {e}", file=sys.stderr)
    # Update cross-check status (sheet just got validated)
    try:
        _run_and_store_cross_check(sheet_id)
    except Exception as cc_err:  # noqa: BLE001
        print(f"[cross-check] sheet {sheet_id} validate: {cc_err}", file=sys.stderr)
    # Learning loop — every 50 validated sheets, mine corrections + gold
    # into learnings. Runs in a background thread; failure is silent.
    try:
        from app.learning.scheduler import maybe_trigger_learning
        maybe_trigger_learning()
    except Exception as le:  # noqa: BLE001
        print(f"[learning] sheet {sheet_id} trigger: {le}", file=sys.stderr)
    return RedirectResponse("/queue", status_code=303)


# --- Round 30 + Round 32: cross-check admin endpoints ---
# Note: Round 32 removed the /dashboard/cross-check page from the UI —
# cross-check runs invisibly inline in /upload (see _run_and_store_cross_check).
# Admin APIs below remain for debug + ref-management.


# ============================================================================
# Mobile flow: capture-only + QTD edit (Round 31)
# Mobile users can ONLY take photos and edit qty/colunas_produzidas. Every
# other interaction with the system happens on desktop.
# ============================================================================

@app.get("/mobile/qtds")
def mobile_qtds(ids: str) -> JSONResponse:
    """Return minimal data needed for the mobile QTD-confirm screen.

    R114 — Para folhas com `cesta_n` no template (Expedição), inclui
    também o número da cesta e `row_fields_extra=['cesta_n']` para a UI
    mostrar a coluna apropriada.
    """
    try:
        sheet_ids = [int(s) for s in ids.split(",") if s.strip()]
    except ValueError:
        raise HTTPException(400, "ids must be comma-separated integers")
    from app.templates_registry import get_template
    out = []
    for sid in sheet_ids:
        sheet = db.get_sheet(sid)
        if sheet is None:
            continue
        sd = sheet.get("sheet_data") or {}
        h = sd.get("header", {}) or {}
        f = sd.get("footer", {}) or {}
        rows = sd.get("rows", []) or []
        # R114 — descobrir o template + campos extra editáveis em mobile
        template_name = db.get_sheet_template_name(sheet)
        try:
            tpl = get_template(template_name)
            extra_fields = [
                fname for fname in ("cesta_n",) if fname in tpl.row_fields
            ]
        except Exception:  # noqa: BLE001
            extra_fields = []
        out.append({
            "sheet_id": sid,
            "status": sheet["status"],
            "operador": h.get("operador") or "",
            "data": h.get("data") or "",
            "template_name": template_name,
            "row_fields_extra": extra_fields,
            "rows": [
                {
                    "row_index": i,
                    "modelo": r.get("modelo", ""),
                    "cliente": r.get("cliente", ""),
                    "of": r.get("of", ""),
                    "qtd": r.get("qtd", ""),
                    # R114 — campo extra (vazio se template não usa cesta)
                    "cesta_n": r.get("cesta_n", ""),
                }
                for i, r in enumerate(rows)
            ],
            "colunas_produzidas": f.get("colunas_produzidas", ""),
            "horas_trabalhadas": f.get("horas_trabalhadas", ""),
        })
    return JSONResponse({"sheets": out})


@app.post("/mobile/qtds-batch")
async def mobile_qtds_batch(request: Request) -> JSONResponse:
    """Apply a batch of qty edits at once. Body is JSON:
        { "edits": [ {sheet_id, field_path, value}, ... ] }

    Restricts field_path to qty/cesta_n/colunas_produzidas only —
    anything else is rejected. Re-cross-checks each affected sheet.

    R123 — já NÃO valida folhas: o auto-validate mobile do R114/R122 foi
    revertido. Validar é um acto humano deliberado, feito no desktop.
    """
    body = await request.json()
    edits = body.get("edits", [])
    if not isinstance(edits, list):
        raise HTTPException(400, "edits must be a list")

    # Whitelist: qty + cesta_n (R114) + footer counters
    allowed_suffixes = (
        ".qtd",
        ".cesta_n",  # R114 — Expedição
        "footer.colunas_produzidas",
        "footer.horas_trabalhadas",
    )
    applied = 0
    affected_sheets: set[int] = set()
    errors: list[dict] = []

    for e in edits:
        try:
            sid = int(e.get("sheet_id"))
            field_path = str(e.get("field_path") or "").strip()
            value = str(e.get("value") if e.get("value") is not None else "")
        except (TypeError, ValueError):
            errors.append({"edit": e, "error": "bad shape"})
            continue
        if not any(field_path == s or field_path.endswith(s) for s in allowed_suffixes):
            errors.append({"edit": e, "error": f"field {field_path} not allowed on mobile"})
            continue
        try:
            db.apply_edit(sid, field_path, value)
            applied += 1
            affected_sheets.add(sid)
        except ValueError as ex:
            errors.append({"edit": e, "error": str(ex)})

    # Re-cross-check each sheet that got edited
    for sid in affected_sheets:
        try:
            _run_and_store_cross_check(sid)
        except Exception:  # noqa: BLE001
            traceback.print_exc()

    return JSONResponse({
        "ok": True,
        "applied": applied,
        "errors": errors,
        "sheets_updated": list(affected_sheets),
    })


# ============================================================================
# Mobile guard — desktop-only pages redirect mobile clients to /capture
# ============================================================================

_DESKTOP_ONLY_PREFIXES = (
    "/sheet/",       # editing/review page (full table edit)
    "/kanbans",      # kanban viewer (operator selector + nav)
    "/dashboard",    # all dashboards (production, workers, jobs, etc.)
    "/queue",        # full sheets list with edit/validate
    "/pair",         # QR code page (only useful on desktop showing the QR)
    "/export",       # bulk Excel export (covers /export and /export/cpis)
    "/excel",        # R66 — continuous data page (desktop-only)
)


@app.middleware("http")
async def _mobile_guard(request: Request, call_next):
    """Redirect mobile clients away from desktop-only pages.

    Allowed on mobile: /capture, /mobile/*, /upload, /image/*, static, admin
    APIs (used by JS only — not human-navigated). Everything else → /capture.

    HTMX/AJAX requests pass through (they need to keep working when mobile JS
    triggers them on /capture itself, e.g. /upload).
    """
    path = request.url.path
    is_mobile = (
        any(p in (request.headers.get("user-agent") or "").lower()
            for p in ("mobile", "iphone", "android", "ipad", "ipod"))
    )
    # Only redirect human navigations — not API/HTMX hits
    is_html_request = (
        request.method == "GET"
        and "text/html" in (request.headers.get("accept") or "")
    )
    if is_mobile and is_html_request:
        for prefix in _DESKTOP_ONLY_PREFIXES:
            if path.startswith(prefix):
                return RedirectResponse("/capture", status_code=303)
    return await call_next(request)


@app.get("/admin/refs-status")
def admin_refs_status() -> JSONResponse:
    """Show current state of the SAP/plan refs (which Excel files loaded,
    when, n_lotes/n_ofs counts). Plus aggregate cross-check summary."""
    return JSONResponse({
        "refs": get_watcher().status(),
        "summary": load_summary(),
    })


@app.get("/admin/queue-status")
def admin_queue_status() -> JSONResponse:
    """R71 — health + depth of the background OCR queue. Used for debug
    and the /admin pages to show worker liveness."""
    return JSONResponse({
        "queue_size": ocr_queue.queue_size(),
        "worker_alive": ocr_queue.worker_alive(),
    })


@app.get("/admin/qwen-tools-test")
def qwen_tools_test() -> JSONResponse:
    """R120 — testa se o modelo Ollama actual invoca function calling.

    Faz um pedido mínimo com uma tool fake ``get_now``. Se a resposta tem
    ``tool_calls[]`` populado, o modelo suporta tools. Se não, sabemos que
    o fallback é a solução correcta.
    """
    from app.config import get_settings
    s = get_settings()
    payload = {
        "model": s.ollama_text_model,
        "messages": [
            {"role": "user", "content": "Qual a hora actual? Usa a tool get_now."}
        ],
        "tools": [{
            "type": "function",
            "function": {
                "name": "get_now",
                "description": "Devolve a data e hora actual.",
                "parameters": {"type": "object", "properties": {}, "required": []},
            },
        }],
        "stream": False,
        "think": False,
        "options": {"temperature": 0.1},
    }
    try:
        with httpx.Client(timeout=60.0) as c:
            r = c.post(f"{str(s.ollama_url).rstrip('/')}/api/chat", json=payload)
            r.raise_for_status()
            data = r.json()
        msg = data.get("message") or {}
        tool_calls = msg.get("tool_calls") or []
        return JSONResponse({
            "model": s.ollama_text_model,
            "supports_tools": len(tool_calls) > 0,
            "tool_calls_in_response": tool_calls,
            "content_preview": (msg.get("content") or "")[:200],
        })
    except Exception as e:  # noqa: BLE001
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/sheet/{sheet_id}/status")
def sheet_status(sheet_id: int) -> JSONResponse:
    """R71 — JSON polling endpoint for the mobile capture flow + desktop
    auto-refresh. Returns minimal fields so the JS can drive its state
    machine without re-rendering the full sheet view.
    """
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404)
    return JSONResponse({
        "id": sheet_id,
        "status": sheet.get("status"),
        "queue_size": ocr_queue.queue_size(),
        "error_message": sheet.get("error_message"),
    })


@app.get("/sheet/{sheet_id}/status-fragment", response_class=HTMLResponse)
def sheet_status_fragment(sheet_id: int) -> Response:
    """R71 — HTMX-friendly partial for the desktop /sheet/{id} pending
    banner. While status is 'pending', returns the banner so HTMX swaps
    it back in (preserving the every-2s trigger). When status transitions
    out of 'pending', returns an empty body with ``HX-Refresh: true``,
    causing the browser to refresh and render the now-extracted view.
    """
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404)
    if sheet.get("status") != "pending":
        return Response(
            content="",
            status_code=200,
            headers={"HX-Refresh": "true"},
        )
    queue_size = ocr_queue.queue_size()
    html = (
        '<div class="alert info" '
        f'hx-get="/sheet/{sheet_id}/status-fragment" '
        'hx-trigger="every 2s" hx-swap="outerHTML">'
        '<b>⏳ A processar OCR...</b><br>'
        '<span class="muted tiny">'
        f'{queue_size} folha(s) na fila — esta página actualiza-se sozinha'
        '</span></div>'
    )
    return HTMLResponse(html)


@app.post("/admin/reload-refs")
def admin_reload_refs() -> JSONResponse:
    """Force-reload SAP + plan_colunas Excel files (skip mtime check).
    Optionally re-cross-check ALL sheets in DB so updated refs propagate."""
    refs = get_watcher().force_reload()
    # R115 — refs novas invalidam o agregado /obras
    try:
        from app.pipeline.obras_status import invalidate_cache as obras_inv
        obras_inv()
    except Exception:  # noqa: BLE001
        pass
    revalidated = 0
    for s in db.list_sheets(limit=10000):
        if s["status"] in ("error", "pending"):
            continue
        try:
            _run_and_store_cross_check(s["id"])
            revalidated += 1
        except Exception:  # noqa: BLE001
            traceback.print_exc()
    return JSONResponse({
        "ok": True,
        "refs_loaded_at": refs.get("loaded_at"),
        "n_lotes": refs.get("stats", {}).get("n_lotes", 0),
        "n_ofs": refs.get("stats", {}).get("n_ofs", 0),
        "sheets_revalidated": revalidated,
    })


# ===================== R104 — página de refs SAP/plan =====================
# Upload de StockSAP.xlsx / plan_colunas_cpis.xlsx com acumulação histórica.

_REFS_FILENAMES = {"stocksap": "StockSAP.xlsx", "plan": "plan_colunas_cpis.xlsx"}

# Progresso da re-validação cross-check em background (1 corrida de cada vez).
_revalidation_state: dict = {
    "running": False, "done": 0, "total": 0,
    "started_at": None, "finished_at": None,
}
_revalidation_lock = threading.Lock()


def _revalidate_all_sheets_bg() -> None:
    """Re-cross-check every non-pending/-error sheet against the freshly
    loaded refs. Runs in a daemon thread so the upload response is instant."""
    try:
        sheets = [
            s for s in db.list_sheets(limit=10000)
            if s["status"] not in ("error", "pending")
        ]
        with _revalidation_lock:
            _revalidation_state.update(
                running=True, done=0, total=len(sheets), finished_at=None,
                started_at=dt.datetime.now().isoformat(timespec="seconds"),
            )
        for s in sheets:
            try:
                _run_and_store_cross_check(s["id"])
            except Exception:  # noqa: BLE001
                traceback.print_exc()
            with _revalidation_lock:
                _revalidation_state["done"] += 1
    finally:
        with _revalidation_lock:
            _revalidation_state["running"] = False
            _revalidation_state["finished_at"] = (
                dt.datetime.now().isoformat(timespec="seconds"))


def _validate_refs_xlsx(path: Path, kind: str) -> str | None:
    """Return ``None`` if the workbook looks like the expected refs file,
    else a human error message. Defensive — a bad upload must never replace
    a good live refs file."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return f"não consegui abrir o Excel ({e})"
    try:
        if kind == "plan":
            ws = (wb["plan_colunas_cpis"]
                  if "plan_colunas_cpis" in wb.sheetnames else wb.active)
            first = next(ws.iter_rows(values_only=True), None)
            hdrs = {str(h).strip().lower() for h in (first or ()) if h}
            if "of" not in hdrs:
                return "falta a coluna 'of' — não parece o plan_colunas_cpis"
        else:  # stocksap
            rows = (wb["Folha1"] if "Folha1" in wb.sheetnames
                    else wb.active).iter_rows(values_only=True)
            header = next(rows, None) or ()
            col0 = str(header[0] or "").strip().lower() if header else ""
            if "lote" not in col0:
                return "1ª coluna não é 'Lote' — não parece o StockSAP"
            data0 = next(rows, None)
            if data0 is None or data0[0] is None:
                return "sem linhas de lote — não parece o StockSAP"
    finally:
        wb.close()
    return None


def _fmt_mtime(ts: float | None) -> str:
    """Epoch float → readable local datetime (for the refs status card)."""
    if not ts:
        return "—"
    return dt.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M")


def _start_revalidation() -> bool:
    """Arranca a re-validação cross-check em background (1 corrida de cada
    vez). Devolve True se arrancou, False se já estava a correr."""
    with _revalidation_lock:
        if _revalidation_state["running"]:
            return False
        _revalidation_state.update(
            running=True, done=0, total=0, finished_at=None,
            started_at=dt.datetime.now().isoformat(timespec="seconds"),
        )
    threading.Thread(target=_revalidate_all_sheets_bg, daemon=True).start()
    return True


@app.get("/refs", response_class=HTMLResponse)
def refs_page(request: Request) -> Response:
    """Página para carregar StockSAP/plan e ver o estado das refs."""
    from app.cross_check import refs_uploads
    refs = get_watcher().get_refs()
    status = get_watcher().status()
    return templates.TemplateResponse(request, "refs.html", {
        "refs_status": status,
        "stats": refs.get("stats", {}),
        "uploads": refs_uploads.recent(),
        "revalidation": dict(_revalidation_state),
        "sap_file_date": _fmt_mtime(status.get("sap", {}).get("mtime")),
        "plan_file_date": _fmt_mtime(status.get("plan", {}).get("mtime")),
        "flash_ok": request.query_params.get("ok"),
        "flash_err": request.query_params.get("err"),
        "active_tab": "refs",
    })


@app.post("/refs/upload")
async def refs_upload(
    kind: str = Form(...),
    file: UploadFile = File(...),
) -> Response:
    """Recebe um StockSAP.xlsx / plan_colunas_cpis.xlsx, valida-o e substitui
    o ficheiro vivo. Recarrega as refs DIRETO do ficheiro (sem acumulação
    histórica). NÃO re-cross-checka folhas — isso é o botão 'Re-validar'."""
    from app.cross_check import refs_uploads
    if kind not in _REFS_FILENAMES:
        raise HTTPException(400, "kind inválido")
    if not file.filename:
        return RedirectResponse("/refs?err=sem+ficheiro", status_code=303)
    if Path(file.filename).suffix.lower() not in (".xlsx", ".xlsm"):
        return RedirectResponse(
            "/refs?err=o+ficheiro+tem+de+ser+.xlsx", status_code=303)

    # R118 — rede de segurança global: qualquer exceção (PermissionError no
    # mkdir, falha do watcher, etc.) é silenciosa hoje e dá página em branco
    # ao operador. Captura e devolve mensagem útil em ?err=...
    try:
        watcher = get_watcher()
        target = watcher.sap_path if kind == "stocksap" else watcher.plan_path
        target.parent.mkdir(parents=True, exist_ok=True)
        # Temp file keeps the .xlsx suffix — openpyxl validates by extension.
        tmp = target.with_name(f"{target.stem}.upload-tmp{target.suffix}")

        bytes_written = 0
        with tmp.open("wb") as f:
            while chunk := await file.read(1024 * 1024):
                bytes_written += len(chunk)
                if bytes_written > _MAX_UPLOAD_BYTES:
                    f.close()
                    tmp.unlink(missing_ok=True)
                    return RedirectResponse(
                        "/refs?err=ficheiro+demasiado+grande", status_code=303)
                f.write(chunk)

        err = _validate_refs_xlsx(tmp, kind)
        if err:
            tmp.unlink(missing_ok=True)
            return RedirectResponse(
                f"/refs?err=ficheiro+rejeitado:+{err}", status_code=303)

        # os.replace falha se o ficheiro vivo estiver aberto (ex.: Excel) — tenta
        # algumas vezes antes de desistir com um erro claro.
        replaced = False
        for _ in range(5):
            try:
                os.replace(tmp, target)
                replaced = True
                break
            except PermissionError:
                await asyncio.sleep(0.3)
        if not replaced:
            tmp.unlink(missing_ok=True)
            return RedirectResponse(
                "/refs?err=ficheiro+em+uso+-+fecha+o+Excel+e+tenta+outra+vez",
                status_code=303)
        refs = get_watcher().force_reload()  # recarrega direto do ficheiro
        # R115 — refs novas invalidam o agregado /obras
        try:
            from app.pipeline.obras_status import invalidate_cache as obras_inv
            obras_inv()
        except Exception:  # noqa: BLE001
            pass
        stats = refs.get("stats", {})
        n_rows = stats.get("n_plan_rows" if kind == "plan" else "n_lotes", 0)
        # R118 — record() é best-effort; nunca falhar o ?ok=
        try:
            refs_uploads.record(kind, target.name, n_rows)
        except Exception:  # noqa: BLE001
            traceback.print_exc()
        return RedirectResponse(f"/refs?ok={kind}+atualizado", status_code=303)
    except Exception as e:  # noqa: BLE001
        # R118 — captura qualquer exceção não tratada e devolve mensagem
        # útil ao operador (antes: silêncio / página em branco).
        traceback.print_exc()
        msg = str(e)[:80].replace("\n", " ").replace("&", "").replace("?", "")
        return RedirectResponse(
            f"/refs?err=erro+inesperado:+{msg}", status_code=303)


@app.post("/refs/revalidate")
def refs_revalidate() -> Response:
    """Botão 'Re-validar folhas' — re-corre o cross-check de TODAS as folhas
    (extracted + validated) contra as refs atuais, em background."""
    if _start_revalidation():
        return RedirectResponse(
            "/refs?ok=re-validacao+iniciada", status_code=303)
    return RedirectResponse(
        "/refs?err=re-validacao+ja+esta+a+correr", status_code=303)


@app.get("/refs/revalidation-status", response_class=HTMLResponse)
def refs_revalidation_status(request: Request) -> Response:
    """Fragmento HTMX com o progresso da re-validação cross-check."""
    return templates.TemplateResponse(request, "_refs_revalidation.html", {
        "revalidation": dict(_revalidation_state),
    })


@app.get("/admin/to-analisar")
def admin_to_analisar(limit: int | None = None) -> JSONResponse:
    """Inbox of cells flagged ANALISAR — for the supervisor's review queue."""
    return JSONResponse(load_to_analisar(limit=limit))


@app.post("/sheet/{sheet_id}/delete")
def sheet_delete(sheet_id: int, request: Request) -> JSONResponse:
    """Round 34 — hard delete sheet + cascade. Desktop only.

    Cascade:
    - sheets row, edits, production_rows
    - image file in data/images/
    - cross-check JSON in C:\\kanban\\nifruka\\03_Cross_Check\\
    - factory CSV in C:\\kanban\\nifruka\\02_Dados_Extraidos\\csv\\
    """
    if _is_mobile_request(request):
        raise HTTPException(403, "Apagar só pode ser feito em desktop")
    try:
        result = db.delete_sheet(sheet_id)
    except ValueError as e:
        raise HTTPException(404, str(e))
    return JSONResponse({"ok": True, "removed": result})


@app.post("/sheet/{sheet_id}/reprocess")
def sheet_reprocess(sheet_id: int) -> RedirectResponse:
    """Round 59/71 — Re-run OCR on a sheet that previously errored.

    Reuses the original uploaded image; no need to re-take photo. Triggered
    by the "↻ Re-processar OCR" button on the error banner in /sheet/{id}.

    R71: flips status back to 'pending' and enqueues to the background
    worker (instead of running OCR synchronously here). /sheet/{id} then
    shows the "A processar..." banner until the worker finishes.
    """
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    img_path = _DATA_DIR / sheet["image_path"]
    if not img_path.exists():
        raise HTTPException(404, "image file missing")
    db.update_status(sheet_id, "pending")
    db.clear_error(sheet_id)
    ocr_queue.enqueue(sheet_id)
    return RedirectResponse(f"/sheet/{sheet_id}", status_code=303)


@app.post("/sheet/{sheet_id}/rotate")
def sheet_rotate(sheet_id: int, request: Request) -> JSONResponse:
    """Round 34c — rotate the served image 90° CW per click. Desktop only.

    Phones don't write EXIF orientation, so the auto-rotation guess can be
    wrong (sometimes upside-down). User clicks "rodar" to cycle through
    0/90/180/270 until kanban is readable. Stored per-sheet.
    """
    if _is_mobile_request(request):
        raise HTTPException(403, "Rodar só pode ser feito em desktop")
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    current = int(sheet.get("image_rotation") or 0)
    new = db.set_image_rotation(sheet_id, current + 90)
    return JSONResponse({"ok": True, "rotation": new})


# R112 — Wizard "Corrigir via OF" -----------------------------------------

@app.get("/sheet/{sheet_id}/of-lookup")
def sheet_of_lookup(
    sheet_id: int,
    of: str = "",
    q: str = "",
    include_done: int = 0,
) -> JSONResponse:
    """R112 — devolve entries do plan_colunas para um OF/OV/modelo.

    R113 — entries são ordenadas por "faltam menos primeiro" e entries
    já fechadas (remaining ≤ 0) são filtradas por defeito.
    `include_done=1` para mostrar todas (recuperação de folhas antigas).

    R128 — auto-detect multi-modo. O parâmetro `q` aceita OF (6 dígitos),
    OV (≥6 dígitos) ou prefixo de modelo (alfanumérico). Ordem de
    tentativa: OF → OV → modelo prefix. `of=` mantido para back-compat
    com clients antigos.
    """
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    query_raw = (q or of or "").strip()
    if not query_raw:
        return JSONResponse({
            "found": False, "mode": "none", "q": "", "of": "", "entries": [],
        })

    from app.pipeline.scoring_engine import normalize_of
    from app.pipeline.of_consumption import sort_entries_by_remaining

    refs = get_watcher().get_refs() or {}
    of_to_entries = refs.get("of_to_entries") or {}
    plan_by_ov = refs.get("plan_by_ov") or {}
    plan_by_modelo_ft = refs.get("plan_by_modelo_ft") or {}

    LIMIT = 50
    mode = "none"
    matched_of = ""
    pooled: list[dict] = []
    n_total_pre_filter = 0
    truncated = False

    q_upper = query_raw.upper()
    is_numeric = query_raw.isdigit()

    # Tier 1 — OF (numérico)
    if is_numeric:
        of_norm = normalize_of(query_raw)
        entries = of_to_entries.get(of_norm) or []
        if entries:
            mode = "of"
            matched_of = of_norm
            n_total_pre_filter = len(entries)
            pooled = [
                {**e, "_of": of_norm, "_orig_idx": i}
                for i, e in enumerate(entries)
            ]

    # Tier 2 — OV (numérico, exact match)
    if mode == "none" and is_numeric:
        ov_entries = plan_by_ov.get(query_raw) or []
        if ov_entries:
            mode = "ov"
            n_total_pre_filter = len(ov_entries)
            # plan_by_ov entries já têm "_of" anotado (ver
            # ref_watcher._derive_plan_indexes). _orig_idx é por OF para
            # apply-of-entry; rebuild aqui contra of_to_entries.
            for e in ov_entries:
                of_of_entry = str(e.get("_of") or "")
                source = of_to_entries.get(of_of_entry) or []
                orig_idx = next(
                    (i for i, se in enumerate(source) if se is e
                     or (se.get("ov") == e.get("ov")
                         and se.get("designacao") == e.get("designacao"))),
                    -1,
                )
                pooled.append({**e, "_orig_idx": orig_idx})

    # Tier 3 — modelo prefix
    if mode == "none":
        matching_keys = [k for k in plan_by_modelo_ft.keys()
                         if k.startswith(q_upper)]
        for k in matching_keys:
            entries = plan_by_modelo_ft.get(k) or []
            for e in entries:
                of_of_entry = str(e.get("_of") or "")
                source = of_to_entries.get(of_of_entry) or []
                orig_idx = next(
                    (i for i, se in enumerate(source) if se is e
                     or (se.get("ov") == e.get("ov")
                         and se.get("designacao") == e.get("designacao"))),
                    -1,
                )
                pooled.append({**e, "_orig_idx": orig_idx})
        if pooled:
            mode = "modelo"
            n_total_pre_filter = len(pooled)

    if mode == "none":
        return JSONResponse({
            "found": False, "mode": "none", "q": query_raw, "of": "",
            "entries": [], "n_entries": 0, "n_total": 0,
        })

    sorted_entries = sort_entries_by_remaining(
        pooled, include_done=bool(include_done),
    )

    if len(sorted_entries) > LIMIT:
        sorted_entries = sorted_entries[:LIMIT]
        truncated = True

    out_entries = []
    for i, e in enumerate(sorted_entries):
        out_entries.append({
            "idx": i,
            "orig_idx": e.get("_orig_idx"),  # R116 — usar este no apply
            "of": str(e.get("_of") or ""),    # R128 — OF por entry (modo modelo/OV)
            "cliente": e.get("cliente", ""),
            "ov": str(e.get("ov", "")),
            "modelo": e.get("designacao", ""),
            "comp_mm": e.get("comp"),
            "lbase": e.get("lbase"),
            "ltopo": e.get("ltopo"),
            "esp": e.get("esp"),
            "material": e.get("material", ""),
            "fechado": bool(e.get("fechado")),
            "remaining": e.get("_remaining"),
            "quanttrp": e.get("_quanttrp"),
            "done": e.get("_done", False),
        })
    return JSONResponse({
        "found": True,
        "mode": mode,
        "q": query_raw,
        "of": matched_of,    # back-compat: vazio nos modos ov/modelo
        "entries": out_entries,
        "n_entries": len(out_entries),
        "n_total": n_total_pre_filter,
        "truncated": truncated,
    })


@app.post("/sheet/{sheet_id}/apply-of-entry")
async def sheet_apply_of_entry(sheet_id: int, request: Request) -> JSONResponse:
    """R112 — aplica os campos de uma entry do plan a uma linha do kanban.

    Body: {row_index, of, entry_idx}. Escreve cliente, modelo, OV,
    comp_mm, lbase, ltopo, esp via apply_edit(source='system'). Não
    toca em qtd, lote, pri, coni, larg_mm (não estão no plan ou são
    por bobine). Re-corre cross-check para refrescar cores.
    """
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    if sheet.get("status") == "validated":
        raise HTTPException(409, "Folha já validada — edits bloqueados")
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        raise HTTPException(400, "Body JSON inválido")
    try:
        row_index = int(body.get("row_index", -1))
        entry_idx = int(body.get("entry_idx", -1))
    except (ValueError, TypeError):
        raise HTTPException(400, "row_index e entry_idx têm de ser inteiros")
    of_raw = str(body.get("of", "")).strip()
    if row_index < 0 or entry_idx < 0 or not of_raw:
        raise HTTPException(400, "row_index, of, entry_idx obrigatórios")

    from app.pipeline.scoring_engine import normalize_of
    of_norm = normalize_of(of_raw)
    refs = get_watcher().get_refs() or {}
    entries = (refs.get("of_to_entries") or {}).get(of_norm) or []
    if not entries or entry_idx >= len(entries):
        raise HTTPException(404, "OF ou entry não encontrados no plan")

    e = entries[entry_idx]
    fields_to_set = {
        "of": of_norm,
        "cliente": e.get("cliente", ""),
        "ov": str(e.get("ov", "")),
        "modelo": e.get("designacao", ""),
        "comp_mm": e.get("comp"),
        "lbase": e.get("lbase"),
        "ltopo": e.get("ltopo"),
        "esp": e.get("esp"),
    }
    applied = []
    skipped = []
    for field, value in fields_to_set.items():
        if value is None or value == "":
            skipped.append(field)
            continue
        path = f"rows[{row_index}].{field}"
        try:
            db.apply_edit(sheet_id, path, str(value), source="system")
            applied.append({"field": field, "value": str(value)})
        except ValueError:
            skipped.append(field)
        except Exception:  # noqa: BLE001
            skipped.append(field)
    try:
        _run_and_store_cross_check(sheet_id)
    except Exception:  # noqa: BLE001
        pass
    # R113 — após aplicar, refresca a cache de consumption (a próxima
    # chamada a /of-lookup vai recomputar baseado neste novo estado).
    try:
        from app.pipeline.of_consumption import invalidate_cache
        invalidate_cache()
    except Exception:  # noqa: BLE001
        pass
    return JSONResponse({
        "ok": True,
        "n_applied": len(applied),
        "applied": applied,
        "skipped": skipped,
        "of_used": of_norm,
    })


@app.post("/sheet/{sheet_id}/recrop")
def sheet_recrop(sheet_id: int) -> JSONResponse:
    """R111 — tenta correr auto-crop outra vez nesta folha.

    Útil quando a primeira tentativa falhou (paper não detectado) e o
    supervisor quer pedir nova tentativa via UI. Idempotente: se já
    existe cropped, sobrepõe; se a detecção falhar agora, devolve
    {ok: False, has_cropped: False} para a UI mostrar mensagem.
    """
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    img_path = _DATA_DIR / sheet["image_path"]
    if not img_path.exists():
        return JSONResponse(
            {"ok": False, "has_cropped": False,
             "error": "Imagem original não encontrada."},
            status_code=404,
        )
    from .image_crop import auto_crop, has_cropped
    result = auto_crop(img_path)
    success = result is not None
    return JSONResponse({
        "ok": success,
        "has_cropped": has_cropped(img_path),
        "message": ("Kanban recortado com sucesso." if success
                    else "Não consegui detectar o kanban na foto. "
                         "Talvez a folha esteja com pouco contraste, "
                         "muito inclinada ou cortada."),
    })


@app.get("/sheet/{sheet_id}/csv")
def sheet_csv(sheet_id: int) -> Response:
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    data = sheet.get("sheet_data") or {}
    csv_text = _to_3block_csv(Path(sheet["image_path"]).name, data)
    return Response(
        content=csv_text,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="sheet_{sheet_id}.csv"'
        },
    )


@app.get("/image/{sheet_id}/original")
def sheet_image_original(sheet_id: int) -> FileResponse:
    """Serve the raw uploaded photo (no auto-crop applied).

    Useful for review when the auto-crop misfires or the user wants to
    see the full photo with background.
    """
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    img_path = _DATA_DIR / sheet["image_path"]
    if not img_path.exists():
        raise HTTPException(404, "image file missing")
    return _serve_image_with_rotation(sheet, img_path)


@app.get("/image/{sheet_id}")
def sheet_image(sheet_id: int) -> FileResponse:
    """Serve the kanban photo. Round 46: prefer auto-cropped version (paper
    only, perspective-corrected) when available; falls back to raw photo.

    Phones take portrait photos of landscape kanbans. Rotating server-side
    (with a disk cache next to the original) avoids CSS transform hacks
    that overflow the layout. EXIF orientation also applied via PIL.
    """
    sheet = db.get_sheet(sheet_id)
    if sheet is None:
        raise HTTPException(404, f"sheet {sheet_id} not found")
    img_path = _DATA_DIR / sheet["image_path"]
    if not img_path.exists():
        raise HTTPException(404, "image file missing")
    # Round 46: prefer auto-cropped version (paper only, no background).
    # Cropped image is ALREADY rotated/oriented correctly by the warp,
    # so we skip _serve_image_with_rotation for it (no extra rotation pass).
    from .image_crop import cropped_path_for
    cropped = cropped_path_for(img_path)
    if cropped.exists():
        # Honour user rotation override on top of the cropped image
        rot_override = int(sheet.get("image_rotation") or 0) % 360
        if rot_override == 0:
            return FileResponse(cropped, media_type="image/jpeg")
        # Apply rotation override to cropped (cache result)
        rot_cache = cropped.with_name(f"{cropped.stem}_r{rot_override}.jpg")
        src_mtime = cropped.stat().st_mtime
        if not rot_cache.exists() or rot_cache.stat().st_mtime < src_mtime:
            from PIL import Image
            with Image.open(cropped) as im:
                if im.mode != "RGB":
                    im = im.convert("RGB")
                im.rotate(-rot_override, expand=True).save(
                    rot_cache, "JPEG", quality=90, optimize=True)
            os.utime(rot_cache, (src_mtime, src_mtime))
        return FileResponse(rot_cache, media_type="image/jpeg")
    # Fallback: raw photo with auto-rotation pass (legacy path).
    return _serve_image_with_rotation(sheet, img_path)


def _serve_image_with_rotation(sheet: dict, img_path: Path) -> FileResponse:
    """Serve raw photo with EXIF transpose + portrait→landscape rotation +
    user rotation override. Used when no auto-cropped version exists."""
    # Cache key includes rotation override (per-sheet, stored in DB).
    # Phones don't write EXIF orientation, so we guess: portrait → rotate
    # 90° CCW. If wrong (upside-down), user clicks "rodar" → override
    # cycles through 90/180/270 and re-renders.
    rot_override = int(sheet.get("image_rotation") or 0) % 360
    suffix = f"_r{rot_override}" if rot_override else "_landscape"
    cache_path = img_path.with_name(img_path.stem + suffix + ".jpg")
    src_mtime = img_path.stat().st_mtime
    if not cache_path.exists() or cache_path.stat().st_mtime < src_mtime:
        from PIL import Image, ImageOps
        with Image.open(img_path) as im:
            im = ImageOps.exif_transpose(im)
            # Auto step: portrait → CCW 90° (most phones in this fleet
            # produce kanban with top-edge towards the LEFT of the sensor)
            if im.height > im.width:
                im = im.rotate(90, expand=True)
            # User override (clockwise quarter-turns)
            if rot_override:
                im = im.rotate(-rot_override, expand=True)
            if im.mode != "RGB":
                im = im.convert("RGB")
            im.save(cache_path, "JPEG", quality=85, optimize=True)
        os.utime(cache_path, (src_mtime, src_mtime))
    return FileResponse(cache_path, media_type="image/jpeg")


@app.get("/kanbans", response_class=HTMLResponse)
def kanban_viewer(
    request: Request,
    operador: str | None = None,
    data: str | None = None,
    setor: str | None = None,
    of: str | None = None,
    status: str | None = None,
    sheet_id: int | None = None,
) -> Response:
    """Desktop kanban viewer with multi-filter (operador + data + setor + of + status).

    Round 34/36: filters combinable via URL params. ``data`` is YYYY-MM-DD.
    ``of`` matches sheets that have at least one row with that OF.
    ``status`` accepts 'extracted' (não validadas) or 'validated'; empty = both.
    Empty filters = all matching sheets.
    """
    operadores = db.list_distinct_operadores()
    setores = db.list_distinct_setores()
    current_of = (of or "").strip() or None
    current_status = status if status in ("extracted", "validated") else None
    statuses = (current_status,) if current_status else ("extracted", "validated")

    if not operadores:
        return templates.TemplateResponse(
            request, "kanban_viewer.html",
            {
                "operadores": [],
                "setores": [],
                "current_operador": None,
                "current_data": data,
                "current_setor": setor,
                "current_of": current_of,
                "current_status": current_status,
                "sheets": [],
                "sheet": None,
                "neighbors": {"position": 0, "total": 0, "prev_id": None, "next_id": None},
                "header": {},
                "rows": [],
                "footer": {},
                "cells_by_path": {},
                "cc_status_by_path": {},
                "cc_ref_by_path": {},
                "cc_obra_concluida_by_path": {},
                "valid_operadores": OPERADORES,
                **_template_ctx_for_sheet(None),  # bobine_formato defaults
            },
        )

    # Resolve operador URL param via case+accent-insensitive match
    current_operador = None
    if operador:
        target_norm = db._normalize_operador(operador)
        for op in operadores:
            if db._normalize_operador(op) == target_norm:
                current_operador = op
                break

    # Apply multi-filter
    sheets = db.list_sheets_filtered(
        operador=current_operador,
        data_iso=data if data and _ISO_DATE_RE.match(data) else None,
        setor=setor,
        of=current_of,
        statuses=statuses,
    )
    if not sheets:
        sheet = None
        neighbors = {"position": 0, "total": 0, "prev_id": None, "next_id": None}
    else:
        # Pick sheet by ID if provided, else first in filtered list
        target_id = sheet_id if sheet_id and any(s["id"] == sheet_id for s in sheets) else sheets[0]["id"]
        sheet = db.get_sheet(target_id)
        # Round 34: neighbors computed within the filtered set so prev/next
        # respects the user's multi-filter (operador + data + setor).
        ids = [s["id"] for s in sheets]
        try:
            idx = ids.index(target_id)
            neighbors = {
                "position": idx + 1,
                "total": len(ids),
                "prev_id": ids[idx - 1] if idx > 0 else None,
                "next_id": ids[idx + 1] if idx + 1 < len(ids) else None,
            }
        except ValueError:
            neighbors = {"position": 0, "total": len(ids), "prev_id": None, "next_id": None}

    cells_by_path = {}
    if sheet and sheet.get("dq_audit"):
        cells_by_path = sheet["dq_audit"].get("cells", {})

    header = (sheet.get("sheet_data") or {}).get("header", {}) if sheet else {}
    rows = (sheet.get("sheet_data") or {}).get("rows", []) if sheet else []
    footer = (sheet.get("sheet_data") or {}).get("footer", {}) if sheet else {}

    # Round 33: cross-check colors per cell
    (cc_status_by_path, cc_ref_by_path, cc_suspended_by_path,
     cc_snapped_by_path, cc_obra_concluida_by_path) = ({}, {}, {}, {}, {})
    if sheet:
        (cc_status_by_path, cc_ref_by_path, cc_suspended_by_path,
         cc_snapped_by_path, cc_obra_concluida_by_path) = (
            _build_cc_maps(sheet["id"])
        )

    # R94 — ISO date pre-fill for validation form (same as sheet_page)
    data_iso_for_validate = db._normalize_data_pt_to_iso(header.get("data")) if header else None

    # R111 — has_cropped flag (mesma lógica do sheet_page)
    from .image_crop import has_cropped as _has_cropped
    sheet_has_cropped = False
    if sheet and sheet.get("image_path"):
        sheet_has_cropped = _has_cropped(_DATA_DIR / sheet["image_path"])

    return templates.TemplateResponse(
        request, "kanban_viewer.html",
        {
            "operadores": operadores,
            "setores": setores,
            "current_operador": current_operador,
            "current_data": data,
            "current_setor": setor,
            "current_of": current_of,
            "current_status": current_status,
            "sheets": sheets,
            "sheet": sheet,
            "neighbors": neighbors,
            "header": header,
            "rows": rows,
            "footer": footer,
            "cells_by_path": cells_by_path,
            "cc_status_by_path": cc_status_by_path,
            "cc_ref_by_path": cc_ref_by_path,
            "cc_suspended_by_path": cc_suspended_by_path,
            "cc_snapped_by_path": cc_snapped_by_path,
            "cc_obra_concluida_by_path": cc_obra_concluida_by_path,
            "valid_operadores": OPERADORES,
            "data_iso_for_validate": data_iso_for_validate,
            "has_cropped": sheet_has_cropped,
            **_template_ctx_for_sheet(sheet),  # per-current-sheet template
        },
    )


@app.get("/queue", response_class=HTMLResponse)
def queue_page(
    request: Request,
    status: str | None = None,
    of: str | None = None,
    operador: str | None = None,
    data: str | None = None,
    captured: str | None = None,
    setor: str | None = None,
) -> Response:
    """Round 36 — OF filter; R81 — operador/data/setor filters (combinable).
    R128 — captured (data de captura, distinta de header.data)."""
    of_filter = (of or "").strip() or None
    operador_filter = (operador or "").strip() or None
    data_filter = (data or "").strip() or None
    captured_filter = (captured or "").strip() or None
    setor_filter = (setor or "").strip() or None

    use_filtered = any([of_filter, operador_filter, data_filter, captured_filter, setor_filter])
    if use_filtered:
        statuses_all = ("pending", "extracted", "validated", "error")
        statuses = (status,) if status and status != "all" else statuses_all
        sheets = db.list_sheets_filtered(
            operador=operador_filter,
            data_iso=data_filter,
            captured_iso=captured_filter,
            setor=setor_filter,
            of=of_filter,
            statuses=statuses,
        )
        # list_sheets_filtered returns oldest first; flip to newest first
        sheets = sorted(sheets, key=lambda s: s.get("captured_at") or "", reverse=True)
    else:
        sheets = db.list_sheets(status=status)

    return templates.TemplateResponse(
        request,
        "queue.html",
        {
            "sheets": sheets,
            "status_filter": status or "all",
            "of_filter": of_filter,
            "operador_filter": operador_filter,
            "data_filter": data_filter,
            "captured_filter": captured_filter,
            "setor_filter": setor_filter,
            "operadores": db.list_distinct_operadores(),
            "setores": db.list_distinct_setores(),
        },
    )


_ISO_DATE_RE = __import__("re").compile(r"^\d{4}-\d{2}-\d{2}$")


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard_today(request: Request, date: str | None = None) -> Response:
    """Default dashboard view: Today (or any date via ?date=YYYY-MM-DD).

    Bad date strings are silently dropped (treated as no-filter) rather
    than passed to SQL — prevents user-visible errors from typos.
    """
    if date and not _ISO_DATE_RE.match(date):
        date = None
    summary = kpis.today_summary(reference_date=date)
    # Timeline window centers on the filter date (or today if no filter)
    daily = kpis.daily_trend(days=7, end_date=date)
    daily_op = kpis.daily_per_operador(days=7, end_date=date)
    return templates.TemplateResponse(
        request,
        "dashboard_today.html",
        {
            "summary": summary,
            "daily": daily,
            "daily_op": daily_op,
            "active_tab": "today",
            "date_filter": date,
        },
    )


@app.get("/export")
def export_excel(
    date_from: str | None = None,
    date_to: str | None = None,
    operador: str | None = None,
    sector: str | None = None,
) -> Response:
    """Round 29 Phase D — Excel multi-sheet bulk export.

    Query params (R69):
    - ``date_from``, ``date_to``: ISO YYYY-MM-DD inclusive range. Both
      omitted = "sempre" (no date filter).
    - ``operador``: optional filter (case-insensitive)
    - ``sector``: optional filter against one of ``PRODUCTION_SECTORS``

    Returns .xlsx with Resumo sheet + 1 sheet per day with sub-tables per
    operador. See export.py for structure.
    """
    df = (date_from or "").strip() or None
    dt_ = (date_to or "").strip() or None
    sec = (sector or "").strip() or None
    # Date validation: both or neither (XOR rejected for UX clarity)
    if bool(df) != bool(dt_):
        raise HTTPException(400, "provide both date_from and date_to, or neither (= sempre)")
    if df and dt_:
        if not _ISO_DATE_RE.match(df) or not _ISO_DATE_RE.match(dt_):
            raise HTTPException(400, "date_from and date_to must be YYYY-MM-DD")
        if dt_ < df:
            raise HTTPException(400, "date_to must be >= date_from")
    if sec and sec not in PRODUCTION_SECTORS:
        raise HTTPException(400, f"sector must be one of {list(PRODUCTION_SECTORS)}")
    xlsx_bytes = export.export_excel(df, dt_, operador, sec)
    filename = export.filename_for(df, dt_, operador, sec)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/excel", response_class=HTMLResponse)
def excel_page(
    request: Request,
    operador: str | None = None,
    of_filter: str | None = Query(None, alias="of"),
    limit: int = Query(500, ge=1, le=5000),
) -> Response:
    """Continuous Excel-style view of every registered production row.

    Mirrors the 17-column CPIS schema (the same shape as the
    `MigracaoNikufraCPIS.xlsx` export). Read-only; clicking "Exportar"
    in the top bar produces the downloadable .xlsx with a period
    selector (1 dia / 1 semana / 1 mês / 3 / 6 meses / 1 ano).
    """
    from app.web.db import conn

    where = ["pr.sheet_iso_date IS NOT NULL"]
    params: list = []
    if operador:
        where.append(
            "(UPPER(pr.operador) = UPPER(?) OR UPPER(s.operador) = UPPER(?))"
        )
        params.extend([operador, operador])
    if of_filter:
        where.append("pr.of = ?")
        params.append(of_filter.strip())

    sql = f"""
        SELECT pr.*,
               s.operador AS validated_operador,
               json_extract(s.sheet_data, '$.header.setor_maquina') AS setor_maquina,
               json_extract(s.sheet_data, '$.header.n_operador') AS n_operador,
               json_extract(s.sheet_data, '$.header.cod_maquina') AS header_cod_maquina,
               json_extract(s.sheet_data, '$.header.pernr') AS header_pernr
          FROM production_rows pr
          JOIN sheets s ON s.id = pr.sheet_id
         WHERE {' AND '.join(where)}
         ORDER BY pr.sheet_iso_date DESC, pr.sheet_id DESC, pr.row_index ASC
         LIMIT ?
    """
    params.append(limit)
    with conn() as c:
        raw_rows = [dict(r) for r in c.execute(sql, params).fetchall()]

    cpis_rows = [export._build_cpis_row(r) for r in raw_rows]

    # Distinct operators for the filter dropdown (reuse db helper)
    operadores = db.list_distinct_operadores()

    # Total count (separate cheap query, no limit) for the row counter
    with conn() as c:
        total_rows = c.execute(
            "SELECT COUNT(*) FROM production_rows WHERE sheet_iso_date IS NOT NULL"
        ).fetchone()[0]

    return templates.TemplateResponse(
        request,
        "excel.html",
        {
            "rows": cpis_rows,
            "columns": export.CPIS_COLUMNS,
            "operadores": operadores,
            "operador_filter": operador or "",
            "of_filter": of_filter or "",
            "limit": limit,
            "shown": len(cpis_rows),
            "total_rows": total_rows,
            "active_tab": "excel",
        },
    )


@app.get("/export/cpis")
def export_cpis(
    date_from: str | None = None,
    date_to: str | None = None,
    operador: str | None = None,
    sector: str | None = None,
) -> Response:
    """CPIS migration export — single-sheet .xlsx matching
    `MigracaoNikufraCPIS.xlsx` (17 columns, `Folha1`).

    One row per kanban production row in the period. Peso/Desperdício
    computed via `geometry.row_waste()` (trapezoidal column, steel ρ=7.85
    g/cm³). `Cód. Máquina` derived from setor_maquina (BOBINE-FORMATO →
    M032, etc.). Query params identical to `/export` (R69: same date /
    sector semantics).
    """
    df = (date_from or "").strip() or None
    dt_ = (date_to or "").strip() or None
    sec = (sector or "").strip() or None
    if bool(df) != bool(dt_):
        raise HTTPException(400, "provide both date_from and date_to, or neither (= sempre)")
    if df and dt_:
        if not _ISO_DATE_RE.match(df) or not _ISO_DATE_RE.match(dt_):
            raise HTTPException(400, "date_from and date_to must be YYYY-MM-DD")
        if dt_ < df:
            raise HTTPException(400, "date_to must be >= date_from")
    if sec and sec not in PRODUCTION_SECTORS:
        raise HTTPException(400, f"sector must be one of {list(PRODUCTION_SECTORS)}")
    xlsx_bytes = export.build_cpis_workbook(df, dt_, operador, sec)
    filename = export.cpis_filename_for(df, dt_, operador, sec)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/dashboard/production", response_class=HTMLResponse)
def dashboard_production(
    request: Request,
    date: str | None = None,
    period: str = "day",
) -> Response:
    """Round 29 Phase C — Production sector flow.

    Round 34: + period selector (day/week/month/year).
    When no date specified AND no data for today, fall back to the
    most-recent-non-future date with data.
    """
    if date and not _ISO_DATE_RE.match(date):
        date = None
    if period not in ("day", "week", "month", "year"):
        period = "day"
    if not date:
        # Try today first; if empty, fall back to most recent NON-FUTURE date.
        from app.web.db import conn
        with conn() as c:
            today_count = c.execute(
                "SELECT COUNT(*) FROM production_rows WHERE sheet_iso_date = DATE('now', 'localtime')"
            ).fetchone()[0]
            if today_count == 0:
                row = c.execute(
                    "SELECT sheet_iso_date FROM production_rows "
                    "WHERE sheet_iso_date IS NOT NULL "
                    "  AND sheet_iso_date <= DATE('now', 'localtime') "
                    "ORDER BY sheet_iso_date DESC LIMIT 1"
                ).fetchone()
                if row and row[0]:
                    date = row[0]
    overview = kpis.production_overview(date=date, period=period)
    return templates.TemplateResponse(
        request,
        "dashboard_production.html",
        {
            "overview": overview,
            "date_filter": date,
            "period": period,
            "active_tab": "production",
        },
    )


@app.get("/dashboard/nesting", response_class=HTMLResponse)
def dashboard_nesting(request: Request) -> Response:
    """Round 54 Fase 7 — Gemini (TPL102) nesting dashboard.

    Aggregates m²/qty across all 3 Gemini machines (GASPARINI / HPE32 /
    HD36). Domain is chapa nesting, distinct from column production —
    deserves its own KPI page.
    """
    from app.gemini import list_gemini_sheets, aggregate_gemini
    db_path = _DATA_DIR / "app.db"
    sheets = list_gemini_sheets(db_path)
    agg = aggregate_gemini(sheets)
    return templates.TemplateResponse(
        request,
        "dashboard_nesting.html",
        {
            "sheets": sheets,
            "agg": agg,
            "active_tab": "nesting",
        },
    )


@app.get("/dashboard/downtime", response_class=HTMLResponse)
def dashboard_downtime(request: Request) -> Response:
    """Round 54 Fase 6 — Paragens (downtime) dashboard.

    Pulls all paragens sheets (template_name=quinadora_pav4_paragens or
    legacy sheets with setor=QUINADORA PAV.4) and aggregates total
    minutes by operador / motivo / resolved status.

    Empty state when no paragens sheets exist yet.
    """
    from app.downtime import list_downtime_sheets, aggregate_downtime
    db_path = _DATA_DIR / "app.db"
    sheets = list_downtime_sheets(db_path)
    agg = aggregate_downtime(sheets)
    return templates.TemplateResponse(
        request,
        "dashboard_downtime.html",
        {
            "sheets": sheets,
            "agg": agg,
            "active_tab": "downtime",
        },
    )


@app.get("/dashboard/workers", response_class=HTMLResponse)
def dashboard_workers(request: Request) -> Response:
    workers = kpis.workers_summary()
    matrix = kpis.operador_cliente_matrix()
    return templates.TemplateResponse(
        request,
        "dashboard_workers.html",
        {"workers": workers, "matrix": matrix, "active_tab": "workers"},
    )


@app.get("/dashboard/workers/{operador}", response_class=HTMLResponse)
def dashboard_worker_detail(request: Request, operador: str) -> Response:
    detail = kpis.worker_detail(operador)
    return templates.TemplateResponse(
        request,
        "dashboard_worker_detail.html",
        {"detail": detail, "active_tab": "workers"},
    )


@app.get("/dashboard/jobs", response_class=HTMLResponse)
def dashboard_jobs(request: Request) -> Response:
    jobs = kpis.jobs_summary()
    leads = kpis.lead_time_per_of()
    return templates.TemplateResponse(
        request,
        "dashboard_jobs.html",
        {"jobs": jobs, "leads": leads, "active_tab": "jobs"},
    )


@app.get("/dashboard/jobs/{of_number:path}", response_class=HTMLResponse)
def dashboard_job_detail(request: Request, of_number: str) -> Response:
    """Use ``:path`` converter so OFs like ``OF262107/70`` (Dossier
    notation) don't get split by the router. Kanban OFs are 6 digits
    and unaffected."""
    detail = kpis.job_detail(of_number)
    return templates.TemplateResponse(
        request,
        "dashboard_job_detail.html",
        {"detail": detail, "active_tab": "jobs"},
    )


@app.get("/dashboard/trends", response_class=HTMLResponse)
def dashboard_trends(request: Request) -> Response:
    daily = kpis.daily_trend(30)
    weekly = kpis.weekly_trend(12)
    monthly = kpis.monthly_trend(12)
    top_clientes = kpis.top_clientes(12)
    top_modelos = kpis.top_modelos(12)
    matrix_modelo = kpis.operador_modelo_matrix(12, top_n=15)
    return templates.TemplateResponse(
        request,
        "dashboard_trends.html",
        {
            "daily": daily, "weekly": weekly, "monthly": monthly,
            "top_clientes": top_clientes, "top_modelos": top_modelos,
            "matrix_modelo": matrix_modelo,
            "active_tab": "trends",
        },
    )


@app.get("/dashboard/alerts", response_class=HTMLResponse)
def dashboard_alerts(request: Request) -> Response:
    data = kpis.alerts()
    return templates.TemplateResponse(
        request,
        "dashboard_alerts.html",
        {"data": data, "active_tab": "alerts"},
    )


@app.get("/pair", response_class=HTMLResponse)
def pair_page(request: Request) -> Response:
    """QR-code pairing: shows a QR pointing at /capture so a phone on the
    same LAN can scan and open the camera page directly without typing
    the IP. The QR is generated client-side from the page URL."""
    return templates.TemplateResponse(request, "pair.html", {})


# ============================================================================
#  Aprendizagens & Regras  (/learnings)  — Round 98 learning engine
# ============================================================================

def _learning_user(request: Request) -> str:
    """Best-effort reviewer identity for the learnings audit trail."""
    return request.headers.get("X-Forwarded-User") or "web"


def _refresh_overlay() -> None:
    """Re-materialise the learned overlay and reload it into the running
    pipeline. Called after any approve / reject / edit / delete so the
    OCR snap sees the change without a restart."""
    try:
        learning_materialize.materialize_overlay()
        learning_scheduler.reload_pipeline()
    except Exception as e:  # noqa: BLE001
        print(f"[learning] overlay refresh failed: {e}", file=sys.stderr)


def _rule_item_response(request: Request, learning_id: int) -> Response:
    rule = learning_store.get_proposal(learning_id)
    if rule is None:
        return HTMLResponse("")
    return templates.TemplateResponse(request, "_rule_item.html", {"r": rule})


@app.get("/learnings", response_class=HTMLResponse)
def learnings_page(
    request: Request,
    status: str = Query(""),
    kind: str = Query(""),
) -> Response:
    """The Aprendizagens & Regras page — rules tab + LLM analyst tab."""
    rules = learning_store.list_proposals(status=status or None, kind=kind or None)
    return templates.TemplateResponse(request, "learnings.html", {
        "rules": rules,
        "filter_status": status,
        "filter_kind": kind,
        "counts": learning_store.count_by_status(),
        "metric_latest": learning_metrics.corrections_per_sheet(),
        "metric_trend": learning_metrics.corrections_trend(),
        "attractors": attractors.compute_attractors(top_n=12),
        "production_sectors": db.list_distinct_setores(),
    })


@app.get("/learnings/rules", response_class=HTMLResponse)
def learnings_rules(
    request: Request,
    status: str = Query(""),
    kind: str = Query(""),
) -> Response:
    """Rules list fragment (HTMX target for filters)."""
    rules = learning_store.list_proposals(status=status or None, kind=kind or None)
    return templates.TemplateResponse(
        request, "_learnings_rules.html", {"rules": rules}
    )


@app.post("/learnings/rules/{learning_id}/approve", response_class=HTMLResponse)
def learnings_rule_approve(request: Request, learning_id: int) -> Response:
    learning_store.approve_proposal(learning_id, _learning_user(request))
    _refresh_overlay()
    return _rule_item_response(request, learning_id)


@app.post("/learnings/rules/{learning_id}/reject", response_class=HTMLResponse)
def learnings_rule_reject(request: Request, learning_id: int) -> Response:
    learning_store.reject_proposal(learning_id, _learning_user(request))
    _refresh_overlay()
    return _rule_item_response(request, learning_id)


@app.post("/learnings/rules/{learning_id}/edit", response_class=HTMLResponse)
def learnings_rule_edit(
    request: Request,
    learning_id: int,
    to: str = Form(""),
    value: str = Form(""),
    weight: str = Form(""),
    note: str = Form(""),
) -> Response:
    rule = learning_store.get_proposal(learning_id)
    if rule is None:
        return HTMLResponse("")
    payload = dict(rule["payload"]) if isinstance(rule["payload"], dict) else {}
    if to.strip():
        payload["to"] = to.strip()
    if value.strip():
        payload["value"] = value.strip()
    if weight.strip():
        try:
            payload["weight"] = float(weight.replace(",", "."))
        except ValueError:
            pass
    learning_store.update_proposal_payload(
        learning_id, payload, _learning_user(request),
        note=note.strip() or None,
    )
    _refresh_overlay()
    return _rule_item_response(request, learning_id)


@app.post("/learnings/rules/{learning_id}/delete", response_class=HTMLResponse)
def learnings_rule_delete(request: Request, learning_id: int) -> Response:
    learning_store.delete_proposal(learning_id)
    _refresh_overlay()
    return HTMLResponse("")


@app.post("/learnings/rules/create")
async def learnings_rule_create(request: Request) -> JSONResponse:
    """Create a learning by hand — used by the LLM tab's 'Adicionar ao
    Separador 1' button. Lands in quarantine awaiting human approval."""
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        return JSONResponse({"ok": False, "error": "JSON inválido"}, status_code=400)
    kind = (body.get("kind") or "").strip()
    payload = body.get("payload")
    if not kind or not isinstance(payload, dict):
        return JSONResponse(
            {"ok": False, "error": "kind/payload em falta"}, status_code=400
        )
    res = learning_store.create_manual_proposal(
        kind, payload,
        field=body.get("field"),
        template_name=body.get("template_name"),
        origin=body.get("origin") or "llm",
        decided_by=_learning_user(request),
    )
    return JSONResponse({"ok": True, "id": res["id"], "created": res["created"]})


@app.post("/learning/run", response_class=HTMLResponse)
def learning_run(
    request: Request,
    status: str = Query(""),
    kind: str = Query(""),
) -> Response:
    """Force a learning cycle now (the 'Minerar agora' button)."""
    try:
        learning_scheduler.run_learning_cycle()
    except Exception as e:  # noqa: BLE001
        print(f"[learning] manual run failed: {e}", file=sys.stderr)
    rules = learning_store.list_proposals(status=status or None, kind=kind or None)
    return templates.TemplateResponse(
        request, "_learnings_rules.html", {"rules": rules}
    )


@app.get("/learning/metrics")
def learning_metrics_json() -> JSONResponse:
    return JSONResponse({
        "latest": learning_metrics.corrections_per_sheet(),
        "trend": learning_metrics.corrections_trend(),
    })


@app.get("/learnings/llm/atratores")
def learnings_atratores() -> JSONResponse:
    return JSONResponse({"attractors": attractors.compute_attractors(top_n=15)})


@app.post("/learnings/llm/chat")
async def learnings_llm_chat(request: Request) -> JSONResponse:
    """One chat turn with the local Ollama analyst. Returns the JSON
    envelope {reply, charts, proposed_rules}."""
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        return JSONResponse(
            {"reply": "Pedido inválido.", "charts": [], "proposed_rules": []},
            status_code=400,
        )
    message = (body.get("message") or "").strip()
    history = body.get("history") if isinstance(body.get("history"), list) else []
    if not message:
        return JSONResponse(
            {"reply": "Escreve uma pergunta.", "charts": [], "proposed_rules": []}
        )
    # llm_assistant.chat() does a blocking HTTP call to Ollama — run it in a
    # worker thread so it never freezes the async event loop.
    envelope = await asyncio.to_thread(llm_assistant.chat, message, history)
    return JSONResponse(envelope)


# ----- R110.C — Agent proposals + policies (endpoints REST) -----
from fastapi.encoders import jsonable_encoder  # noqa: E402

# R117: `kernel` import foi promovido para o topo do ficheiro.


def _json_ok(data: dict) -> JSONResponse:
    """JSONResponse com datetime → ISO automático."""
    return JSONResponse(jsonable_encoder(data))


# ----- R110.E — Kernel state visibility -----

@app.get("/agent/kernel/state")
def kernel_state_endpoint() -> JSONResponse:
    return _json_ok({"state": kernel.get_state(),
                     "total_events": kernel.count_events()})


@app.get("/agent/kernel/events")
def kernel_events_endpoint(limit: int = 50) -> JSONResponse:
    return _json_ok({"events": kernel.list_recent_events(limit=limit)})


@app.get("/agent/proposals")
def agent_proposals_list(
    status: str = "",
    kind: str = "",
    limit: int = 50,
) -> JSONResponse:
    """Lista propostas do agente Qwen. Filtros opcionais status/kind."""
    rows = db.list_proposals(status=status, kind=kind, limit=limit)
    return _json_ok({"proposals": rows, "count": len(rows)})


@app.post("/agent/proposals/{proposal_id}/approve")
def agent_proposal_approve(proposal_id: int) -> JSONResponse:
    """Aceita uma proposta: corre eval gate, promove nova policy_version,
    aplica template_overlay se for o caso."""
    from app.pipeline import policy_engine
    version_id = policy_engine.promote_policy_from_proposal(
        proposal_id, created_by="human-approval"
    )
    if version_id is None:
        return JSONResponse(
            {"status": "error", "error": "Proposta não encontrada."},
            status_code=404,
        )
    proposal = db.get_proposal(proposal_id)
    kernel.emit_event("proposal_decided",
                      {"proposal_id": proposal_id, "decision": "accepted"})
    kernel.emit_event("policy_promoted", {"version": version_id})
    return _json_ok({
        "status": "ok",
        "proposal_id": proposal_id,
        "policy_version": version_id,
        "proposal": proposal,
    })


@app.post("/agent/proposals/{proposal_id}/reject")
def agent_proposal_reject(proposal_id: int) -> JSONResponse:
    from app.pipeline import policy_engine
    ok = policy_engine.reject_proposal(proposal_id, decided_by="human")
    if not ok:
        return JSONResponse(
            {"status": "error", "error": "Proposta não encontrada."},
            status_code=404,
        )
    kernel.emit_event("proposal_decided",
                      {"proposal_id": proposal_id, "decision": "rejected"})
    return JSONResponse({"status": "ok", "proposal_id": proposal_id})


@app.get("/agent/policies")
def agent_policies_list(limit: int = 30) -> JSONResponse:
    return _json_ok({"versions": db.list_policy_versions(limit=limit)})


@app.get("/agent/policies/active")
def agent_policies_active() -> JSONResponse:
    return _json_ok({"policy": db.get_active_policy_version()})


@app.post("/agent/policies/rollback")
def agent_policies_rollback() -> JSONResponse:
    """Reverte para a parent_version da activa actual."""
    from app.pipeline import policy_engine
    new_version = policy_engine.rollback_to_parent(reason="manual-rollback")
    if new_version is None:
        return JSONResponse(
            {"status": "error",
             "error": "Sem versão anterior para reverter."},
            status_code=400,
        )
    kernel.emit_event("policy_rolled_back", {"version": new_version})
    return JSONResponse({"status": "ok", "active_version": new_version})


@app.get("/agent/circuit-breaker")
def agent_circuit_breaker() -> JSONResponse:
    """Verifica estado do circuit breaker (recommended action). Não age sozinho."""
    from app.pipeline import policy_engine
    return JSONResponse(policy_engine.check_circuit_breaker())


@app.get("/agent/sessions")
def agent_sessions_list(limit: int = 50) -> JSONResponse:
    return _json_ok({"sessions": db.list_qwen_sessions(limit=limit)})


@app.get("/agent/charts")
def agent_charts_list(pinned_only: bool = False, limit: int = 30) -> JSONResponse:
    return _json_ok({
        "charts": db.list_qwen_charts(limit=limit, pinned_only=pinned_only)
    })


@app.post("/agent/charts/{chart_id}/pin")
def agent_chart_pin(chart_id: int) -> JSONResponse:
    ok = db.set_chart_pinned(chart_id, True)
    if not ok:
        return JSONResponse({"status": "error",
                             "error": "Chart não encontrado."},
                            status_code=404)
    return JSONResponse({"status": "ok"})


@app.post("/agent/charts/{chart_id}/unpin")
def agent_chart_unpin(chart_id: int) -> JSONResponse:
    ok = db.set_chart_pinned(chart_id, False)
    if not ok:
        return JSONResponse({"status": "error",
                             "error": "Chart não encontrado."},
                            status_code=404)
    return JSONResponse({"status": "ok"})


# ----- R115 — /obras (ponto de situação por cliente e OV) -----

@app.get("/obras", response_class=HTMLResponse)
def obras_page(
    request: Request,
    closed: int = 0,
    q: str = "",
    tab: str = "cliente",
    ov: str = "",
) -> Response:
    """Página de ponto-de-situação. Junta o plan (refs) com production_rows."""
    from app.pipeline import obras_status
    data = obras_status.compute_obras_status(include_closed=bool(closed))
    return templates.TemplateResponse(request, "obras.html", {
        "data": data,
        "include_closed": bool(closed),
        "initial_search": q,
        "initial_tab": "ov" if tab == "ov" else "cliente",
        "initial_ov": ov,
    })


@app.get("/api/obras/summary")
def api_obras_summary(closed: int = 0) -> JSONResponse:
    from app.pipeline import obras_status
    return _json_ok(obras_status.compute_obras_status(include_closed=bool(closed)))


@app.get("/api/obras/ov/{ov_str}")
def api_obras_ov(ov_str: str) -> JSONResponse:
    from app.pipeline import obras_status
    det = obras_status.get_ov_detail(ov_str)
    if det is None:
        return JSONResponse({"status": "error",
                             "error": f"OV {ov_str} não encontrada no plan"},
                            status_code=404)
    return _json_ok(det)


@app.post("/api/obras/refresh")
def api_obras_refresh() -> JSONResponse:
    from app.pipeline import obras_status
    obras_status.invalidate_cache()
    return JSONResponse({"status": "ok"})


# ----- CSV builder (mirrors ocr6.write_csv but to string) -----

def _to_3block_csv(filename: str, data: dict) -> str:
    """Produce CSV string in the Metalogalva 3-block format.

    Round 54 — template-aware. The block headers + column lists are
    derived from the TemplateSpec for this sheet:
      - block 2 heading: ``template.csv_block_label`` (e.g. "TABELA DE
        PRODUÇÃO" for bobine, "TABELA DE PARAGENS" for paragens,
        "TABELA DE NESTING" for Gemini).
      - block 2 columns: ``template.row_fields`` (uppercased).
      - block 3 (footer) columns: ``template.footer_fields``. Templates
        with no footer (paragens) omit the block entirely.

    For BOBINE-FORMATO this produces byte-identical output to the prior
    R53 implementation so the factory validator (which knows only this
    schema) keeps parsing legacy + new bobine sheets identically.
    """
    import csv
    import io

    # Lazy import — avoids circular if templates_registry ever imports main
    from app.templates_registry import (
        DEFAULT_TEMPLATE, detect_template, get_template,
    )

    # Resolve template — explicit > inferred > default
    tname = (data or {}).get("template_name")
    if tname:
        template = get_template(tname)
    else:
        setor = ((data or {}).get("header") or {}).get("setor_maquina", "")
        template = detect_template(setor) if setor else DEFAULT_TEMPLATE

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\r\n")
    h = data.get("header", {}) or {}
    f_ = data.get("footer", {}) or {}

    # --- Block 1: CABEÇALHO (identical for all templates) ---
    w.writerow(["### CABEÇALHO"])
    w.writerow(["FICHEIRO", "DATA", "OPERADOR", "N_OPERADOR", "SETOR_MAQUINA"])
    w.writerow([
        filename, h.get("data", ""), h.get("operador", ""),
        h.get("n_operador", ""), h.get("setor_maquina", ""),
    ])
    w.writerow([])

    # --- Block 2: TABELA (per-template label + columns) ---
    w.writerow([f"### {template.csv_block_label}"])
    # Column header — common prefix (FICHEIRO/DATA/OPERADOR) + template fields
    column_header = ["FICHEIRO", "DATA", "OPERADOR"] + [
        f.upper() for f in template.row_fields
    ]
    w.writerow(column_header)
    for row in data.get("rows", []) or []:
        row_out = [filename, h.get("data", ""), h.get("operador", "")]
        row_out.extend(str(row.get(f, "") or "") for f in template.row_fields)
        w.writerow(row_out)
    w.writerow([])

    # --- Block 3: RODAPÉ (skipped when template has no footer fields) ---
    if template.footer_fields:
        w.writerow(["### RODAPÉ"])
        footer_header = ["FICHEIRO", "DATA", "OPERADOR"] + [
            f.upper() for f in template.footer_fields
        ]
        w.writerow(footer_header)
        footer_row = [filename, h.get("data", ""), h.get("operador", "")]
        footer_row.extend(str(f_.get(f, "") or "") for f in template.footer_fields)
        w.writerow(footer_row)

    # UTF-8 BOM prefix (Excel friendly) — preserved from R53.
    return "﻿" + buf.getvalue()
