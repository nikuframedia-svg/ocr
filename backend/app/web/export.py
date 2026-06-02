"""Round 29 Phase D — Excel multi-sheet export.

Output structure:
- `Resumo`: period totals, per-day summary, per-operator summary
- One sheet per production day (`DD-MM-YYYY`):
  - Day header row (date, total qty, total sheets, total OFs)
  - Sub-tables per operator: their rows for that day in the standard
    13-column kanban format

Reads from `production_rows` (denormalized) for fast iteration.

CPIS migration export — see `build_cpis_workbook()`. Produces a single
flat sheet matching the current `MigracaoNikufraCPIS.xlsx` template.
"""
from __future__ import annotations

import datetime as dt
import io
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.production.weights import calculate_row_weights, find_plan_entry
from app.templates_registry import detect_template

from .db import conn
from .kpis import _derive_cod_maquina

# Order of columns in the per-day per-operator sub-tables (matches kanban)
ROW_COLUMNS = [
    ("pri", "PRI"),
    ("cliente", "CLIENTE"),
    ("ov", "OV"),
    ("of", "OF"),
    ("modelo", "MODELO"),
    ("qtd", "QTD"),
    ("comp_mm", "COMP_MM"),
    ("larg_mm", "LARG_MM"),
    ("lote", "LOTE"),
    ("coni", "CONI"),
    ("esp", "ESP"),
    ("lbase", "LBASE"),
    ("ltopo", "LTOPO"),
]


# ---- styles ------------------------------------------------------------

_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

_FONT_BASE = Font(name="Inter", size=10)
_FONT_BOLD = Font(name="Inter", size=10, bold=True)
_FONT_HEADER = Font(name="Inter", size=11, bold=True, color="FFFFFF")
_FONT_TITLE = Font(name="Inter", size=14, bold=True, color="16140F")
_FONT_HERO = Font(name="Inter", size=20, bold=True, color="16140F")

_FILL_HEADER = PatternFill("solid", fgColor="2F5597")
_FILL_OPERATOR = PatternFill("solid", fgColor="DCE6F1")
_FILL_DAYHEAD = PatternFill("solid", fgColor="F0E2C9")


def _filter_by_sector(rows: list[dict], sector: str | None) -> list[dict]:
    """R69: post-filter rows whose detected template phase matches ``sector``.

    Reuses the canonical ``detect_template()`` which already handles all
    alias variants (exact + substring) — same logic as engine/UI/dashboards.
    Rows with unknown setor_maquina fall back to bobine_formato → kept
    only if ``sector`` is "Bobine Formato".
    """
    if not sector:
        return rows
    return [
        r for r in rows
        if detect_template(r.get("setor_maquina") or "").phase == sector
    ]


def _query_rows(
    date_from: str | None,
    date_to: str | None,
    operador: str | None,
    sector: str | None = None,
    validated_only: bool = False,
) -> list[dict]:
    """Return all production_rows in [date_from, date_to] (inclusive),
    optionally filtered by operador and sector. Joined with sheets to get
    setor_maquina + canonical sheets.operador (post-validate).

    R69: dates may be None (= "sempre"); sector applied via post-filter
    using ``detect_template`` (canonical alias matching).

    R130: ``validated_only=True`` restringe a folhas com ``sheets.status =
    'validated'`` (operador carregou em "Validar"). Default False = inclui
    rascunhos.
    """
    where: list[str] = []
    params: list = []
    if date_from and date_to:
        where.append("pr.sheet_iso_date BETWEEN ? AND ?")
        params.extend([date_from, date_to])
    if operador:
        where.append(
            "(UPPER(pr.operador) = UPPER(?) OR UPPER(s.operador) = UPPER(?))"
        )
        params.extend([operador, operador])
    if validated_only:
        where.append("s.status = 'validated'")

    sql = """
        SELECT pr.*,
               s.operador AS validated_operador,
               json_extract(s.sheet_data, '$.header.setor_maquina') AS setor_maquina,
               json_extract(s.sheet_data, '$.footer.colunas_produzidas') AS colunas_produzidas,
               json_extract(s.sheet_data, '$.footer.horas_trabalhadas') AS horas_trabalhadas
          FROM production_rows pr
          JOIN sheets s ON s.id = pr.sheet_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY pr.sheet_iso_date ASC, pr.operador ASC, pr.sheet_id ASC, pr.row_index ASC"

    with conn() as c:
        rows = [dict(r) for r in c.execute(sql, params).fetchall()]
    return _filter_by_sector(rows, sector)


def _operador_canon(row: dict) -> str:
    """Use the validated operador if available, else the row's stored operador."""
    return (row.get("validated_operador") or row.get("operador") or "—").strip() or "—"


def _date_iso_to_pt(iso: str) -> str:
    """YYYY-MM-DD → DD-MM-YYYY for sheet titles + day headers."""
    if not iso or len(iso) != 10:
        return iso or ""
    return f"{iso[8:10]}-{iso[5:7]}-{iso[0:4]}"


def _style_cell(cell, *, bold=False, fill=None, align="left"):
    cell.font = _FONT_BOLD if bold else _FONT_BASE
    cell.border = _BORDER
    if fill is not None:
        cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center")


def _write_resumo(
    ws,
    all_rows: list[dict],
    date_from: str | None,
    date_to: str | None,
    sector: str | None = None,
) -> None:
    """First sheet: period summary + per-day summary + per-operator summary.

    R69: shows "Sempre" when dates are None; adds a SETOR row when filtered.
    """
    ws.title = "Resumo"

    # Title
    ws["A1"] = "Metalogalva — Resumo de Produção"
    ws["A1"].font = _FONT_TITLE
    ws.merge_cells("A1:F1")

    ws["A2"] = "Período"
    if date_from and date_to:
        ws["B2"] = f"{_date_iso_to_pt(date_from)} a {_date_iso_to_pt(date_to)}"
    else:
        ws["B2"] = "Sempre"
    ws["A2"].font = _FONT_BOLD

    # R69: SETOR row when a sector filter is applied
    if sector:
        ws["A3"] = "Setor"
        ws["B3"] = sector
        ws["A3"].font = _FONT_BOLD

    # Aggregate
    total_qty = sum(r["qtd"] or 0 for r in all_rows)
    sheet_ids = {r["sheet_id"] for r in all_rows}
    of_set = {r["of"] for r in all_rows if r["of"]}
    operadores = {_operador_canon(r) for r in all_rows}

    ws["A4"] = "TOTAL COLUNAS"
    ws["B4"] = total_qty
    ws["A4"].font = _FONT_BOLD
    ws["B4"].font = _FONT_HERO

    ws["A5"] = "TOTAL KANBANS"
    ws["B5"] = len(sheet_ids)
    ws["A5"].font = _FONT_BOLD

    ws["A6"] = "OFs ÚNICAS"
    ws["B6"] = len(of_set)
    ws["A6"].font = _FONT_BOLD

    ws["A7"] = "OPERADORES"
    ws["B7"] = len(operadores)
    ws["A7"].font = _FONT_BOLD

    # Per-day summary
    by_day: dict[str, dict] = defaultdict(lambda: {"qtd": 0, "sheets": set(), "ofs": set(), "ops": set()})
    for r in all_rows:
        d = by_day[r["sheet_iso_date"]]
        d["qtd"] += r["qtd"] or 0
        d["sheets"].add(r["sheet_id"])
        if r["of"]:
            d["ofs"].add(r["of"])
        d["ops"].add(_operador_canon(r))

    row_idx = 9
    ws.cell(row=row_idx, column=1, value="Por dia").font = _FONT_BOLD
    row_idx += 1
    headers = ["DATA", "COLUNAS", "KANBANS", "OFs", "OPERADORES"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=ci, value=h)
        _style_cell(cell, bold=True, fill=_FILL_HEADER)
        cell.font = _FONT_HEADER
    row_idx += 1
    for d_iso in sorted(by_day):
        d = by_day[d_iso]
        ws.cell(row=row_idx, column=1, value=_date_iso_to_pt(d_iso))
        ws.cell(row=row_idx, column=2, value=d["qtd"])
        ws.cell(row=row_idx, column=3, value=len(d["sheets"]))
        ws.cell(row=row_idx, column=4, value=len(d["ofs"]))
        ws.cell(row=row_idx, column=5, value=len(d["ops"]))
        for ci in range(1, 6):
            _style_cell(ws.cell(row=row_idx, column=ci))
        row_idx += 1

    # Per-operator summary
    row_idx += 2
    by_op: dict[str, dict] = defaultdict(lambda: {"qtd": 0, "sheets": set(), "ofs": set(), "days": set()})
    for r in all_rows:
        op = _operador_canon(r)
        by_op[op]["qtd"] += r["qtd"] or 0
        by_op[op]["sheets"].add(r["sheet_id"])
        if r["of"]:
            by_op[op]["ofs"].add(r["of"])
        by_op[op]["days"].add(r["sheet_iso_date"])

    ws.cell(row=row_idx, column=1, value="Por operador").font = _FONT_BOLD
    row_idx += 1
    headers = ["OPERADOR", "COLUNAS", "KANBANS", "OFs", "DIAS"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=ci, value=h)
        _style_cell(cell, bold=True, fill=_FILL_HEADER)
        cell.font = _FONT_HEADER
    row_idx += 1
    for op in sorted(by_op, key=lambda x: -by_op[x]["qtd"]):
        d = by_op[op]
        ws.cell(row=row_idx, column=1, value=op)
        ws.cell(row=row_idx, column=2, value=d["qtd"])
        ws.cell(row=row_idx, column=3, value=len(d["sheets"]))
        ws.cell(row=row_idx, column=4, value=len(d["ofs"]))
        ws.cell(row=row_idx, column=5, value=len(d["days"]))
        for ci in range(1, 6):
            _style_cell(ws.cell(row=row_idx, column=ci))
        row_idx += 1

    # Column widths
    for ci, width in enumerate([22, 14, 12, 10, 12, 14], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width


def _write_day_sheet(wb: openpyxl.Workbook, day_iso: str, day_rows: list[dict]) -> None:
    """Per-day worksheet: day-header row + sub-tables per operator."""
    title = _date_iso_to_pt(day_iso)  # 'DD-MM-YYYY' (Excel allows '-' in titles)
    ws = wb.create_sheet(title=title[:31])  # Excel max 31 chars

    # Day header
    qtd = sum(r["qtd"] or 0 for r in day_rows)
    sheets_n = len({r["sheet_id"] for r in day_rows})
    ofs_n = len({r["of"] for r in day_rows if r["of"]})

    ws["A1"] = f"Produção — {title}"
    ws["A1"].font = _FONT_TITLE
    ws.merge_cells("A1:F1")

    ws["A2"] = "TOTAL COLUNAS"
    ws["B2"] = qtd
    ws["A2"].font = _FONT_BOLD
    ws["B2"].font = _FONT_HERO

    ws["C2"] = "KANBANS"
    ws["D2"] = sheets_n
    ws["C2"].font = _FONT_BOLD
    ws["D2"].font = _FONT_BOLD

    ws["E2"] = "OFs"
    ws["F2"] = ofs_n
    ws["E2"].font = _FONT_BOLD
    ws["F2"].font = _FONT_BOLD

    # Group by operator
    by_op: dict[str, list[dict]] = defaultdict(list)
    for r in day_rows:
        by_op[_operador_canon(r)].append(r)

    row_idx = 4
    for op in sorted(by_op):
        # Operator header bar
        ws.cell(row=row_idx, column=1, value=op).font = _FONT_BOLD
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=len(ROW_COLUMNS))
        for ci in range(1, len(ROW_COLUMNS) + 1):
            ws.cell(row=row_idx, column=ci).fill = _FILL_OPERATOR
        row_idx += 1

        # Column headers
        for ci, (_, label) in enumerate(ROW_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=ci, value=label)
            _style_cell(cell, bold=True, fill=_FILL_HEADER)
            cell.font = _FONT_HEADER
        row_idx += 1

        # Rows
        for r in by_op[op]:
            for ci, (key, _) in enumerate(ROW_COLUMNS, start=1):
                v = r.get(key)
                cell = ws.cell(row=row_idx, column=ci, value=v if v is not None else "")
                _style_cell(cell)
            row_idx += 1

        row_idx += 1  # blank line between operators

    # Column widths — generous defaults for Inter font
    widths = [6, 14, 11, 10, 14, 7, 10, 10, 12, 10, 8, 9, 9]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def export_excel(
    date_from: str | None,
    date_to: str | None,
    operador: str | None = None,
    sector: str | None = None,
    validated_only: bool = False,
) -> bytes:
    """Build and return the .xlsx file bytes for the given period.

    R69: dates may be None (= "sempre"); sector is one of
    ``PRODUCTION_SECTORS`` or None (= all sectors).
    R130: ``validated_only=True`` exclui folhas em rascunho (status != 'validated').
    """
    rows = _query_rows(date_from, date_to, operador, sector, validated_only)

    wb = openpyxl.Workbook()
    ws = wb.active
    _write_resumo(ws, rows, date_from, date_to, sector)

    # Group by day, write one sheet per day with data
    by_day: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        if r["sheet_iso_date"]:
            by_day[r["sheet_iso_date"]].append(r)
    for day_iso in sorted(by_day):
        _write_day_sheet(wb, day_iso, by_day[day_iso])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _slugify(value: str) -> str:
    """Lowercase, replace whitespace with underscore, strip non-ascii."""
    s = value.strip().lower().replace(" ", "_")
    return "".join(ch for ch in s if ch.isalnum() or ch == "_")


def filename_for(
    date_from: str | None,
    date_to: str | None,
    operador: str | None = None,
    sector: str | None = None,
    validated_only: bool = False,
) -> str:
    """Slugged filename for the download. R69: supports "sempre" + sector.
    R130: sufixo `_validadas` quando o filtro de validação está activo."""
    period = (
        f"{date_from}_{date_to}"
        if (date_from and date_to)
        else "sempre"
    )
    sector_suffix = f"_{_slugify(sector)}" if sector else ""
    op_suffix = ""
    if operador:
        slug = operador.upper().replace(" ", "")
        slug = "".join(ch for ch in slug if ch.isalnum())
        op_suffix = f"_{slug}"
    validated_suffix = "_validadas" if validated_only else ""
    return f"metalogalva_{period}{sector_suffix}{op_suffix}{validated_suffix}.xlsx"


# ============================================================================
#  CPIS migration export (MigracaoNikufraCPIS.xlsx schema)
# ============================================================================

# Column order must match the user-supplied template exactly (folha "Folha1").
CPIS_COLUMNS: tuple[tuple[str, str], ...] = (
    ("data", "Data"),
    ("cod_funcionario", "Cód. Funcionário"),
    ("nome_funcionario", "Nome Funcionário"),
    ("setor_maquina_desc", "Setor / Máquina Desc."),
    ("cod_maquina", "Cód. Máquina"),
    ("of", "OF"),
    ("ov", "OV"),
    ("cliente", "Cliente"),
    ("modelo", "Modelo"),
    ("qtd", "QTD"),
    # R97 — Qtd Metros (Soldline/Laser); empty para outros.
    ("qtd_metros", "Qtd Metros"),
    # R86 — Gemini-only fields (HPE32/GASPARINI/HD36). Empty in non-Gemini rows.
    ("m2", "M²"),
    ("nesting", "Nesting"),
    # R97 — Cesta Nº (Expedição); Duração (calc fim-inicio para Gemini/Paragens)
    ("cesta_n", "Cesta Nº"),
    ("duracao", "Duração"),
    ("comp_mm", "Comprimento (mm)"),
    ("larg_mm", "Largura (mm)"),
    ("esp_mm", "Espessura (mm)"),
    ("coni", "CONI"),
    # R72 — 4 colunas weight-related em toneladas + nº chapas explícito
    # como pediu o user (2 fórmulas separadas, ambos pesos lado-a-lado).
    ("n_chapas", "Nº Chapas"),
    ("peso_consumido_t", "Peso Consumido (t)"),
    ("peso_produzido_t", "Peso Produzido (t)"),
    ("desperdicio_t", "Desperdício (t)"),
    ("desperdicio_pct", "% Desperdício"),
)


def _query_cpis_rows(
    date_from: str | None,
    date_to: str | None,
    operador: str | None,
    sector: str | None = None,
    validated_only: bool = False,
) -> list[dict]:
    """Pull production_rows + header.n_operador joined from sheets.

    Mirrors `_query_rows` but adds `n_operador` (header field, not on
    production_rows). R69: dates may be None (= "sempre"); sector applied
    via post-filter using ``detect_template``.

    R130: ``validated_only=True`` restringe a folhas com ``sheets.status =
    'validated'``.
    """
    where: list[str] = []
    params: list = []
    if date_from and date_to:
        where.append("pr.sheet_iso_date BETWEEN ? AND ?")
        params.extend([date_from, date_to])
    if operador:
        where.append(
            "(UPPER(pr.operador) = UPPER(?) OR UPPER(s.operador) = UPPER(?))"
        )
        params.extend([operador, operador])
    if validated_only:
        where.append("s.status = 'validated'")

    sql = """
        SELECT pr.*,
               s.operador AS validated_operador,
               json_extract(s.sheet_data, '$.header.setor_maquina') AS setor_maquina,
               json_extract(s.sheet_data, '$.header.n_operador') AS n_operador,
               json_extract(s.sheet_data, '$.header.cod_maquina') AS header_cod_maquina,
               json_extract(s.sheet_data, '$.header.pernr') AS header_pernr
          FROM production_rows pr
          JOIN sheets s ON s.id = pr.sheet_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += (
        " ORDER BY pr.sheet_iso_date ASC, pr.operador ASC,"
        "          pr.sheet_id ASC, pr.row_index ASC"
    )
    with conn() as c:
        rows = [dict(r) for r in c.execute(sql, params).fetchall()]
    return _filter_by_sector(rows, sector)


def _iso_to_date(iso: str | None) -> dt.date | None:
    if not iso or len(iso) != 10:
        return None
    try:
        return dt.date(int(iso[0:4]), int(iso[5:7]), int(iso[8:10]))
    except ValueError:
        return None


def _to_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(str(v).strip())
        except (TypeError, ValueError):
            return None


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except (TypeError, ValueError):
        return None


def _parse_time_to_minutes(s: str | None) -> int | None:
    """R97 — Parse a free-form time string to total minutes-since-midnight.

    Operators write times in many forms:
      "08:00"  → 480       (standard HH:MM)
      "9h15"   → 555       (Portuguese vernacular "9 horas e 15")
      "11h"    → 660       (only hour, no minutes)
      "9+15"   → 555       (OCR misread of "h" or ":")
      "09.30"  → 570       (dot separator)
      "1430"   → 870       (4-digit no separator)

    Returns None when unparseable.
    """
    if s is None:
        return None
    raw = str(s).strip().lower()
    if not raw:
        return None
    # Normalize separators to ":"
    import re as _re
    norm = _re.sub(r"[h\+\.\s,]", ":", raw)
    # Remove trailing colon (e.g. "11:" from "11h")
    norm = norm.rstrip(":")
    # Pure-digit 4-char like "1430" → "14:30"
    if norm.isdigit() and len(norm) == 4:
        norm = f"{norm[:2]}:{norm[2:]}"
    # Pure-digit ≤2 chars → just hours
    if norm.isdigit():
        try:
            h = int(norm)
            if 0 <= h <= 23:
                return h * 60
        except ValueError:
            return None
        return None
    parts = norm.split(":")
    try:
        h = int(parts[0]) if parts[0] else 0
        m = int(parts[1]) if len(parts) > 1 and parts[1] else 0
    except ValueError:
        return None
    if not (0 <= h <= 23 and 0 <= m <= 59):
        return None
    return h * 60 + m


def _compute_duracao(inicio: str | None, fim: str | None) -> str:
    """R97 — Duration HH:MM from inicio/fim strings.

    Accepts various time formats via ``_parse_time_to_minutes``. Returns
    "" when either side is missing/malformed. Handles past-midnight
    rollover (fim < inicio → add 24h).

    Used in CPIS export so Gemini (HPE32/GASPARINI) and Paragens rows
    show duration even though it's not persisted in DB.
    """
    mi = _parse_time_to_minutes(inicio)
    mf = _parse_time_to_minutes(fim)
    if mi is None or mf is None:
        return ""
    diff = mf - mi
    if diff < 0:
        diff += 24 * 60
    if diff == 0:
        return ""
    return f"{diff // 60:02d}:{diff % 60:02d}"


def _derive_ov_from_plan(row: dict, refs: dict | None) -> str:
    """R96 — Return the row's OV, falling back to plan_colunas when empty.

    Gemini templates (HPE32, GASPARINI) don't have OV in their row schema
    but each OF has an OV in `plan_colunas_cpis.xlsx`. When OCR row.ov is
    empty, look up the plan to get the canonical OV.
    """
    ov_val = (row.get("ov") or "").strip()
    if ov_val:
        return ov_val
    entry = find_plan_entry(row, refs)
    if not entry:
        return ""
    return str(entry.get("ov") or "").strip()


def _derive_cliente_from_plan(row: dict, refs: dict | None) -> str:
    """Return row cliente, falling back to the same plan match used for OV."""
    cliente_val = (row.get("cliente") or "").strip()
    if cliente_val:
        return cliente_val
    entry = find_plan_entry(row, refs)
    if not entry:
        return ""
    return str(entry.get("cliente") or "").strip()


def _build_cpis_row(row: dict, refs: dict | None = None) -> dict:
    """Project a production_rows dict into the current CPIS schema.

    Weight metrics come from ``app.production.weights`` so CPIS, dashboards
    and previews share the same source-priority rules.

    R96/R139: ``refs`` lets us derive missing OV/Cliente from the plan when
    the kanban doesn't capture those fields (Gemini/Acabamento TPL086).
    """
    qtd = _to_int(row.get("qtd"))
    weights = calculate_row_weights(row, refs=refs)
    peso_consumido_t = (
        weights.peso_consumido_kg / 1000.0
        if weights.peso_consumido_kg is not None else None
    )
    peso_produzido_t = (
        weights.peso_produzido_kg / 1000.0
        if weights.peso_produzido_kg is not None else None
    )
    desperdicio_t = (
        weights.desperdicio_kg / 1000.0
        if weights.desperdicio_kg is not None else None
    )

    setor_maquina = row.get("setor_maquina") or ""
    # Prefer OCR-extracted cod_maquina if the header has it (future kanbans).
    # Fall back to the derivation table for legacy sheets.
    cod_maquina = (
        (row.get("header_cod_maquina") or "").strip()
        or _derive_cod_maquina(setor_maquina)
    )

    nome = (row.get("validated_operador") or row.get("operador") or "").strip()

    # R70 + R79 — prefer canonical pernr (8 digits, e.g. "10000095") from
    # header.pernr (set by snap_operador). R79 fallback: when header.pernr
    # is empty (legacy sheet not yet snapped), DERIVE pernr structurally
    # from n_operador via the "1000" + zfill(4) rule. This guarantees the
    # CPIS export always shows 8-digit pernr regardless of when sheet was
    # uploaded.
    pernr_str = (row.get("header_pernr") or "").strip()
    if not pernr_str:
        cod_raw = str(row.get("n_operador") or "").strip().lstrip("0") or "0"
        try:
            cod_int = int(cod_raw)
            if cod_int > 0:
                pernr_str = f"1000{cod_int:04d}" if cod_int < 10000 else str(cod_int)
        except (ValueError, TypeError):
            pass
    cod_funcionario = _to_int(pernr_str) if pernr_str else None

    return {
        "data": _iso_to_date(row.get("sheet_iso_date")),
        "cod_funcionario": cod_funcionario,
        "nome_funcionario": nome,
        "setor_maquina_desc": setor_maquina,
        "cod_maquina": cod_maquina,
        "of": row.get("of") or "",
        # R96/R139 — derive OV/Cliente from plan when the sheet schema
        # does not capture them (Gemini/Acabamento TPL086).
        "ov": _derive_ov_from_plan(row, refs),
        "cliente": _derive_cliente_from_plan(row, refs),
        "modelo": row.get("modelo") or "",
        "qtd": qtd,
        # R97 — qtd_metros (Soldline/Laser). NULL/0 para outros templates.
        "qtd_metros": _to_float(row.get("qtd_metros")),
        # R86 — Gemini-only fields. Empty/NULL for non-Gemini rows.
        "m2": _to_float(row.get("m2")),
        "nesting": row.get("nesting") or "",
        # R97 — cesta_n (Expedição); duracao calc fim-inicio (Gemini/Paragens).
        "cesta_n": row.get("cesta_n") or "",
        "duracao": _compute_duracao(row.get("inicio"), row.get("fim")),
        "comp_mm": _to_int(
            weights.comp_mm if weights.comp_mm is not None else row.get("comp_mm")
        ),
        "larg_mm": _to_int(
            weights.larg_mm if weights.larg_mm is not None else row.get("larg_mm")
        ),
        "esp_mm": weights.esp_mm if weights.esp_mm is not None else _to_float(row.get("esp")),
        "coni": row.get("coni") or "",
        # R72 — 4 weight-related fields (vs old peso_kg/desperdicio_kg)
        "n_chapas": weights.n_chapas,
        "peso_consumido_t": (
            round(peso_consumido_t, 3) if peso_consumido_t is not None else None
        ),
        "peso_produzido_t": (
            round(peso_produzido_t, 3) if peso_produzido_t is not None else None
        ),
        "desperdicio_t": round(desperdicio_t, 3) if desperdicio_t is not None else None,
        "desperdicio_pct": (
            round(weights.desperdicio_pct, 2)
            if weights.desperdicio_pct is not None else None
        ),
    }


def build_cpis_workbook(
    date_from: str | None,
    date_to: str | None,
    operador: str | None = None,
    sector: str | None = None,
    validated_only: bool = False,
) -> bytes:
    """Return .xlsx bytes matching MigracaoNikufraCPIS.xlsx schema.

    Single sheet (`Folha1`) with the CPIS column layout. One row per kanban
    row in the period. Excel-native types: dates as date objects, numerics
    as numbers.

    R69: dates may be None (= "sempre"); sector filter via canonical
    ``detect_template`` post-fetch.
    R130: ``validated_only=True`` exclui folhas em rascunho (status != 'validated').

    Empty-period: still returns a valid file with just the header row, so
    the user can confirm column layout.
    """
    raw_rows = _query_cpis_rows(date_from, date_to, operador, sector, validated_only)
    # R96 — load refs once so _build_cpis_row can derive OV from plan
    # when the row doesn't have it (Gemini templates).
    try:
        from app.cross_check import get_watcher
        refs = get_watcher().get_refs()
    except Exception:
        refs = None
    cpis_rows = [_build_cpis_row(r, refs=refs) for r in raw_rows]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Folha1"

    # Header row (matches template exactly)
    for ci, (_, label) in enumerate(CPIS_COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=label)
        _style_cell(cell, bold=True, fill=_FILL_HEADER)
        cell.font = _FONT_HEADER

    # Data rows
    for ri, cpis in enumerate(cpis_rows, start=2):
        for ci, (key, _) in enumerate(CPIS_COLUMNS, start=1):
            v = cpis.get(key)
            cell = ws.cell(row=ri, column=ci, value=v)
            _style_cell(cell)
            if key == "data" and isinstance(v, dt.date):
                cell.number_format = "DD-MM-YYYY"
            elif key in ("peso_consumido_t", "peso_produzido_t", "desperdicio_t"):
                cell.number_format = "0.000"   # toneladas com 3 casas
            elif key == "desperdicio_pct":
                cell.number_format = "0.00"
            elif key == "esp_mm":
                cell.number_format = "0.0"
            elif key == "n_chapas":
                cell.number_format = "0"       # integer

    # Column widths — one entry per CPIS column.
    widths = [
        12, 16, 24, 24, 12, 10, 10, 20, 24, 8, 12, 10,
        14, 10, 10, 16, 14, 12, 10, 10, 16, 16, 14, 12,
    ]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def cpis_filename_for(
    date_from: str | None,
    date_to: str | None,
    operador: str | None = None,
    sector: str | None = None,
    validated_only: bool = False,
) -> str:
    """Slugged filename for the CPIS download. R69: "sempre" + sector.
    R130: sufixo `_validadas` quando o filtro de validação está activo."""
    period = (
        f"{date_from}_{date_to}"
        if (date_from and date_to)
        else "sempre"
    )
    sector_suffix = f"_{_slugify(sector)}" if sector else ""
    op_suffix = ""
    if operador:
        slug = "".join(ch for ch in operador.upper().replace(" ", "") if ch.isalnum())
        op_suffix = f"_{slug}"
    validated_suffix = "_validadas" if validated_only else ""
    return f"MigracaoNikufraCPIS_{period}{sector_suffix}{op_suffix}{validated_suffix}.xlsx"
