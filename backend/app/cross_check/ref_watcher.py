"""RefWatcher — live mtime-based reload of SAP + plan_colunas refs.

Watches ``C:\\kanban\\nifruka\\04_Documentacao\\StockSAP.xlsx`` and
``plan_colunas_cpis.xlsx``. When either's mtime changes, re-mines the
data into in-memory structures (and persists to ``lexicons/sap_plan_mined.json``
for backward-compat with scripts/predict_validation.py).

Cheap detection: ``Path.stat().st_mtime`` is ~microseconds. Called from
the hot path (/upload, /edit) on every request. Reload itself takes
~2-3s for ~10k OFs; only fires when files actually change.

API:
    watcher = get_watcher()           # singleton
    refs = watcher.get_refs()         # dict with all keys (loads if needed)
    watcher.force_reload()            # for /admin/reload-refs

Refs dict shape (matches old ``sap_plan_mined.json``):
    {
        "lotes_sap": frozenset[str],
        "lotes_sap_full": dict[str, dict],     # lote → {qtd, esp, larg, desc}
        "ofs_plan_str": frozenset[str],         # OF strings
        "of_to_entries": dict[str, list[dict]], # OF → plan entries
        "of_to_ovs": dict[str, frozenset[str]], # OF → set of OVs
        "of_to_designacoes": dict[str, tuple],  # OF → designacao strings (upper)
        "clientes_plan": frozenset[str],
        "loaded_at": ISO timestamp,
        "sap_mtime": float,
        "plan_mtime": float,
        "stats": {n_lotes, n_ofs, n_plan_rows, n_clientes},
    }
"""
from __future__ import annotations

import json
import os
import re
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from app.cross_check.of_utils import normalize_of

# R64 — overridable via KANBAN_DOC_DIR env var. Lets the laptop runtime
# point at a local refs folder (or `kanban_refs/04_Documentacao` inside
# the repo) instead of the hardcoded desktop path.
_DEFAULT_DOC_DIR = Path(
    os.environ.get("KANBAN_DOC_DIR", r"C:\kanban\nifruka\04_Documentacao")
)
_REPO_ROOT = Path(__file__).resolve().parents[3]

# Round 35 — SAP exports lotes inconsistently: some rows have a leading
# letter prefix (M for Metalogalva) and others don't. Operators always
# write the M prefix on kanbans. We index each lote under both forms.
_LOTE_PATTERN = re.compile(r"^([A-Z])?(\d{2}B\d+)$")


def _empty_refs() -> dict[str, Any]:
    """Empty refs structure used when source files missing (server still
    works, just nothing to validate against)."""
    return {
        "lotes_sap": frozenset(),
        "lotes_sap_full": {},
        "ofs_plan_str": frozenset(),
        "of_to_entries": {},
        "of_to_ovs": {},
        "of_to_designacoes": {},
        "clientes_plan": frozenset(),
        "ovs_plan": frozenset(),
        "modelos_plan": frozenset(),
        # R70 — SAP employee list keyed by cod (int 1-4 digits = pernr suffix)
        "colaboradores": {},
        # R85 — machine catalog (maquinas.xlsx). Maps kanban-written labels
        # (desigkanban: HPE32 / GUIFIL / LASER) to canonical codmaq (M024 /
        # M067 / M030) + sector (dessec: CORTE / QUINAGEM / SOLDADURA).
        "maquinas_by_kanban": {},
        "maquinas_by_codmaq": {},
        # R91 — operator aliases lexicon (OCR-corrupt name → canonical
        # cod/pernr/sname). Updated when an operator manually corrects a
        # header.operador field via the UI; next OCR of the same corrupt
        # name resolves directly via this lexicon.
        "operador_aliases": {},
        "loaded_at": None,
        "sap_mtime": 0.0,
        "plan_mtime": 0.0,
        "colab_mtime": 0.0,
        "maquinas_mtime": 0.0,
        "aliases_mtime": 0.0,
        "stats": {"n_lotes": 0, "n_ofs": 0, "n_plan_rows": 0, "n_clientes": 0, "n_colaboradores": 0, "n_maquinas": 0, "n_ofs_file": 0, "n_lotes_file": 0},
        "available": False,
    }


def _mine_colaboradores(colab_path: Path) -> dict[int, dict[str, str]]:
    """R70 — load ListaColaboradores.xlsx into ``cod (int) → {pernr, sname}``.

    Header is ``pernr | sname | cod``. sname is UPPERCASE ASCII (no acentos).
    Returns empty dict if file missing or malformed.
    """
    import openpyxl

    if not colab_path.exists():
        return {}
    out: dict[int, dict[str, str]] = {}
    wb = openpyxl.load_workbook(colab_path, read_only=True, data_only=True)
    ws = wb["Export"] if "Export" in wb.sheetnames else wb.active
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if r is None or r[0] is None:
            continue
        try:
            cod = int(r[2]) if r[2] is not None else None
        except (TypeError, ValueError):
            continue
        if cod is None:
            continue
        pernr = str(r[0]).strip()
        sname = str(r[1] or "").strip().upper()
        if not sname:
            continue
        # First entry wins on cod collisions (should not happen — cod is PK)
        out.setdefault(cod, {"pernr": pernr, "sname": sname})
    wb.close()
    return out


def _phase_columns(hdrs: dict[str, int]) -> list[str]:
    """R106 — the phase columns of plan_colunas: every column between
    ``quanttrp`` and ``esp`` (esp exclusive), in sheet order. On the current
    export that is ``bf, c, q, s, r, a, exp``."""
    qi, ei = hdrs.get("quanttrp"), hdrs.get("esp")
    if qi is None or ei is None:
        return []
    return sorted((name for name, idx in hdrs.items() if qi < idx < ei),
                  key=lambda n: hdrs[n])


def _to_num(v: object) -> float:
    """Lenient numeric coercion — blanks / non-numbers become 0."""
    try:
        return float(v)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return 0.0


def _fase_incompleta(quanttrp: object, fases: dict[str, object]) -> bool:
    """R106 — a plan row is *concluded* when every phase equals ``quanttrp``.
    Returns True (still in production → priority) when some phase differs."""
    q = _to_num(quanttrp)
    if q <= 0:
        return False
    return any(_to_num(v) != q for v in fases.values())


def _derive_plan_indexes(of_to_entries: dict[str, list[dict]]) -> dict[str, Any]:
    """Build every plan-derived index + count from an ``OF → entries`` map.
    Pure function — no I/O."""
    of_to_ovs = {
        of_str: frozenset(e["ov"] for e in entries if e.get("ov"))
        for of_str, entries in of_to_entries.items()
    }
    of_to_designacoes = {
        of_str: tuple(e["designacao"].upper().strip()
                      for e in entries if e.get("designacao"))
        for of_str, entries in of_to_entries.items()
    }
    # R83 — global sets + inverted indexes for OF-independent / holistic
    # plan-row search when the OCR's OF doesn't bate plan.
    clientes: set[str] = set()
    ovs_global: set[str] = set()
    modelo_fts: set[str] = set()
    cli_idx: dict[str, list[dict]] = {}
    mod_idx: dict[str, list[dict]] = {}
    n_rows = 0
    for of_str, entries in of_to_entries.items():
        for e in entries:
            n_rows += 1
            cli = (e.get("cliente") or "").strip().upper()
            if cli:
                clientes.add(cli)
            ov_v = (e.get("ov") or "").strip()
            if ov_v:
                ovs_global.add(ov_v)
            desig = (e.get("designacao") or "").strip()
            ft = desig.split(" - ", 1)[0].strip().upper() if desig else ""
            if ft:
                modelo_fts.add(ft)
            e_with_of = {**e, "_of": of_str}
            if cli:
                cli_idx.setdefault(cli, []).append(e_with_of)
            if ft:
                mod_idx.setdefault(ft, []).append(e_with_of)
    return {
        "ofs_plan_str": frozenset(of_to_entries.keys()),
        "of_to_entries": of_to_entries,
        "of_to_ovs": of_to_ovs,
        "of_to_designacoes": of_to_designacoes,
        "clientes_plan": frozenset(clientes),
        "ovs_plan": frozenset(ovs_global),
        "modelos_plan": frozenset(modelo_fts),
        "plan_by_cliente": cli_idx,
        "plan_by_modelo_ft": mod_idx,
        "n_ofs": len(of_to_entries),
        "n_clientes": len(clientes),
        "n_ovs": len(ovs_global),
        "n_modelos_fts": len(modelo_fts),
        "n_plan_rows": n_rows,
    }


def _mine_from_excel(
    sap_path: Path,
    plan_path: Path,
    colab_path: Path | None = None,
) -> dict[str, Any]:
    """Re-mine StockSAP + plan_colunas + ListaColaboradores Excel files into
    in-memory refs. Mirror of scripts/mine_sap_plan.py logic (kept inline
    to avoid cross-script dependency)."""
    import openpyxl

    refs = _empty_refs()

    # ---- StockSAP: column0=Lote, col1=Qtd, col2=Espessura, col3=Largura, col4=Desc
    stock_full: dict[str, dict] = {}
    if sap_path.exists():
        wb = openpyxl.load_workbook(sap_path, read_only=True, data_only=True)
        ws = wb["Folha1"] if "Folha1" in wb.sheetnames else wb.active
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if r[0] is None:
                break
            lote = str(r[0]).strip().upper()
            if not lote:
                continue
            entry = {
                "qtd": r[1],
                "esp": r[2],
                "larg": r[3],
                "desc": r[4] if len(r) > 4 else None,
            }
            stock_full[lote] = entry
            # Round 35 — alias under the other prefix variant. setdefault
            # so a real lote row never overwrites another's alias.
            m = _LOTE_PATTERN.match(lote)
            if m:
                prefix, suffix = m.groups()
                if prefix:
                    stock_full.setdefault(suffix, entry)        # add no-prefix alias
                else:
                    stock_full.setdefault("M" + lote, entry)    # add M-prefix alias
        wb.close()
        refs["stats"]["n_lotes_file"] = len(stock_full)
        refs["sap_mtime"] = sap_path.stat().st_mtime

    # ---- plan_colunas: cliente, ov, of, designacao, quanttrp, bf, esp, lbase, ltopo, ltotal, comp, ...
    of_to_entries: dict[str, list[dict]] = {}
    if plan_path.exists():
        wb = openpyxl.load_workbook(plan_path, read_only=True, data_only=True)
        ws = wb["plan_colunas_cpis"] if "plan_colunas_cpis" in wb.sheetnames else wb.active
        hdrs: dict[str, int] = {}
        phase_cols: list[str] = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                for j, h in enumerate(r):
                    if h:
                        hdrs[str(h).strip().lower()] = j
                phase_cols = _phase_columns(hdrs)
                continue
            if r[0] is None:
                break
            try:
                of_int = int(float(str(r[hdrs["of"]])))
            except (ValueError, TypeError, KeyError):
                continue
            cliente = str(r[hdrs.get("cliente", -1)] or "").strip().upper()
            ov = str(r[hdrs.get("ov", -1)] or "").strip()
            desig = str(r[hdrs.get("designacao", -1)] or "").strip()

            # Round 43 Sol 2: track `fechado` flag (0=active, 1=closed)
            fechado_raw = r[hdrs.get("fechado", -1)] if "fechado" in hdrs else None
            fechado_flag = "1" if str(fechado_raw or "").strip() in ("1", "1.0", "True") else "0"

            # R106 — quanttrp (final qty) + phase columns. A row is "concluded"
            # when every phase == quanttrp; rows that aren't get cross-check
            # priority (still in production).
            quanttrp = r[hdrs["quanttrp"]] if "quanttrp" in hdrs else None
            fases = {name: r[hdrs[name]] for name in phase_cols}

            entry = {
                # R106 — OF is always a 6-digit string.
                "of": normalize_of(of_int),
                "cliente": cliente,
                "ov": ov,
                "designacao": desig,
                "esp": r[hdrs["esp"]] if "esp" in hdrs else None,
                "lbase": r[hdrs["lbase"]] if "lbase" in hdrs else None,
                "ltopo": r[hdrs["ltopo"]] if "ltopo" in hdrs else None,
                "ltotal": r[hdrs["ltotal"]] if "ltotal" in hdrs else None,
                "comp": r[hdrs["comp"]] if "comp" in hdrs else None,
                # Round 38: load npecas (canonical pieces-per-bobine count)
                # + material for weight + material consistency checks.
                "npecas": r[hdrs["npecas"]] if "npecas" in hdrs else None,
                "material": str(r[hdrs["material"]] or "").strip() if "material" in hdrs else None,
                # Round 43: closed-OF flag, used for snap disambiguation
                "fechado": fechado_flag,
                # R106: phase tracking
                "quanttrp": quanttrp,
                "fases": fases,
                "fase_incompleta": _fase_incompleta(quanttrp, fases),
            }
            of_to_entries.setdefault(normalize_of(of_int), []).append(entry)
        wb.close()
        refs["stats"]["n_ofs_file"] = len(of_to_entries)
        refs["plan_mtime"] = plan_path.stat().st_mtime

    # R106 — refs come straight from the current Excel files (the R104
    # cumulative historical merge was removed).
    refs["lotes_sap_full"] = stock_full
    refs["lotes_sap"] = frozenset(stock_full.keys())
    _plan_idx = _derive_plan_indexes(of_to_entries)
    for _k in ("ofs_plan_str", "of_to_entries", "of_to_ovs",
               "of_to_designacoes", "clientes_plan", "ovs_plan",
               "modelos_plan", "plan_by_cliente", "plan_by_modelo_ft"):
        refs[_k] = _plan_idx[_k]
    refs["stats"]["n_plan_rows"] = _plan_idx["n_plan_rows"]
    refs["stats"]["n_ofs"] = _plan_idx["n_ofs"]
    refs["stats"]["n_clientes"] = _plan_idx["n_clientes"]
    refs["stats"]["n_ovs"] = _plan_idx["n_ovs"]
    refs["stats"]["n_modelos_fts"] = _plan_idx["n_modelos_fts"]

    # R70 — ListaColaboradores.xlsx (pernr/sname/cod) for operator snap
    if colab_path and colab_path.exists():
        colabs = _mine_colaboradores(colab_path)
        if colabs:
            refs["colaboradores"] = colabs
            refs["colab_mtime"] = colab_path.stat().st_mtime
            refs["stats"]["n_colaboradores"] = len(colabs)

    # R85 — maquinas.xlsx (cod / desmaq / desigkanban / sector). Lives in
    # the same docs dir as the other Excel refs. Optional: if missing,
    # cod_maquina auto-fill simply skips.
    maq_path = colab_path.parent / "maquinas.xlsx" if colab_path else None
    if maq_path and maq_path.exists():
        wb = openpyxl.load_workbook(maq_path, read_only=True, data_only=True)
        ws = wb.active
        by_kanban: dict[str, dict] = {}
        by_codmaq: dict[str, dict] = {}
        hdrs: dict[str, int] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                for j, h in enumerate(row):
                    if h:
                        hdrs[str(h).strip().lower()] = j
                continue
            if row[0] is None:
                continue
            rec = {
                "codmaq": str(row[hdrs["codmaq"]] or "").strip().upper(),
                "desmaq": str(row[hdrs.get("desmaq", -1)] or "").strip(),
                "desigkanban": (
                    str(row[hdrs["desigkanban"]] or "").strip()
                    if "desigkanban" in hdrs else ""
                ),
                "codsec": str(row[hdrs.get("codsec", -1)] or "").strip(),
                "dessec": str(row[hdrs.get("dessec", -1)] or "").strip(),
                "colunaexcel": str(row[hdrs.get("colunaexcel", -1)] or "").strip(),
            }
            if rec["codmaq"]:
                by_codmaq[rec["codmaq"]] = rec
            if rec["desigkanban"]:
                by_kanban[rec["desigkanban"].upper()] = rec
        wb.close()
        refs["maquinas_by_kanban"] = by_kanban
        refs["maquinas_by_codmaq"] = by_codmaq
        refs["maquinas_mtime"] = maq_path.stat().st_mtime
        refs["stats"]["n_maquinas"] = len(by_codmaq)

    # R91 — operador aliases lexicon (memorized OCR-corrupt → canonical
    # snaps). Empty dict if file missing/invalid. Path: <repo>/lexicons/
    # operador_aliases.json. Hot-reload happens when file mtime changes.
    aliases_path = _REPO_ROOT / "lexicons" / "operador_aliases.json"
    if aliases_path.exists():
        try:
            refs["operador_aliases"] = json.loads(
                aliases_path.read_text(encoding="utf-8")
            )
        except (json.JSONDecodeError, OSError):
            refs["operador_aliases"] = {}
        try:
            refs["aliases_mtime"] = aliases_path.stat().st_mtime
        except OSError:
            refs["aliases_mtime"] = 0.0
    else:
        refs["operador_aliases"] = {}
        refs["aliases_mtime"] = 0.0

    refs["stats"]["n_lotes"] = len(refs["lotes_sap"])
    refs["loaded_at"] = datetime.now(tz=timezone.utc).isoformat(timespec="seconds")
    refs["available"] = bool(
        refs["lotes_sap"] or refs["ofs_plan_str"] or refs["colaboradores"]
    )
    return refs


def _persist_refs_status(refs: dict, doc_dir: Path, *, prev_stats: dict | None = None) -> None:
    """Write _refs_status.json next to the Excel files for human visibility.
    Includes diff vs previous load (lotes added/removed, etc.)."""
    diff = {}
    if prev_stats:
        diff = {
            "lotes_added": refs["stats"]["n_lotes"] - prev_stats.get("n_lotes", 0),
            "ofs_added": refs["stats"]["n_ofs"] - prev_stats.get("n_ofs", 0),
            "clientes_added": refs["stats"]["n_clientes"] - prev_stats.get("n_clientes", 0),
        }
    status = {
        "loaded_at": refs["loaded_at"],
        "available": refs["available"],
        "stock_sap": {
            "mtime": (
                datetime.fromtimestamp(refs["sap_mtime"], tz=timezone.utc).isoformat(timespec="seconds")
                if refs["sap_mtime"] else None
            ),
            "n_lotes": refs["stats"]["n_lotes"],
        },
        "plan_colunas": {
            "mtime": (
                datetime.fromtimestamp(refs["plan_mtime"], tz=timezone.utc).isoformat(timespec="seconds")
                if refs["plan_mtime"] else None
            ),
            "n_ofs": refs["stats"]["n_ofs"],
            "n_rows": refs["stats"]["n_plan_rows"],
            "n_clientes": refs["stats"]["n_clientes"],
        },
        "diff_vs_previous": diff,
    }
    try:
        (doc_dir / "_refs_status.json").write_text(
            json.dumps(status, indent=2, ensure_ascii=False), encoding="utf-8"
        )
    except OSError:
        pass  # non-fatal — the in-memory refs still work


def _persist_legacy_mined(refs: dict, repo_root: Path) -> None:
    """Mirror the old ``lexicons/sap_plan_mined.json`` so scripts/predict_validation
    + scripts/resnap_round28 keep working without changes."""
    legacy_path = repo_root / "lexicons" / "sap_plan_mined.json"
    if not legacy_path.parent.exists():
        return
    legacy = {
        "lotes_sap": sorted(refs["lotes_sap"]),
        "lotes_sap_full": refs["lotes_sap_full"],
        "ofs_plan": sorted(int(x) for x in refs["ofs_plan_str"]),
        "of_to_entries": refs["of_to_entries"],
        "clientes_plan": sorted(refs["clientes_plan"]),
        "stats": {
            "n_lotes_sap": refs["stats"]["n_lotes"],
            "n_ofs_plan": refs["stats"]["n_ofs"],
            "n_plan_rows": refs["stats"]["n_plan_rows"],
            "n_clientes_plan": refs["stats"]["n_clientes"],
        },
    }
    try:
        legacy_path.write_text(
            json.dumps(legacy, indent=2, ensure_ascii=False), encoding="utf-8"
        )
    except OSError:
        pass


class RefWatcher:
    """mtime-based watcher. Thread-safe singleton via ``get_watcher()``."""

    def __init__(self, doc_dir: Path = _DEFAULT_DOC_DIR, repo_root: Path = _REPO_ROOT):
        self.doc_dir = Path(doc_dir)
        self.repo_root = Path(repo_root)
        self.sap_path = self.doc_dir / "StockSAP.xlsx"
        self.plan_path = self.doc_dir / "plan_colunas_cpis.xlsx"
        # R70 — ListaColaboradores.xlsx (SAP employee export)
        self.colab_path = self.doc_dir / "ListaColaboradores.xlsx"
        # R85 — maquinas.xlsx (cod / desmaq / desigkanban / sector catalog)
        self.maq_path = self.doc_dir / "maquinas.xlsx"
        self._refs: dict[str, Any] = _empty_refs()
        self._lock = threading.Lock()
        self._last_check_ts = 0.0
        self._check_interval = 1.0  # seconds — debounce mtime stat calls

    def _current_mtimes(self) -> tuple[float, float, float, float, float]:
        sap_m = self.sap_path.stat().st_mtime if self.sap_path.exists() else 0.0
        plan_m = self.plan_path.stat().st_mtime if self.plan_path.exists() else 0.0
        colab_m = self.colab_path.stat().st_mtime if self.colab_path.exists() else 0.0
        maq_m = self.maq_path.stat().st_mtime if self.maq_path.exists() else 0.0
        aliases_path = _REPO_ROOT / "lexicons" / "operador_aliases.json"
        aliases_m = aliases_path.stat().st_mtime if aliases_path.exists() else 0.0
        return sap_m, plan_m, colab_m, maq_m, aliases_m

    def _needs_reload(self) -> bool:
        sap_m, plan_m, colab_m, maq_m, aliases_m = self._current_mtimes()
        return (
            sap_m != self._refs.get("sap_mtime", 0.0)
            or plan_m != self._refs.get("plan_mtime", 0.0)
            or colab_m != self._refs.get("colab_mtime", 0.0)
            or maq_m != self._refs.get("maquinas_mtime", 0.0)
            or aliases_m != self._refs.get("aliases_mtime", 0.0)
        )

    def get_refs(self) -> dict[str, Any]:
        """Return current refs. Cheap (debounced mtime check ≤ 1/second).
        Reload if files changed."""
        now = time.monotonic()
        # Skip the stat call if we just checked
        if now - self._last_check_ts < self._check_interval and self._refs.get("available"):
            return self._refs
        with self._lock:
            self._last_check_ts = now
            if self._needs_reload():
                prev_stats = dict(self._refs.get("stats", {})) if self._refs.get("available") else None
                self._refs = _mine_from_excel(self.sap_path, self.plan_path, self.colab_path)
                _persist_refs_status(self._refs, self.doc_dir, prev_stats=prev_stats)
                _persist_legacy_mined(self._refs, self.repo_root)
        return self._refs

    def force_reload(self) -> dict[str, Any]:
        """Skip mtime check, force reload + persist. Used by /admin/reload-refs."""
        with self._lock:
            prev_stats = dict(self._refs.get("stats", {})) if self._refs.get("available") else None
            self._refs = _mine_from_excel(self.sap_path, self.plan_path, self.colab_path)
            _persist_refs_status(self._refs, self.doc_dir, prev_stats=prev_stats)
            _persist_legacy_mined(self._refs, self.repo_root)
            self._last_check_ts = time.monotonic()
        return self._refs

    def status(self) -> dict[str, Any]:
        """Lightweight status dict for /admin/refs-status."""
        refs = self.get_refs()
        return {
            "available": refs["available"],
            "loaded_at": refs["loaded_at"],
            "sap": {
                "path": str(self.sap_path),
                "exists": self.sap_path.exists(),
                "mtime": refs["sap_mtime"],
                "n_lotes": refs["stats"]["n_lotes"],
            },
            "plan": {
                "path": str(self.plan_path),
                "exists": self.plan_path.exists(),
                "mtime": refs["plan_mtime"],
                "n_ofs": refs["stats"]["n_ofs"],
                "n_rows": refs["stats"]["n_plan_rows"],
                "n_clientes": refs["stats"]["n_clientes"],
            },
        }


_singleton: RefWatcher | None = None


def get_watcher() -> RefWatcher:
    """Process-wide singleton."""
    global _singleton
    if _singleton is None:
        _singleton = RefWatcher()
    return _singleton
