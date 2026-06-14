"""Automatic importer for reference workbooks dropped in a shared folder.

The live OCR/cross-check refs still come from the canonical files watched by
``RefWatcher``. This module only copies validated workbooks from an external
folder into that canonical docs folder.
"""
from __future__ import annotations

import os
import shutil
import threading
import time
import traceback
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path, PureWindowsPath
from typing import Any

from app.config import _dotenv_value
from app.cross_check import refs_uploads
from app.cross_check.ref_watcher import RefWatcher, file_sha256, get_watcher
from app.pipeline.scoring_engine import normalize_of

_REPO_ROOT = Path(__file__).resolve().parents[3]

DEFAULT_IMPORT_DIR = Path(r"F:\ocr\files")
IMPORT_DIR_ENV_VARS = ("KANBAN_REFS_IMPORT_DIR", "OCR_REFS_IMPORT_DIR")
IMPORT_INTERVAL_ENV = "KANBAN_REFS_IMPORT_INTERVAL_SEC"
DEFAULT_IMPORT_INTERVAL_SECONDS = 15 * 60

REF_KINDS: dict[str, dict[str, str]] = {
    "plan": {
        "filename": "plan_colunas_cpis.xlsx",
        "watcher_attr": "plan_path",
        "sha_key": "plan_sha256",
        "count_key": "n_plan_rows",
    },
    "stocksap": {
        "filename": "StockSAP.xlsx",
        "watcher_attr": "sap_path",
        "sha_key": "sap_sha256",
        "count_key": "n_lotes",
    },
    "maquinas": {
        "filename": "maquinas.xlsx",
        "watcher_attr": "maq_path",
        "sha_key": "maquinas_sha256",
        "count_key": "n_maquinas",
    },
    "colaboradores": {
        "filename": "ListaColaboradores.xlsx",
        "watcher_attr": "colab_path",
        "sha_key": "colab_sha256",
        "count_key": "n_colaboradores",
    },
}


@dataclass(frozen=True)
class RefCandidate:
    kind: str
    path: Path
    info: dict[str, Any]
    sha256: str
    mtime: float
    size: int

    def as_dict(self) -> dict[str, Any]:
        out = {
            "kind": self.kind,
            "file": str(self.path),
            "filename": self.path.name,
            "sha256": self.sha256,
            "mtime": self.mtime,
            "size": self.size,
        }
        out.update(self.info)
        return out


@dataclass
class _ImportedRef:
    candidate: RefCandidate
    target: Path
    backup: Path | None
    target_existed: bool

    def as_dict(self) -> dict[str, Any]:
        out = self.candidate.as_dict()
        out["target"] = str(self.target)
        return out


_state_lock = threading.Lock()
_import_lock = threading.Lock()
_thread_lock = threading.Lock()
_thread: threading.Thread | None = None
_state: dict[str, Any] = {
    "enabled": False,
    "running": False,
    "source_dir": "",
    "interval_seconds": None,
    "last_run_at": None,
    "last_ok": None,
    "last_error": None,
    "last_result": None,
}


def _config_value(name: str) -> str | None:
    return os.environ.get(name) or _dotenv_value(_REPO_ROOT, name)


def _is_windows_absolute(raw: str | Path) -> bool:
    return PureWindowsPath(str(raw)).is_absolute()


def _resolve_path(raw: str | Path) -> Path:
    path = Path(raw)
    if path.is_absolute() or _is_windows_absolute(raw):
        return path
    return _REPO_ROOT / path


def configured_import_dir() -> Path:
    """Return the configured source folder.

    Production default is ``F:\\ocr\\files``. In dev, callers can override with
    ``KANBAN_REFS_IMPORT_DIR`` or ``OCR_REFS_IMPORT_DIR``.
    """
    for env_name in IMPORT_DIR_ENV_VARS:
        val = _config_value(env_name)
        if val:
            return _resolve_path(val)
    return DEFAULT_IMPORT_DIR


def configured_interval_seconds() -> int:
    raw = _config_value(IMPORT_INTERVAL_ENV)
    if not raw:
        return DEFAULT_IMPORT_INTERVAL_SECONDS
    try:
        return max(30, int(float(raw)))
    except (TypeError, ValueError):
        return DEFAULT_IMPORT_INTERVAL_SECONDS


def inspect_refs_xlsx(path: Path, kind: str) -> tuple[str | None, dict[str, Any]]:
    """Validate a refs workbook and return lightweight file stats."""
    import openpyxl

    try:
        size = Path(path).stat().st_size
    except OSError as e:
        return f"ficheiro não existe ou não é legível ({e})", {
            "n_rows": 0, "n_ofs": 0, "n_ovs": 0, "n_lotes": 0,
            "n_maquinas": 0, "n_colaboradores": 0, "size": 0,
        }

    info: dict[str, Any] = {
        "n_rows": 0, "n_ofs": 0, "n_ovs": 0, "n_lotes": 0,
        "n_maquinas": 0, "n_colaboradores": 0, "size": size,
    }
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return f"não consegui abrir o Excel ({e})", info
    try:
        if kind == "plan":
            ws = (
                wb["plan_colunas_cpis"]
                if "plan_colunas_cpis" in wb.sheetnames else wb.active
            )
            first = next(ws.iter_rows(values_only=True), None)
            hdrs = {
                str(h).strip().lower(): i
                for i, h in enumerate(first or ())
                if h
            }
            required = {"of", "ov", "quanttrp"}
            missing = sorted(required - set(hdrs))
            if missing:
                return (
                    "falta a coluna "
                    + ", ".join(repr(m) for m in missing)
                    + " — não parece o plan_colunas_cpis"
                ), info
            ofs: set[str] = set()
            ovs: set[str] = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                of_raw = row[hdrs["of"]] if hdrs["of"] < len(row) else None
                ov_raw = row[hdrs["ov"]] if hdrs["ov"] < len(row) else None
                if of_raw in (None, "") and ov_raw in (None, ""):
                    continue
                info["n_rows"] += 1
                of_norm = normalize_of(of_raw)
                if of_norm:
                    ofs.add(of_norm)
                ov = str(ov_raw or "").strip()
                if ov:
                    ovs.add(ov)
            info["n_ofs"] = len(ofs)
            info["n_ovs"] = len(ovs)
            info["n_plan_rows"] = info["n_rows"]
            if info["n_rows"] <= 0:
                return "sem linhas de plano — ficheiro vazio", info
            if info["n_ofs"] <= 0:
                return "sem OFs válidas — não parece o plan_colunas_cpis", info
        elif kind == "stocksap":
            rows = (
                wb["Folha1"] if "Folha1" in wb.sheetnames else wb.active
            ).iter_rows(values_only=True)
            header = next(rows, None) or ()
            col0 = str(header[0] or "").strip().lower() if header else ""
            if "lote" not in col0:
                return "1ª coluna não é 'Lote' — não parece o StockSAP", info
            for row in rows:
                if row and row[0] not in (None, ""):
                    info["n_rows"] += 1
            info["n_lotes"] = info["n_rows"]
            if info["n_rows"] <= 0:
                return "sem linhas de lote — não parece o StockSAP", info
        elif kind == "maquinas":
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            first = next(rows, None)
            hdrs = {
                str(h).strip().lower(): i
                for i, h in enumerate(first or ())
                if h
            }
            required = {"codmaq", "desmaq", "codsec", "dessec", "colunaexcel"}
            missing = sorted(required - set(hdrs))
            if missing:
                return (
                    "falta a coluna "
                    + ", ".join(repr(m) for m in missing)
                    + " — não parece o maquinas.xlsx"
                ), info
            for row in rows:
                if not row:
                    continue
                cod_idx = hdrs["codmaq"]
                codmaq = row[cod_idx] if cod_idx < len(row) else None
                if codmaq not in (None, ""):
                    info["n_rows"] += 1
            info["n_maquinas"] = info["n_rows"]
            if info["n_maquinas"] <= 0:
                return "sem máquinas válidas — não parece o maquinas.xlsx", info
        elif kind == "colaboradores":
            ws = wb["Export"] if "Export" in wb.sheetnames else wb.active
            rows = ws.iter_rows(values_only=True)
            first = next(rows, None)
            hdrs = {
                str(h).strip().lower(): i
                for i, h in enumerate(first or ())
                if h
            }
            required = {"pernr", "sname", "cod"}
            missing = sorted(required - set(hdrs))
            if missing:
                return (
                    "falta a coluna "
                    + ", ".join(repr(m) for m in missing)
                    + " — não parece a ListaColaboradores"
                ), info
            for row in rows:
                if not row:
                    continue
                cod_idx = hdrs["cod"]
                sname_idx = hdrs["sname"]
                cod = row[cod_idx] if cod_idx < len(row) else None
                sname = row[sname_idx] if sname_idx < len(row) else None
                if cod not in (None, "") and sname not in (None, ""):
                    info["n_rows"] += 1
            info["n_colaboradores"] = info["n_rows"]
            if info["n_colaboradores"] <= 0:
                return (
                    "sem colaboradores válidos — não parece a ListaColaboradores",
                    info,
                )
        else:
            return "tipo de refs inválido", info
    finally:
        wb.close()
    return None, info


def _filename_hint(path: Path) -> str | None:
    name = path.name.lower()
    if "stock" in name or "sap" in name:
        return "stocksap"
    if "plan" in name or "colunas" in name or "cpis" in name:
        return "plan"
    if "maquina" in name or "maq" in name:
        return "maquinas"
    if "colab" in name or "lista" in name or "operador" in name:
        return "colaboradores"
    return None


def _workbook_files(source_dir: Path) -> list[Path]:
    out: list[Path] = []
    for path in sorted(source_dir.iterdir(), key=lambda p: p.name.lower()):
        if not path.is_file():
            continue
        name = path.name
        if name.startswith("~$") or name.endswith(".tmp"):
            continue
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            continue
        out.append(path)
    return out


def scan_source_dir(source_dir: Path) -> dict[str, Any]:
    """Classify valid reference workbooks in ``source_dir`` by content."""
    candidates: list[RefCandidate] = []
    rejected: list[dict[str, Any]] = []
    files = _workbook_files(Path(source_dir))
    for path in files:
        valid: dict[str, dict[str, Any]] = {}
        errors: dict[str, str] = {}
        for kind in REF_KINDS:
            err, info = inspect_refs_xlsx(path, kind)
            if err is None:
                valid[kind] = info
            else:
                errors[kind] = err
        if not valid:
            rejected.append({
                "file": str(path),
                "filename": path.name,
                "errors": errors,
            })
            continue
        hint = _filename_hint(path)
        kind = hint if hint in valid else max(
            valid,
            key=lambda k: int(valid[k].get("n_rows") or 0),
        )
        st = path.stat()
        candidates.append(RefCandidate(
            kind=kind,
            path=path,
            info=valid[kind],
            sha256=file_sha256(path),
            mtime=st.st_mtime,
            size=st.st_size,
        ))
    return {
        "source_dir": str(source_dir),
        "scanned": len(files),
        "candidates": [c.as_dict() for c in candidates],
        "rejected": rejected,
    }


def _latest_by_kind(candidates: list[RefCandidate]) -> dict[str, RefCandidate]:
    selected: dict[str, RefCandidate] = {}
    for cand in candidates:
        current = selected.get(cand.kind)
        if current is None:
            selected[cand.kind] = cand
            continue
        if (cand.mtime, cand.size, cand.path.name) > (
            current.mtime, current.size, current.path.name
        ):
            selected[cand.kind] = cand
    return selected


def _target_for_kind(watcher: RefWatcher, kind: str) -> Path:
    return Path(getattr(watcher, REF_KINDS[kind]["watcher_attr"]))


def _copy_into_target(cand: RefCandidate, target: Path) -> _ImportedRef:
    target.parent.mkdir(parents=True, exist_ok=True)
    token = f"{os.getpid()}-{time.time_ns()}"
    tmp = target.with_name(f"{target.stem}.autoimport-tmp-{token}{target.suffix}")
    backup = target.with_name(
        f"{target.stem}.autoimport-prevbak-{token}{target.suffix}"
    )
    shutil.copy2(cand.path, tmp)
    err, _info = inspect_refs_xlsx(tmp, cand.kind)
    if err:
        tmp.unlink(missing_ok=True)
        raise RuntimeError(
            f"{cand.path.name}: ficheiro rejeitado depois da cópia ({err})"
        )

    target_existed = target.exists()
    backup_path: Path | None = None
    if target_existed:
        shutil.copy2(target, backup)
        backup_path = backup

    replaced = False
    try:
        for _ in range(5):
            try:
                os.replace(tmp, target)
                replaced = True
                break
            except PermissionError:
                time.sleep(0.3)
        if not replaced:
            raise PermissionError("ficheiro ativo em uso")
        active_sha = file_sha256(target)
        if active_sha != cand.sha256:
            raise RuntimeError("ficheiro ativo ficou com hash diferente da origem")
        return _ImportedRef(
            candidate=cand,
            target=target,
            backup=backup_path,
            target_existed=target_existed,
        )
    except Exception:
        tmp.unlink(missing_ok=True)
        if backup_path is not None and backup_path.exists():
            os.replace(backup_path, target)
        elif not target_existed:
            target.unlink(missing_ok=True)
        raise


def _restore_imports(imported: list[_ImportedRef]) -> None:
    for item in imported:
        try:
            if item.backup is not None and item.backup.exists():
                os.replace(item.backup, item.target)
            elif not item.target_existed:
                item.target.unlink(missing_ok=True)
        except Exception:  # noqa: BLE001
            traceback.print_exc()


def _discard_backups(imported: list[_ImportedRef]) -> None:
    for item in imported:
        if item.backup is not None:
            item.backup.unlink(missing_ok=True)


def _invalidate_refs_dependents() -> None:
    try:
        from app.pipeline.obras_status import invalidate_cache as obras_inv
        obras_inv()
    except Exception:  # noqa: BLE001
        pass
    try:
        from app.pipeline.of_consumption import invalidate_cache as of_inv
        of_inv()
    except Exception:  # noqa: BLE001
        pass


def _record_imports(imported: list[_ImportedRef], refs: dict[str, Any]) -> None:
    stats = refs.get("stats", {}) or {}
    for item in imported:
        kind = item.candidate.kind
        info = item.candidate.info
        count_key = REF_KINDS[kind]["count_key"]
        n_rows = int(
            stats.get(count_key) or info.get(count_key) or info.get("n_rows") or 0
        )
        try:
            refs_uploads.record(
                kind,
                item.candidate.path.name,
                n_rows,
                sha256=item.candidate.sha256,
                n_ofs=(
                    stats.get("n_ofs") or info.get("n_ofs")
                    if kind == "plan" else None
                ),
                n_ovs=(
                    stats.get("n_ovs") or info.get("n_ovs")
                    if kind == "plan" else None
                ),
                size=item.target.stat().st_size,
            )
        except Exception:  # noqa: BLE001
            traceback.print_exc()


def _state_result(result: dict[str, Any]) -> dict[str, Any]:
    compact = dict(result)
    if len(compact.get("rejected") or []) > 20:
        compact["rejected"] = compact["rejected"][:20]
        compact["rejected_truncated"] = True
    if len(compact.get("candidates") or []) > 20:
        compact["candidates"] = compact["candidates"][:20]
        compact["candidates_truncated"] = True
    return compact


def _remember_result(result: dict[str, Any]) -> dict[str, Any]:
    with _state_lock:
        _state["last_ok"] = bool(result.get("ok"))
        _state["last_error"] = result.get("error") or (
            "; ".join(
                str(e.get("error"))
                for e in result.get("errors", [])
                if e.get("error")
            )
            or None
        )
        _state["last_result"] = _state_result(result)
    return result


def import_refs_from_dir(
    source_dir: Path,
    *,
    watcher: RefWatcher | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Import the newest valid workbook of each ref kind from ``source_dir``."""
    if not _import_lock.acquire(blocking=False):
        return {"ok": False, "error": "importação de refs já está em curso"}
    with _state_lock:
        _state["running"] = True
        _state["last_run_at"] = datetime.now(tz=timezone.utc).isoformat(
            timespec="seconds"
        )
        _state["source_dir"] = str(source_dir)
    try:
        source = Path(source_dir)
        if not source.exists() or not source.is_dir():
            result = {
                "ok": False,
                "source_dir": str(source),
                "error": "pasta de importação não existe",
                "imported": [],
                "skipped": [],
                "errors": [],
                "rejected": [],
                "candidates": [],
            }
            return _remember_result(result)

        scan = scan_source_dir(source)
        candidates = [
            RefCandidate(
                kind=c["kind"],
                path=Path(c["file"]),
                info={k: v for k, v in c.items() if k.startswith("n_") or k == "size"},
                sha256=c["sha256"],
                mtime=float(c["mtime"]),
                size=int(c["size"]),
            )
            for c in scan["candidates"]
        ]
        selected = _latest_by_kind(candidates)
        watcher = watcher or get_watcher()
        skipped: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []
        to_import: list[tuple[RefCandidate, Path]] = []
        for kind, cand in sorted(selected.items()):
            target = _target_for_kind(watcher, kind)
            if target.exists() and file_sha256(target) == cand.sha256:
                skipped.append({
                    "kind": kind,
                    "file": str(cand.path),
                    "filename": cand.path.name,
                    "target": str(target),
                    "reason": "igual ao ficheiro ativo",
                    "sha256": cand.sha256,
                })
                continue
            to_import.append((cand, target))

        imported: list[_ImportedRef] = []
        if dry_run:
            result = {
                "ok": True,
                "dry_run": True,
                "source_dir": str(source),
                "scanned": scan["scanned"],
                "candidates": scan["candidates"],
                "rejected": scan["rejected"],
                "selected": [c.as_dict() for c, _target in to_import],
                "skipped": skipped,
                "imported": [],
                "errors": [],
            }
            return _remember_result(result)

        for cand, target in to_import:
            try:
                imported.append(_copy_into_target(cand, target))
            except Exception as e:  # noqa: BLE001
                errors.append({
                    "kind": cand.kind,
                    "file": str(cand.path),
                    "filename": cand.path.name,
                    "target": str(target),
                    "error": str(e),
                })

        refs: dict[str, Any] = {}
        if errors and imported:
            _restore_imports(imported)
            watcher.force_reload()
            imported = []
        elif imported:
            refs = watcher.force_reload()
            mismatched = [
                item for item in imported
                if refs.get(REF_KINDS[item.candidate.kind]["sha_key"])
                != item.candidate.sha256
            ]
            if mismatched:
                _restore_imports(imported)
                refs = watcher.force_reload()
                errors.append({
                    "error": (
                        "refs carregadas não batem o hash dos ficheiros "
                        "importados; rollback aplicado"
                    ),
                    "kinds": [item.candidate.kind for item in mismatched],
                })
                imported = []
            else:
                _discard_backups(imported)
                _invalidate_refs_dependents()
                _record_imports(imported, refs)

        result = {
            "ok": not errors,
            "source_dir": str(source),
            "scanned": scan["scanned"],
            "candidates": scan["candidates"],
            "rejected": scan["rejected"],
            "skipped": skipped,
            "imported": [item.as_dict() for item in imported],
            "errors": errors,
            "refs_loaded_at": refs.get("loaded_at") if refs else None,
        }
        return _remember_result(result)
    finally:
        with _state_lock:
            _state["running"] = False
        _import_lock.release()


def import_refs_from_config(*, watcher: RefWatcher | None = None) -> dict[str, Any]:
    return import_refs_from_dir(configured_import_dir(), watcher=watcher)


def _run_loop(source_dir: Path, interval_seconds: int) -> None:
    while True:
        try:
            result = import_refs_from_dir(source_dir)
        except Exception as e:  # noqa: BLE001
            traceback.print_exc()
            result = {
                "ok": False,
                "source_dir": str(source_dir),
                "error": str(e),
                "imported": [],
                "skipped": [],
                "errors": [{"error": str(e)}],
            }
        _remember_result(result)
        time.sleep(interval_seconds)


def start_background_importer(
    *,
    source_dir: Path | None = None,
    interval_seconds: int | None = None,
) -> bool:
    """Start the polling importer for an available or absolute source path."""
    global _thread
    source = Path(source_dir) if source_dir is not None else configured_import_dir()
    interval = interval_seconds or configured_interval_seconds()
    with _thread_lock:
        if _thread is not None and _thread.is_alive():
            return True
        with _state_lock:
            _state.update({
                "enabled": False,
                "source_dir": str(source),
                "interval_seconds": interval,
            })
        if (not source.exists() or not source.is_dir()) and not source.is_absolute():
            with _state_lock:
                _state["last_error"] = "pasta de importação não existe"
            return False
        with _state_lock:
            _state.update({
                "enabled": True,
                "source_dir": str(source),
                "interval_seconds": interval,
                "last_error": (
                    None if source.exists() else "pasta de importação não existe"
                ),
            })
        _thread = threading.Thread(
            target=_run_loop,
            args=(source, interval),
            name="refs-importer",
            daemon=True,
        )
        _thread.start()
        return True


def status() -> dict[str, Any]:
    with _state_lock:
        out = dict(_state)
    out["thread_alive"] = _thread is not None and _thread.is_alive()
    return out
