"""rev00 (13/04/2026) — testes do novo formato de kanbans.

Cobre: coluna SUCATA (registry/prompt/CSV/schema), TURNO no header por defeito,
novos aliases de máquina + endurecimento do MTG-token, template genérico
`paragens`, routing frente/verso por pista de página + flag de revisão, e o
filtro de paragens do downtime dirigido pelo registry.
"""
from __future__ import annotations

import io
import json
import re
from pathlib import Path

import openpyxl
import pytest
from app.templates_registry import (
    TEMPLATES,
    detect_template,
    get_template,
)
from app.web import ocr_runner

# ---------------------------------------------------------------------------
# A — Registry: SUCATA + TURNO + aliases + paragens genérico
# ---------------------------------------------------------------------------

_PRODUCTION = [t for t in TEMPLATES.values() if t.has_production_rows]


@pytest.mark.parametrize("tpl", _PRODUCTION, ids=lambda t: t.name)
def test_production_templates_have_turno_in_header(tpl):
    assert "turno" in tpl.header_fields


@pytest.mark.parametrize(
    "tpl",
    [t for t in _PRODUCTION if t.name not in ("acabamento", "bobine_formato")],
    ids=lambda t: t.name,
)
def test_production_templates_end_with_sucata(tpl):
    # SUCATA é sempre a última coluna no papel rev00.
    assert tpl.row_fields[-1] == "sucata"


def test_acabamento_untouched_no_sucata():
    # TPL086 é família diferente; não entrou no lote rev00.
    assert "sucata" not in get_template("acabamento").row_fields


def test_bobine_v3_replaces_sucata_with_fecho():
    tpl = get_template("bobine_formato")
    assert tpl.row_fields[-1] == "fecho"
    assert "sucata" not in tpl.row_fields
    legacy = get_template("bobine_formato_legacy")
    assert legacy.row_fields[-1] == "sucata"
    assert legacy.setor_aliases == ()
    assert legacy.internal is True


def test_generic_paragens_registered():
    p = get_template("paragens")
    assert p.name == "paragens"
    assert p.has_production_rows is False
    assert p.row_fields == ("motivo", "inicio", "fim", "duracao", "resolvido")
    assert p.cross_check_fields == ()
    assert p.footer_fields == ()
    assert "turno" in p.header_fields
    assert "setor_maquina" in p.header_fields  # mantém identidade da máquina


def test_paragens_never_auto_detected():
    # setor_aliases=() → nunca sai de detect_template; só via pista/side-detect.
    assert detect_template("MÁQUINA DE FUSTES").name != "paragens"
    assert detect_template("QUINADORA PAV.4").name != "paragens"


def test_maq_fustes_paragens_kept_for_legacy():
    # Folhas já persistidas continuam a resolver.
    assert get_template("maq_fustes_paragens").name == "maq_fustes_paragens"


# ---------------------------------------------------------------------------
# B — Detecção: novos rótulos rev00 + endurecimento do MTG-token
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    "label, expected",
    [
        ("ROBOT MTG2", "robot"),
        ("LASER MTG2", "laser"),
        ("GUILHOTINA 6M", "guilhotina"),
        ("SOLDLINE 4", "soldline"),
        ("QUINADORA PAV.8", "quinadora_pav8"),
        ("MÁQUINA DE FUSTES", "maq_fustes"),
    ],
)
def test_rev00_labels_resolve(label, expected):
    assert detect_template(label).name == expected


def test_garbled_base_with_mtg_suffix_not_acabamento():
    # base mal lida + "MTG2" limpo NÃO deve cair em acabamento (regressão rev00).
    assert detect_template("R0B0T MTG2").name != "acabamento"
    assert detect_template("LAZER MTG2").name != "acabamento"


def test_bare_mtg_token_still_acabamento():
    # Um setor que é SÓ "MTG2" continua a ser acabamento (comportamento antigo).
    assert detect_template("MTG2").name == "acabamento"
    assert detect_template("ACABAMENTO MTG2").name == "acabamento"


# ---------------------------------------------------------------------------
# B — Prompt + schema + CSV: SUCATA
# ---------------------------------------------------------------------------

def test_prompt_columns_include_fecho_not_sucata():
    from app.pipeline.prompt_builder import build_prompt

    prompt = build_prompt(get_template("bobine_formato"))
    assert "FECHO" in prompt
    assert '"fecho":""' in prompt
    assert '"sucata":""' not in prompt


def test_paragens_prompt_uses_paragens_rules():
    from app.pipeline.prompt_builder import build_prompt

    prompt = build_prompt(get_template("paragens"))
    assert "MOTIVO DA PARAGEM" in prompt
    # não deve pedir o schema de produção (QTD/OF de 6 dígitos)
    assert "OF is EXACTLY 6 digits" not in prompt


def test_row_schema_has_sucata():
    from app.pipeline.inference.schemas import Row

    assert "sucata" in Row.model_fields
    assert Row(sucata="3").sucata == "3"
    assert Row(fecho="x").fecho == "X"


def test_default_ocr_normalizer_preserves_and_normalizes_fecho():
    import ocr6

    normalized = ocr6.normalize_extraction({
        "header": {"setor_maquina": "BOBINE-FORMATO", "turno": "M"},
        "rows": [{"of": "262107", "fecho": "x"}],
        "footer": {},
    })
    assert normalized["rows"][0]["fecho"] == "X"
    assert normalized["header"]["turno"] == "M"


def test_default_run_pipeline_keeps_fecho_from_real_ocr_path(monkeypatch):
    """Regression: default _run_ocr used to omit FECHO from ROW_FIELDS."""
    import ocr6

    payload = json.dumps({
        "header": {
            "operador": "ANA",
            "n_operador": "12",
            "setor_maquina": "BOBINE-FORMATO",
            "data": "27-07-2026",
            "turno": "M",
        },
        "rows": [{
            "ov": "2410001",
            "of": "262107",
            "modelo": "CFC5F45RIV",
            "comp_mm": "6000",
            "larg_mm": "250",
            "lote": "M26B0307",
            "fecho": "x",
        }],
        "footer": {"colunas_produzidas": "1", "horas_trabalhadas": "8"},
    })
    monkeypatch.setattr(ocr6, "image_to_base64", lambda *_a, **_k: "image")
    monkeypatch.setattr(
        ocr6,
        "ollama_request",
        lambda *_a, **_k: (
            payload,
            {"eval_count": 10, "eval_duration_ms": 10},
        ),
    )

    result = ocr_runner.run_pipeline(Path("bobine.jpg"), page_hint="F")

    assert result["template_name"] == "bobine_formato"
    assert result["raw"]["header"]["turno"] == "M"
    assert result["raw"]["rows"][0]["fecho"] == "X"
    assert result["current"]["rows"][0]["fecho"] == "X"


def test_strict_grammar_accepts_sucata():
    from app.pipeline.inference.schemas_strict import KANBAN_JSON_SCHEMA

    props = KANBAN_JSON_SCHEMA["properties"]["rows"]["items"]["properties"]
    assert "sucata" in props
    # não em required — outros templates continuam válidos
    assert "sucata" not in KANBAN_JSON_SCHEMA["properties"]["rows"]["items"]["required"]


@pytest.mark.parametrize(
    ("value", "valid"),
    [("", True), ("X", True), ("x", True), ("1", False), ("sim", False)],
)
def test_strict_grammar_fecho_is_blank_or_x(value, valid):
    from app.pipeline.inference.schemas_strict import KANBAN_JSON_SCHEMA

    pattern = (
        KANBAN_JSON_SCHEMA["properties"]["rows"]["items"]["properties"]
        ["fecho"]["pattern"]
    )
    assert bool(re.fullmatch(pattern, value)) is valid


def test_factory_csv_includes_sucata_column():
    from app.web.main import _to_3block_csv

    data = {
        "template_name": "bobine_formato_legacy",
        "header": {"operador": "X", "data": "10-05-2026", "setor_maquina": "BOBINE-FORMATO"},
        "rows": [{"of": "262107", "modelo": "OMEGA", "qtd": "5", "sucata": "2"}],
        "footer": {"colunas_produzidas": "5", "horas_trabalhadas": "8"},
    }
    csv_text = _to_3block_csv("kanban.jpg", data)
    assert "SUCATA" in csv_text
    # o valor 2 sai na linha
    assert ";2" in csv_text or ",2" in csv_text or "2\r" in csv_text


def test_factory_csv_new_bobine_uses_fecho_not_sucata():
    from app.web.main import _to_3block_csv

    data = {
        "template_name": "bobine_formato",
        "header": {"operador": "X", "data": "27-07-2026"},
        "rows": [{"of": "262107", "fecho": "X"}],
        "footer": {},
    }
    csv_text = _to_3block_csv("kanban.jpg", data)
    assert "FECHO" in csv_text
    assert "SUCATA" not in csv_text
    assert csv_text.count(";X") >= 1


def test_aggregate_xlsx_keeps_legacy_sucata_and_new_fecho_separate(monkeypatch):
    from app.web import export

    common = {
        "sheet_iso_date": "2026-07-27",
        "operador": "ANA",
        "validated_operador": "ANA",
        "of": "262107",
        "qtd": 1,
    }
    rows = [
        {**common, "sheet_id": 1, "sucata": 2, "fecho": None},
        {**common, "sheet_id": 2, "sucata": None, "fecho": "X"},
    ]
    monkeypatch.setattr(export, "_query_rows", lambda *_a, **_k: rows)

    workbook = openpyxl.load_workbook(io.BytesIO(export.export_excel(None, None)))
    day = workbook["27-07-2026"]
    headers = [day.cell(5, col).value for col in range(1, len(export.ROW_COLUMNS) + 1)]
    sucata_col = headers.index("SUCATA") + 1
    fecho_col = headers.index("FECHO") + 1

    assert day.cell(6, sucata_col).value == 2
    assert day.cell(6, fecho_col).value in (None, "")
    assert day.cell(7, sucata_col).value in (None, "")
    assert day.cell(7, fecho_col).value == "X"


# ---------------------------------------------------------------------------
# C — Historical compatibility migration/rebuild
# ---------------------------------------------------------------------------

def test_bobine_legacy_migration_is_idempotent_and_preserves_raw(
    tmp_path, monkeypatch,
):
    from app.web import db

    monkeypatch.setattr(db, "_DB_PATH", tmp_path / "migration.db")
    db.init_db()

    raw_payloads = {
        "explicit": (
            '{"template_name":"bobine_formato","header":{"setor_maquina":'
            '"BOBINE-FORMATO"},"rows":[{"sucata":"2"}],"footer":{}}'
        ),
        "inferred": (
            '{"header":{"setor_maquina":"BOBINE-FORMATO"},'
            '"rows":[{"sucata":"3"}],"footer":{}}'
        ),
        "other": (
            '{"header":{"setor_maquina":"GUILHOTINA"},'
            '"rows":[{"sucata":"1"}],"footer":{}}'
        ),
    }
    sheet_data = {
        "explicit": json.loads(raw_payloads["explicit"]),
        "inferred": json.loads(raw_payloads["inferred"]),
        "other": json.loads(raw_payloads["other"]),
    }
    ids: dict[str, int] = {}
    with db.conn() as c:
        for name in ("explicit", "inferred", "other"):
            cur = c.execute(
                "INSERT INTO sheets (image_path, status, raw_extraction, sheet_data) "
                "VALUES (?, 'extracted', ?, ?)",
                (
                    f"{name}.jpg",
                    raw_payloads[name],
                    json.dumps(sheet_data[name], ensure_ascii=False),
                ),
            )
            ids[name] = cur.lastrowid
        c.execute(
            "DELETE FROM app_migrations WHERE name = ?",
            ("bobine_formato_v3_fecho",),
        )

    db.init_db()
    after_first = {
        name: db.get_sheet(sid)["sheet_data"]
        for name, sid in ids.items()
    }
    with db.conn() as c:
        stored_raw = {
            name: c.execute(
                "SELECT raw_extraction FROM sheets WHERE id = ?", (sid,)
            ).fetchone()["raw_extraction"]
            for name, sid in ids.items()
        }
        stored_sheet_data = {
            name: c.execute(
                "SELECT sheet_data FROM sheets WHERE id = ?", (sid,)
            ).fetchone()["sheet_data"]
            for name, sid in ids.items()
        }

    assert after_first["explicit"]["template_name"] == "bobine_formato_legacy"
    assert after_first["inferred"]["template_name"] == "bobine_formato_legacy"
    assert "template_name" not in after_first["other"]
    assert stored_raw == raw_payloads

    db.init_db()
    with db.conn() as c:
        after_second = {
            name: c.execute(
                "SELECT sheet_data FROM sheets WHERE id = ?", (sid,)
            ).fetchone()["sheet_data"]
            for name, sid in ids.items()
        }
        marker_count = c.execute(
            "SELECT COUNT(*) AS n FROM app_migrations WHERE name = ?",
            ("bobine_formato_v3_fecho",),
        ).fetchone()["n"]
    assert after_second == stored_sheet_data
    assert marker_count == 1


def test_rebuild_from_immutable_raw_keeps_legacy_template(tmp_path, monkeypatch):
    from app.web import db, main

    monkeypatch.setattr(db, "_DB_PATH", tmp_path / "rebuild.db")
    db.init_db()
    sid = db.insert_sheet("legacy.jpg")
    raw = {
        "template_name": "bobine_formato",
        "header": {"setor_maquina": "BOBINE-FORMATO"},
        "rows": [{"of": "262107", "sucata": "2"}],
        "footer": {},
    }
    current = {**raw, "template_name": "bobine_formato_legacy"}
    db.update_extraction(sid, raw, {}, current)

    assert main._rebuild_sheet_data_from_raw(sid, db.get_sheet(sid)) is False
    # Force a material difference so the rebuild writes a new editable copy.
    db.apply_edit(sid, "rows[0].sucata", "3", source="system")
    assert main._rebuild_sheet_data_from_raw(sid, db.get_sheet(sid)) is True

    sheet = db.get_sheet(sid)
    assert sheet["raw_extraction"]["template_name"] == "bobine_formato"
    assert sheet["raw_extraction"]["rows"][0]["sucata"] == "2"
    assert sheet["sheet_data"]["template_name"] == "bobine_formato_legacy"
    assert sheet["sheet_data"]["rows"][0]["sucata"] == "2"


# ---------------------------------------------------------------------------
# E — Downtime: filtro dirigido pelo registry
# ---------------------------------------------------------------------------

def test_downtime_paragens_names_registry_driven():
    from app.downtime import _paragens_template_names

    names = _paragens_template_names()
    assert "paragens" in names
    assert "maq_fustes_paragens" in names
    # nenhum template de produção
    assert "bobine_formato" not in names
    assert "quinadora_pav4_paragens" not in names  # nome morto, já não existe


# ---------------------------------------------------------------------------
# C — run_pipeline: routing frente/verso (pista autoritativa / fast-path / flag)
# ---------------------------------------------------------------------------

def _mock_pipeline(monkeypatch, *, setor="GUILHOTINA", rows=None):
    """Isola a lógica de lado: _run_ocr e helpers devolvem fakes."""
    pass1 = {"header": {"setor_maquina": setor}, "rows": rows or []}
    monkeypatch.setattr(ocr_runner, "_run_ocr", lambda *a, **k: (pass1, None))
    monkeypatch.setattr(ocr_runner, "_merge_pass2_into_pass1", lambda p1, p2: p1)
    monkeypatch.setattr(ocr_runner, "_build_current_and_dq", lambda raw, tpl: ({}, {}))


def _detect_must_not_run(*_a, **_k):
    raise AssertionError("_detect_side não devia correr aqui")


def test_hint_verso_routes_to_paragens_without_mini_ocr(monkeypatch):
    _mock_pipeline(monkeypatch, rows=[])
    monkeypatch.setattr(ocr_runner, "_detect_side", _detect_must_not_run)
    res = ocr_runner.run_pipeline(Path("x.jpg"), page_hint="V")
    assert res["template_name"] == "paragens"
    assert res["side"] == "V"
    assert res["side_source"] == "hint"
    assert res["needs_review"] is False


def test_hint_frente_keeps_production_without_mini_ocr(monkeypatch):
    _mock_pipeline(monkeypatch, rows=[])
    monkeypatch.setattr(ocr_runner, "_detect_side", _detect_must_not_run)
    res = ocr_runner.run_pipeline(Path("x.jpg"), page_hint="F")
    assert res["template_name"] == "guilhotina"
    assert res["side"] == "F"
    assert res["side_source"] == "hint"


def test_persisted_legacy_hint_survives_reprocess(monkeypatch):
    calls: list[str | None] = []

    def fake_run(_path, template=None):
        calls.append(template.name if template is not None else None)
        if template is None:
            return (
                {"header": {"setor_maquina": "BOBINE-FORMATO"}, "rows": [], "footer": {}},
                None,
            )
        return (
            {
                "header": {"setor_maquina": "BOBINE-FORMATO"},
                "rows": [{"of": "262107", "sucata": "2"}],
                "footer": {},
            },
            None,
        )

    monkeypatch.setattr(ocr_runner, "_run_ocr", fake_run)
    monkeypatch.setattr(ocr_runner, "_detect_side", _detect_must_not_run)

    res = ocr_runner.run_pipeline(
        Path("legacy.jpg"),
        page_hint="F",
        template_name_hint="bobine_formato_legacy",
    )

    assert calls == [None, "bobine_formato_legacy"]
    assert res["template_name"] == "bobine_formato_legacy"
    assert res["template_detection"]["source"] == "persisted_template"
    assert res["current"]["rows"][0]["sucata"] == "2"
    assert "fecho" not in res["current"]["rows"][0]


def test_persisted_legacy_hint_still_routes_verso_to_paragens(monkeypatch):
    calls: list[str | None] = []

    def fake_run(_path, template=None):
        calls.append(template.name if template is not None else None)
        if template is None:
            return (
                {"header": {"setor_maquina": "BOBINE-FORMATO"}, "rows": [], "footer": {}},
                None,
            )
        return (
            {
                "header": {"setor_maquina": "BOBINE-FORMATO"},
                "rows": [{"motivo": "AVARIA", "inicio": "09:00"}],
                "footer": {},
            },
            None,
        )

    monkeypatch.setattr(ocr_runner, "_run_ocr", fake_run)
    monkeypatch.setattr(ocr_runner, "_detect_side", _detect_must_not_run)

    res = ocr_runner.run_pipeline(
        Path("legacy-verso.jpg"),
        page_hint="V",
        template_name_hint="bobine_formato_legacy",
    )

    assert calls == [None, "paragens"]
    assert res["template_name"] == "paragens"
    assert res["side"] == "V"
    assert res["current"]["rows"][0]["motivo"] == "AVARIA"


def test_confident_frente_skips_mini_ocr(monkeypatch):
    _mock_pipeline(
        monkeypatch,
        rows=[{"of": "262107", "modelo": "OMEGA"}, {"of": "262559", "modelo": "CGC2"}],
    )
    monkeypatch.setattr(ocr_runner, "_detect_side", _detect_must_not_run)
    res = ocr_runner.run_pipeline(Path("x.jpg"))  # sem pista
    assert res["template_name"] == "guilhotina"
    assert res["side"] == "F"
    assert res["side_source"] == "structure"


def test_no_hint_detect_verso_routes_to_paragens(monkeypatch):
    _mock_pipeline(monkeypatch, rows=[])  # sem sinal de frente
    monkeypatch.setattr(ocr_runner, "_detect_side", lambda *_a, **_k: "V")
    res = ocr_runner.run_pipeline(Path("x.jpg"))
    assert res["template_name"] == "paragens"
    assert res["side"] == "V"
    assert res["side_source"] == "detect"
    assert res["needs_review"] is False


def test_no_hint_indeterminate_flags_needs_review(monkeypatch):
    _mock_pipeline(monkeypatch, rows=[])
    monkeypatch.setattr(ocr_runner, "_detect_side", lambda *_a, **_k: "?")
    res = ocr_runner.run_pipeline(Path("x.jpg"))
    # extrai como frente mas fica marcada p/ revisão (não deposita produção)
    assert res["side"] == "F"
    assert res["needs_review"] is True
    assert res["review_reason"] == "side_indeterminate"


# --- check estrutural INLINE da pista vs Pass-1 (substitui o cross-check async) ---

_PROD_ROWS = [{"of": "262107", "modelo": "OMEGA"}, {"of": "262559", "modelo": "CGC2"}]


def test_hint_v_but_frente_flags_review(monkeypatch):
    # pista=V mas o Pass-1 tem cara de frente → provável frente perdida → revisão.
    _mock_pipeline(monkeypatch, rows=_PROD_ROWS)
    monkeypatch.setattr(ocr_runner, "_detect_side", _detect_must_not_run)
    res = ocr_runner.run_pipeline(Path("x.jpg"), page_hint="V")
    assert res["template_name"] == "paragens"      # routing respeita a pista
    assert res["side"] == "V"
    assert res["needs_review"] is True
    assert res["review_reason"] == "side_hint_conflict"


def test_hint_f_but_nonproduction_flags_review(monkeypatch):
    # pista=F mas ≥2 linhas preenchidas sem identidade de produção → provável verso.
    _mock_pipeline(monkeypatch, rows=[{"a": "x"}, {"b": "y"}])
    monkeypatch.setattr(ocr_runner, "_detect_side", _detect_must_not_run)
    res = ocr_runner.run_pipeline(Path("x.jpg"), page_hint="F")
    assert res["template_name"] == "guilhotina"    # routing respeita a pista
    assert res["side"] == "F"
    assert res["needs_review"] is True
    assert res["review_reason"] == "side_hint_conflict"


def test_hint_f_with_production_not_flagged(monkeypatch):
    _mock_pipeline(monkeypatch, rows=_PROD_ROWS)
    monkeypatch.setattr(ocr_runner, "_detect_side", _detect_must_not_run)
    res = ocr_runner.run_pipeline(Path("x.jpg"), page_hint="F")
    assert res["needs_review"] is False


def test_hint_v_with_empty_pass1_not_flagged(monkeypatch):
    # pista=V com Pass-1 vazio (verso real) → sem conflito, sem revisão.
    _mock_pipeline(monkeypatch, rows=[])
    monkeypatch.setattr(ocr_runner, "_detect_side", _detect_must_not_run)
    res = ocr_runner.run_pipeline(Path("x.jpg"), page_hint="V")
    assert res["template_name"] == "paragens"
    assert res["needs_review"] is False
