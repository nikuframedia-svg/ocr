"""CPIS migration export — schema + row-mapping tests.

The output of `build_cpis_workbook` must match the current CPIS column
layout at the header level (exact labels, `Folha1` sheet name).
"""
from __future__ import annotations

import datetime as dt
import io

import openpyxl
from app.web.export import (
    CPIS_COLUMNS,
    _build_cpis_row,
    _iso_to_date,
    cpis_filename_for,
)

# Cabeçalhos exatos do template (transcritos de
# /Users/martimnicolau/Downloads/MigracaoNikufraCPIS.xlsx).
EXPECTED_HEADER_LABELS = [
    "Data",
    "Cód. Funcionário",
    "Nome Funcionário",
    "Setor / Máquina Desc.",
    "Cód. Máquina",
    "OF",
    "OV",
    "Cliente",
    "Modelo",
    "QTD",
    "Qtd Metros",
    "M²",
    "Nesting",
    "Cesta Nº",
    "Duração",
    "Comprimento (mm)",
    "Largura (mm)",
    "Espessura (mm)",
    "CONI",
    "Nº Chapas",
    "Peso Consumido (t)",
    "Peso Produzido (t)",
    "Desperdício (t)",
    "% Desperdício",
    # R261 — Sucata (rev00) acrescentada no fim do schema CPIS.
    "Sucata",
    # Lote acrescentado no fim para preservar as posições anteriores.
    "Lote",
]


def _refs():
    maquinas = {
        "M045": {
            "codmaq": "M045",
            "desmaq": "QUINADORA ADIRA 14M P8",
            "desigkanban": "QUINADORA P8",
            "colunaexcel": "Q",
        },
        "M040": {
            "codmaq": "M040",
            "desmaq": "QUINADORA CÓNICA P8",
            "desigkanban": "",
            "colunaexcel": "Q",
        },
    }
    return {
        "of_to_entries": {
            "999999": [{
                "of": "999999",
                "ov": "100200",
                "cliente": "ENEDIS",
                "designacao": "CGC2E10D",
                "comp": 5000,
                "lbase": 200,
                "ltopo": 150,
                "esp": 2.6,
                "npecas": 6,
                "pesounit": 17.85875,
            }],
        },
        "plan_by_ov": {
            "100200": [{
                "_of": "999999",
                "of": "999999",
                "ov": "100200",
                "cliente": "ENEDIS",
                "designacao": "CGC2E10D",
                "comp": 5000,
                "lbase": 200,
                "ltopo": 150,
                "esp": 2.6,
                "npecas": 6,
                "pesounit": 17.85875,
            }],
        },
        "lotes_sap_full": {
            "L1": {"esp": 2.6, "larg": 1500},
        },
        "maquinas_by_codmaq": maquinas,
        "maquinas_by_kanban": {
            "QUINADORA P8": maquinas["M045"],
            "QUINADORA CÓNICA P8": maquinas["M040"],
        },
    }


def test_cpis_columns_match_template_exactly() -> None:
    """The header labels must match the CPIS export contract."""
    labels = [label for _, label in CPIS_COLUMNS]
    assert labels == EXPECTED_HEADER_LABELS


def test_iso_to_date_parses_valid_iso() -> None:
    assert _iso_to_date("2026-04-09") == dt.date(2026, 4, 9)
    assert _iso_to_date(None) is None
    assert _iso_to_date("") is None
    assert _iso_to_date("not-a-date") is None


def test_build_cpis_row_bobine_formato_with_waste() -> None:
    """Typical BOBINE-FORMATO row → CPIS fields populated."""
    raw = {
        "sheet_iso_date": "2026-04-09",
        "operador": "JÚLIO LIMA",
        "validated_operador": "JÚLIO LIMA",
        "n_operador": "0537",
        "setor_maquina": "BOBINE-FORMATO",
        "header_cod_maquina": None,
        "of": "999999",
        "ov": "100200",
        "cliente": "ENEDIS",
        "modelo": "CGC2E10D",
        "qtd": 5,
        "comp_mm": 9999,
        "larg_mm": 999,
        "lote": "L1",
        "lbase": 200,
        "ltopo": 150,
        "esp": 26.0,
        "coni": "10",
        "sucata": "3",
    }
    out = _build_cpis_row(raw, refs=_refs())

    assert out["data"] == dt.date(2026, 4, 9)
    assert out["cod_funcionario"] == 10000537
    assert out["nome_funcionario"] == "JÚLIO LIMA"
    assert out["setor_maquina_desc"] == "BOBINE-FORMATO"
    assert out["cod_maquina"] == "M032"  # derived from setor
    assert out["of"] == "999999"
    assert out["ov"] == "100200"
    assert out["cliente"] == "ENEDIS"
    assert out["modelo"] == "CGC2E10D"
    assert out["qtd"] == 5
    assert out["comp_mm"] == 5000
    assert out["larg_mm"] == 1500
    assert out["esp_mm"] == 2.6
    assert out["coni"] == "10"
    assert out["n_chapas"] == 1
    assert out["peso_consumido_t"] == 0.153
    assert out["peso_produzido_t"] == 0.089
    assert out["desperdicio_t"] == 0.064
    assert out["desperdicio_pct"] is not None
    assert out["sucata"] == 3  # R261
    assert out["lote"] == "L1"


def test_build_cpis_row_sucata_empty_is_none() -> None:
    """Linhas sem sucata exportam célula vazia (None), não 0."""
    raw = {
        "sheet_iso_date": "2026-04-09",
        "n_operador": "1",
        "setor_maquina": "BOBINE-FORMATO",
        "operador": "X",
        "qtd": 5,
    }
    out = _build_cpis_row(raw)
    assert out["sucata"] is None
    assert out["lote"] == ""


def test_build_cpis_row_header_cod_maquina_wins_over_derivation() -> None:
    """If OCR captures cod_maquina directly, prefer it over the table."""
    raw = {
        "sheet_iso_date": "2026-04-09",
        "n_operador": "1",
        "setor_maquina": "UNKNOWN-MACHINE",  # not in derivation table
        "header_cod_maquina": "M999",
        "operador": "X",
    }
    out = _build_cpis_row(raw)
    assert out["cod_maquina"] == "M999"


def test_build_cpis_row_missing_geometry_returns_none_for_weights() -> None:
    """Rows without enough geometry skip Peso/Desperdício gracefully."""
    raw = {
        "sheet_iso_date": "2026-04-09",
        "n_operador": "1",
        "setor_maquina": "BOBINE-FORMATO",
        "operador": "X",
        "qtd": 5,
        # No plan/SAP/OCR geometry, so the shared weight layer returns empty.
    }
    out = _build_cpis_row(raw)
    assert out["peso_consumido_t"] is None
    assert out["peso_produzido_t"] is None
    assert out["desperdicio_t"] is None
    assert out["desperdicio_pct"] is None
    # Other fields still present
    assert out["qtd"] == 5
    assert out["setor_maquina_desc"] == "BOBINE-FORMATO"


def test_build_cpis_row_acabamento_produced_only_from_plan() -> None:
    """Acabamento TPL086 não cria consumo/desperdício, só peso produzido."""
    raw = {
        "sheet_iso_date": "2026-05-25",
        "operador": "JÚLIO LIMA",
        "validated_operador": "JÚLIO LIMA",
        "n_operador": "0537",
        "setor_maquina": "ACABAMENTO MTG4",
        "header_cod_maquina": None,
        "of": "999999",
        "modelo": "CGC2E10D",
        "qtd": 10,
    }
    out = _build_cpis_row(raw, refs=_refs())

    assert out["cod_maquina"] == "M061"
    assert out["ov"] == "100200"
    assert out["cliente"] == "ENEDIS"
    assert out["peso_produzido_t"] == 0.179
    assert out["n_chapas"] is None
    assert out["peso_consumido_t"] is None
    assert out["desperdicio_t"] is None
    assert out["desperdicio_pct"] is None


def test_build_cpis_row_quinadora_derives_cod_maquina_from_refs() -> None:
    raw = {
        "sheet_iso_date": "2026-05-25",
        "n_operador": "0537",
        "setor_maquina": "QUINADORA PAV.8",
        "header_cod_maquina": None,
        "operador": "JÚLIO LIMA",
    }
    out = _build_cpis_row(raw, refs=_refs())

    assert out["cod_maquina"] == "M045"


def test_build_cpis_row_expedicao_produced_weight_by_of() -> None:
    raw = {
        "sheet_iso_date": "2026-05-25",
        "n_operador": "0537",
        "setor_maquina": "EXPEDIÇÃO",
        "header_cod_maquina": None,
        "operador": "JÚLIO LIMA",
        "of": "999999",
        "modelo": "CGC2E10D",
        "qtd": 10,
    }
    out = _build_cpis_row(raw, refs=_refs())

    assert out["peso_produzido_t"] == 0.179
    assert out["n_chapas"] is None
    assert out["peso_consumido_t"] is None
    assert out["desperdicio_t"] is None
    assert out["desperdicio_pct"] is None


def test_build_cpis_row_expedicao_produced_weight_by_unique_ov_model() -> None:
    raw = {
        "sheet_iso_date": "2026-05-25",
        "n_operador": "0537",
        "setor_maquina": "EXPEDIÇÃO",
        "header_cod_maquina": None,
        "operador": "JÚLIO LIMA",
        "of": "",
        "ov": "100200",
        "modelo": "CGC2E10D",
        "qtd": 10,
    }
    out = _build_cpis_row(raw, refs=_refs())

    assert out["of"] == ""
    assert out["ov"] == "100200"
    assert out["cliente"] == "ENEDIS"
    assert out["peso_produzido_t"] == 0.179
    assert out["peso_consumido_t"] is None


def test_build_cpis_row_expedicao_wrong_existing_of_uses_ov_model() -> None:
    refs = _refs()
    refs["of_to_entries"]["888888"] = [{
        "of": "888888",
        "ov": "999000",
        "cliente": "OTHER",
        "designacao": "OUTRA PECA",
        "pesounit": 99,
    }]
    raw = {
        "sheet_iso_date": "2026-05-25",
        "n_operador": "0537",
        "setor_maquina": "EXPEDIÇÃO",
        "header_cod_maquina": None,
        "operador": "JÚLIO LIMA",
        "of": "888888",
        "ov": "100200",
        "modelo": "CGC2E10D",
        "qtd": 10,
    }
    out = _build_cpis_row(raw, refs=refs)

    assert out["of"] == "888888"
    assert out["peso_produzido_t"] == 0.179
    assert out["peso_consumido_t"] is None


def test_build_cpis_row_expedicao_ambiguous_ov_model_leaves_weight_empty() -> None:
    refs = _refs()
    refs["plan_by_ov"]["100200"].append({
        "_of": "888888",
        "of": "888888",
        "ov": "100200",
        "cliente": "ENEDIS",
        "designacao": "CGC2E10D",
        "pesounit": 99,
    })
    raw = {
        "sheet_iso_date": "2026-05-25",
        "n_operador": "0537",
        "setor_maquina": "EXPEDIÇÃO",
        "operador": "JÚLIO LIMA",
        "of": "",
        "ov": "100200",
        "modelo": "CGC2E10D",
        "qtd": 10,
    }
    out = _build_cpis_row(raw, refs=refs)

    assert out["peso_produzido_t"] is None
    assert out["peso_consumido_t"] is None


def test_cpis_filename_includes_operador_slug() -> None:
    f = cpis_filename_for("2026-04-01", "2026-04-30")
    assert f == "MigracaoNikufraCPIS_2026-04-01_2026-04-30.xlsx"

    f_op = cpis_filename_for("2026-04-01", "2026-04-30", "JÚLIO LIMA")
    assert f_op.startswith("MigracaoNikufraCPIS_2026-04-01_2026-04-30_")
    assert f_op.endswith(".xlsx")
    # Diacritics stripped, spaces removed
    assert "JLIOLIMA" in f_op or "JULIOLIMA" in f_op or "LIMA" in f_op


def test_cpis_filename_for_selected_dates_is_compact_and_deterministic() -> None:
    selected = ("2026-07-29", "2026-07-23", "2026-07-27", "2026-07-23")
    filename = cpis_filename_for(
        None,
        None,
        selected_dates=selected,
    )
    assert filename == "MigracaoNikufraCPIS_3-dias_2026-07-23_2026-07-29.xlsx"

    one_day = cpis_filename_for(
        None,
        None,
        selected_dates=("2026-07-23",),
    )
    assert one_day == "MigracaoNikufraCPIS_1-dia_2026-07-23.xlsx"

    filtered = cpis_filename_for(
        None,
        None,
        operador="OPERADOR TESTE",
        sector="Bobine Formato",
        validated_only=True,
        selected_dates=("2026-07-23", "2026-07-29"),
    )
    assert filtered == (
        "MigracaoNikufraCPIS_2-dias_2026-07-23_2026-07-29"
        "_bobine_formato_OPERADORTESTE_validadas.xlsx"
    )


def test_workbook_header_row_matches_template() -> None:
    """End-to-end: build a workbook from a synthetic row and read it back.

    Validates: sheet name = Folha1, first row = expected labels, second
    row has the synthetic data in the right slots.
    """
    # Inline-import to avoid pulling kpis/db at module-load time
    from app.web import export

    # Stub out the DB query — we don't need real data for this test
    synthetic = [{
        "sheet_iso_date": "2026-04-09",
        "operador": "JÚLIO LIMA",
        "validated_operador": "JÚLIO LIMA",
        "n_operador": "0537",
        "setor_maquina": "BOBINE-FORMATO",
        "header_cod_maquina": None,
        "of": "999999",
        "ov": "100200",
        "cliente": "ENEDIS",
        "modelo": "CGC2E10D",
        "qtd": 5,
        "comp_mm": 5000,
        "larg_mm": 1500,
        "lbase": 200,
        "ltopo": 150,
        "esp": 3.0,
        "coni": "10",
        "lote": "M26B0330",
    }]
    original_query = export._query_cpis_rows
    export._query_cpis_rows = lambda *a, **kw: synthetic  # type: ignore[attr-defined]
    try:
        xlsx_bytes = export.build_cpis_workbook("2026-04-01", "2026-04-30")
    finally:
        export._query_cpis_rows = original_query  # type: ignore[attr-defined]

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == ["Folha1"]
    ws = wb["Folha1"]
    header_row = [c.value for c in ws[1]]
    assert header_row == EXPECTED_HEADER_LABELS

    data_row = [c.value for c in ws[2]]
    data_by_header = dict(zip(header_row, data_row, strict=True))
    # openpyxl reads date cells back as datetime; compare the date portion.
    assert isinstance(data_row[0], (dt.date, dt.datetime))
    assert (data_row[0].date() if isinstance(data_row[0], dt.datetime)
            else data_row[0]) == dt.date(2026, 4, 9)
    assert data_by_header["Cód. Funcionário"] == 10000537
    assert data_by_header["Nome Funcionário"] == "JÚLIO LIMA"
    assert data_by_header["Setor / Máquina Desc."] == "BOBINE-FORMATO"
    assert data_by_header["Cód. Máquina"] == "M032"
    assert data_by_header["OF"] == "999999"
    assert data_by_header["Cliente"] == "ENEDIS"
    assert data_by_header["QTD"] == 5
    assert data_by_header["Comprimento (mm)"] == 5000
    assert data_by_header["Largura (mm)"] == 1500
    assert data_by_header["Espessura (mm)"] is not None
    assert data_by_header["Nº Chapas"] is not None
    assert data_by_header["Peso Consumido (t)"] > 0
    assert data_by_header["Desperdício (t)"] >= 0
    assert header_row[-1] == "Lote"
    assert data_row[-1] == "M26B0330"


def test_workbook_empty_period_still_has_header() -> None:
    """Zero rows → workbook with just the header row, valid for download."""
    from app.web import export

    original_query = export._query_cpis_rows
    export._query_cpis_rows = lambda *a, **kw: []  # type: ignore[attr-defined]
    try:
        xlsx_bytes = export.build_cpis_workbook("2026-04-01", "2026-04-30")
    finally:
        export._query_cpis_rows = original_query  # type: ignore[attr-defined]

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Folha1"]
    assert [c.value for c in ws[1]] == EXPECTED_HEADER_LABELS
    # No data rows
    assert ws.max_row == 1
