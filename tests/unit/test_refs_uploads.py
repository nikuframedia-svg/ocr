"""Tests for Round 106 — refs upload log, OF normalisation, phase priority."""
from __future__ import annotations

import json
import shutil
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.cross_check import ref_importer, ref_watcher, refs_uploads, storage
from app.pipeline.scoring_engine import ENGINE_VERSION, normalize_of
from app.web import main


@pytest.fixture
def log_path(tmp_path, monkeypatch):
    monkeypatch.setattr(refs_uploads, "_LOG_PATH", tmp_path / "refs_uploads.json")
    return tmp_path / "refs_uploads.json"


def _write_plan(path, rows, extra_cols=()):
    wb = Workbook()
    ws = wb.active
    ws.title = "plan_colunas_cpis"
    ws.append([
        "cliente", "ov", "of", "designacao", "quanttrp",
        "bf", "c", "q", "s", "r", "a", "exp",
        "esp", "lbase", "ltopo", "comp",
        *extra_cols,
    ])
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_stocksap(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Folha1"
    ws.append(["Lote", "Qtd", "Espessura", "Largura", "Desc"])
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_maquinas(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "codmaq", "desmaq", "desigkanban", "codsec", "ativo",
        "ordem", "operacao", "centrotrab", "dessec", "colunaexcel",
    ])
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_colaboradores(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(["pernr", "sname", "cod"])
    for row in rows:
        ws.append(row)
    wb.save(path)


# --- upload log -----------------------------------------------------------

def test_upload_log_records_most_recent_first(log_path):
    assert refs_uploads.recent() == []
    refs_uploads.record(
        "plan", "plan_colunas_cpis.xlsx", 16617,
        sha256="abc123", n_ofs=5291, n_ovs=5000, size=123456,
    )
    refs_uploads.record("stocksap", "StockSAP.xlsx", 2770)
    rows = refs_uploads.recent()
    assert len(rows) == 2
    assert rows[0]["kind"] == "stocksap"        # most recent first
    assert rows[1]["kind"] == "plan"
    assert rows[1]["n_rows"] == 16617
    assert rows[1]["sha256"] == "abc123"
    assert rows[1]["n_ofs"] == 5291
    assert rows[1]["n_ovs"] == 5000
    assert rows[1]["size"] == 123456


# --- OF normalisation -----------------------------------------------------

def test_normalize_of():
    assert normalize_of("52306") == "052306"     # zero-pad short
    assert normalize_of("254877") == "254877"    # already 6
    assert normalize_of(254877) == "254877"      # int input
    assert normalize_of("05000A") == "05000A"    # has letters — untouched
    assert normalize_of("2502343") == "2502343"  # 7 digits — untouched
    assert normalize_of("") == ""
    assert normalize_of(None) == ""


# --- phase helpers --------------------------------------------------------

def test_phase_columns_detected_between_quanttrp_and_esp():
    hdrs = {"cliente": 0, "ov": 1, "of": 2, "designacao": 3, "quanttrp": 4,
            "bf": 5, "c": 6, "q": 7, "s": 8, "r": 9, "a": 10, "exp": 11,
            "esp": 12, "lbase": 13}
    assert ref_watcher._phase_columns(hdrs) == ["bf", "c", "q", "s", "r", "a", "exp"]


def test_phase_columns_empty_when_no_quanttrp():
    assert ref_watcher._phase_columns({"of": 0, "esp": 1}) == []


def test_fase_incompleta():
    # all phases == quanttrp → concluded
    assert ref_watcher._fase_incompleta(20, {"bf": 20, "c": 20, "exp": 20}) is False
    # one phase below quanttrp → still in production
    assert ref_watcher._fase_incompleta(20, {"bf": 20, "c": 20, "exp": 10}) is True
    # blank phase counts as 0 → incomplete
    assert ref_watcher._fase_incompleta(20, {"bf": 20, "c": None}) is True
    # quanttrp 0 / blank → cannot judge → not flagged
    assert ref_watcher._fase_incompleta(0, {"bf": 0}) is False
    assert ref_watcher._fase_incompleta(None, {"bf": 5}) is False


# --- plan upload audit -----------------------------------------------------

def test_ref_watcher_tracks_plan_hash_and_snapshot(tmp_path):
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    plan = doc_dir / "plan_colunas_cpis.xlsx"
    _write_plan(plan, [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
        ["B", "2603978", "263349", "CAC4E10C", 2, 0, 0, 0, 0, 0, 0, 0, 3, 600, 200, 9000],
    ])

    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    refs = watcher.force_reload()
    sha = ref_watcher.file_sha256(plan)

    assert refs["plan_sha256"] == sha
    assert refs["plan_size"] == plan.stat().st_size
    assert refs["stats"]["n_plan_rows"] == 2
    assert refs["stats"]["n_ofs"] == 2
    assert refs["stats"]["n_ovs"] == 2

    snap = ref_watcher.refs_snapshot(refs, watcher.plan_path)
    assert snap["plan_sha256"] == sha
    assert snap["plan_rows"] == 2
    assert snap["plan_ofs"] == 2
    assert snap["plan_ovs"] == 2
    assert snap["files"]["plan"]["sha256"] == sha
    assert snap["files"]["plan"]["rows"] == 2

    status = json.loads((doc_dir / "_refs_status.json").read_text())
    assert status["plan_colunas"]["sha256"] == sha
    assert status["plan_colunas"]["n_ovs"] == 2


def test_ref_watcher_tracks_all_ref_workbook_hashes_and_snapshot(tmp_path):
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    plan = doc_dir / "plan_colunas_cpis.xlsx"
    sap = doc_dir / "StockSAP.xlsx"
    maquinas = doc_dir / "maquinas.xlsx"
    colabs = doc_dir / "ListaColaboradores.xlsx"
    _write_plan(plan, [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ])
    _write_stocksap(sap, [["M26B001", 1, 2.6, 1250, "bobine"]])
    _write_maquinas(maquinas, [
        ["M061", "ACABAMENTO MTG2", "ACABAMENTO MTG2", "A", 1, 1, "", "", "ACABAMENTO", "a"],
        ["M062", "ACABAMENTO MTG4", "ACABAMENTO MTG4", "A", 1, 2, "", "", "ACABAMENTO", "a"],
    ])
    _write_colaboradores(colabs, [
        ["0000000537", "JULIO LIMA", 537],
        ["0000000123", "ANA SILVA", 123],
    ])

    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    refs = watcher.force_reload()
    snap = ref_watcher.refs_snapshot(refs, watcher.plan_path)

    assert snap["files"]["plan"]["sha256"] == ref_watcher.file_sha256(plan)
    assert snap["files"]["stocksap"]["sha256"] == ref_watcher.file_sha256(sap)
    assert snap["files"]["stocksap"]["lotes"] == 2  # lote + alias sem M
    assert snap["files"]["maquinas"]["sha256"] == ref_watcher.file_sha256(maquinas)
    assert snap["files"]["maquinas"]["maquinas"] == 2
    assert snap["files"]["colaboradores"]["sha256"] == ref_watcher.file_sha256(colabs)
    assert snap["files"]["colaboradores"]["colaboradores"] == 2

    status = watcher.status()
    assert status["sap"]["sha256"] == ref_watcher.file_sha256(sap)
    assert status["maquinas"]["n_maquinas"] == 2
    assert status["colaboradores"]["n_colaboradores"] == 2


def test_ref_importer_imports_latest_valid_refs_by_content(tmp_path, log_path):
    doc_dir = tmp_path / "docs"
    source_dir = tmp_path / "source"
    doc_dir.mkdir()
    source_dir.mkdir()
    _write_plan(doc_dir / "plan_colunas_cpis.xlsx", [
        ["OLD", "2600000", "260000", "OLDMODEL", 1, 0, 0, 0, 0, 0, 0, 0, 4, 600, 200, 9000],
    ])
    _write_stocksap(doc_dir / "StockSAP.xlsx", [["M26B001", 1, 2.6, 1250, "old"]])
    _write_maquinas(doc_dir / "maquinas.xlsx", [
        ["M001", "OLD", "OLD", "A", 1, 1, "", "", "OLD", "a"],
    ])
    _write_colaboradores(doc_dir / "ListaColaboradores.xlsx", [
        ["0000000001", "OLD USER", 1],
    ])
    _write_plan(source_dir / "plan_colunas_cpis (4).xlsx", [
        ["MTG GMBH", "2602568", "222414", "SCD301J07", 1, 0, 0, 0, 0, 0, 0, 0, 3, 610, 210, 9100],
        ["LE HAVRE", "2512130", "260108", "CFH2F07RI", 9, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ])
    _write_stocksap(source_dir / "stock_export_noite.xlsx", [
        ["M26B999", 1, 3.0, 1500, "new"],
    ])
    _write_maquinas(source_dir / "mapa_maquinas.xlsx", [
        ["M030", "LASER", "LASER", "C", 1, 2, "", "", "CORTE", "c"],
    ])
    _write_colaboradores(source_dir / "lista_pessoas.xlsx", [
        ["0000000537", "JULIO LIMA", 537],
        ["0000000123", "ANA SILVA", 123],
    ])
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    watcher.force_reload()

    result = ref_importer.import_refs_from_dir(source_dir, watcher=watcher)

    assert result["ok"] is True
    assert {item["kind"] for item in result["imported"]} == {
        "plan", "stocksap", "maquinas", "colaboradores",
    }
    refs = watcher.get_refs()
    assert refs["plan_sha256"] == ref_watcher.file_sha256(source_dir / "plan_colunas_cpis (4).xlsx")
    assert refs["sap_sha256"] == ref_watcher.file_sha256(source_dir / "stock_export_noite.xlsx")
    assert refs["maquinas_sha256"] == ref_watcher.file_sha256(source_dir / "mapa_maquinas.xlsx")
    assert refs["colab_sha256"] == ref_watcher.file_sha256(source_dir / "lista_pessoas.xlsx")
    assert refs["stats"]["n_plan_rows"] == 2
    assert refs["stats"]["n_ofs"] == 2
    assert refs["stats"]["n_lotes"] == 2
    assert refs["stats"]["n_maquinas"] == 1
    assert refs["stats"]["n_colaboradores"] == 2
    assert {u["kind"] for u in refs_uploads.recent()[:4]} == {
        "plan", "stocksap", "maquinas", "colaboradores",
    }


def test_ref_importer_rejects_invalid_refs_without_replacing_active(tmp_path, log_path):
    doc_dir = tmp_path / "docs"
    source_dir = tmp_path / "source"
    doc_dir.mkdir()
    source_dir.mkdir()
    active_plan = doc_dir / "plan_colunas_cpis.xlsx"
    _write_plan(active_plan, [
        ["ACTIVE", "2600000", "260000", "ACTIVE_MODEL", 1, 0, 0, 0, 0, 0, 0, 0, 4, 600, 200, 9000],
    ])
    bad = Workbook()
    bad.active.append(["of", "cliente"])
    bad.active.append(["222414", "MTG GMBH"])
    bad.save(source_dir / "plan_colunas_cpis.xlsx")
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    before_sha = ref_watcher.file_sha256(active_plan)

    result = ref_importer.import_refs_from_dir(source_dir, watcher=watcher)

    assert result["ok"] is True
    assert result["imported"] == []
    assert result["rejected"]
    assert ref_watcher.file_sha256(active_plan) == before_sha
    refs = watcher.force_reload()
    assert refs["plan_sha256"] == before_sha
    assert refs["stats"]["n_plan_rows"] == 1


def test_ref_importer_skips_same_hash(tmp_path, log_path):
    doc_dir = tmp_path / "docs"
    source_dir = tmp_path / "source"
    doc_dir.mkdir()
    source_dir.mkdir()
    rows = [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ]
    _write_plan(doc_dir / "plan_colunas_cpis.xlsx", rows)
    shutil.copy2(doc_dir / "plan_colunas_cpis.xlsx", source_dir / "plan_noite.xlsx")
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)

    result = ref_importer.import_refs_from_dir(source_dir, watcher=watcher)

    assert result["ok"] is True
    assert result["imported"] == []
    assert result["skipped"][0]["kind"] == "plan"


def test_ref_importer_preserves_windows_configured_paths(monkeypatch):
    monkeypatch.setenv("KANBAN_REFS_IMPORT_DIR", r"F:\ocr\files")
    assert str(ref_importer.configured_import_dir()) == r"F:\ocr\files"

    monkeypatch.setenv("KANBAN_REFS_IMPORT_DIR", r"\\srv-planeamento\ocr\files")
    assert str(ref_importer.configured_import_dir()) == r"\\srv-planeamento\ocr\files"


def test_ref_importer_default_is_external_windows_folder(monkeypatch):
    for env_name in ref_importer.IMPORT_DIR_ENV_VARS:
        monkeypatch.delenv(env_name, raising=False)
    monkeypatch.setattr(ref_importer, "_config_value", lambda _name: None)

    path = ref_importer.configured_import_dir()

    assert str(path) == r"F:\ocr\files"
    assert ref_importer._is_absolute_import_path(path) is True


def test_ref_importer_starts_for_windows_absolute_source_even_if_missing(monkeypatch):
    state = {
        "enabled": False,
        "running": False,
        "source_dir": "",
        "interval_seconds": None,
        "last_run_at": None,
        "last_ok": None,
        "last_error": None,
        "last_result": None,
    }
    started: list[bool] = []

    class FakeThread:
        def __init__(self, *args, **kwargs):
            self.args = args
            self.kwargs = kwargs

        def start(self):
            started.append(True)

        def is_alive(self):
            return bool(started)

    monkeypatch.setattr(ref_importer, "_state", state)
    monkeypatch.setattr(ref_importer, "_thread", None)
    monkeypatch.setattr(ref_importer.threading, "Thread", FakeThread)

    ok = ref_importer.start_background_importer(
        source_dir=Path(r"F:\ocr\files"),
        interval_seconds=30,
    )

    assert ok is True
    assert started == [True]
    status = ref_importer.status()
    assert status["enabled"] is True
    assert status["thread_alive"] is True
    assert status["source_dir"] == r"F:\ocr\files"
    assert status["last_error"] == "pasta de importação não existe"


def test_plan_inspection_rejects_missing_required_columns(tmp_path):
    bad = tmp_path / "bad_plan.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["of", "cliente"])
    ws.append(["263348", "COBELBA"])
    wb.save(bad)

    err, info = main._inspect_refs_xlsx(bad, "plan")

    assert err is not None
    assert "ov" in err
    assert "quanttrp" in err
    assert info["n_rows"] == 0


def test_refs_inspection_accepts_machines_and_collaborators(tmp_path):
    maquinas = tmp_path / "maquinas.xlsx"
    colabs = tmp_path / "ListaColaboradores.xlsx"
    _write_maquinas(maquinas, [
        ["M061", "ACABAMENTO MTG2", "ACABAMENTO MTG2", "A", 1, 1, "", "", "ACABAMENTO", "a"],
        ["M030", "LASER", "LASER", "C", 1, 2, "", "", "CORTE", "c"],
    ])
    _write_colaboradores(colabs, [
        ["0000000537", "JULIO LIMA", 537],
        ["0000000123", "ANA SILVA", 123],
    ])

    maq_err, maq_info = main._inspect_refs_xlsx(maquinas, "maquinas")
    colab_err, colab_info = main._inspect_refs_xlsx(colabs, "colaboradores")

    assert maq_err is None
    assert maq_info["n_maquinas"] == 2
    assert colab_err is None
    assert colab_info["n_colaboradores"] == 2


def test_refs_inspection_rejects_invalid_machines_and_collaborators(tmp_path):
    bad_maquinas = tmp_path / "bad_maquinas.xlsx"
    bad_colabs = tmp_path / "bad_colabs.xlsx"
    wb = Workbook()
    wb.active.append(["codmaq", "desmaq"])
    wb.save(bad_maquinas)
    wb = Workbook()
    wb.active.append(["pernr", "sname"])
    wb.save(bad_colabs)

    maq_err, _ = main._inspect_refs_xlsx(bad_maquinas, "maquinas")
    colab_err, _ = main._inspect_refs_xlsx(bad_colabs, "colaboradores")

    assert maq_err is not None
    assert "colunaexcel" in maq_err
    assert colab_err is not None
    assert "cod" in colab_err


def test_refs_upload_replaces_active_plan_and_reload_uses_same_hash(
    tmp_path, monkeypatch, log_path
):
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    upload_plan = tmp_path / "upload_plan.xlsx"
    _write_plan(upload_plan, [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
        ["B", "2603978", "263349", "CAC4E10C", 2, 0, 0, 0, 0, 0, 0, 0, 3, 600, 200, 9000],
        ["C", "2603979", "263350", "CAC4E10D", 3, 0, 0, 0, 0, 0, 0, 0, 3, 610, 210, 9100],
    ])
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    monkeypatch.setattr(main, "get_watcher", lambda: watcher)

    client = TestClient(main.app)
    with upload_plan.open("rb") as f:
        resp = client.post(
            "/refs/upload",
            data={"kind": "plan"},
            files={"file": ("plan_colunas_cpis.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            follow_redirects=False,
        )

    assert resp.status_code == 303
    active_sha = ref_watcher.file_sha256(watcher.plan_path)
    assert active_sha == ref_watcher.file_sha256(upload_plan)
    refs = watcher.get_refs()
    assert refs["plan_sha256"] == active_sha
    assert refs["stats"]["n_plan_rows"] == 3
    assert refs["stats"]["n_ofs"] == 3
    assert refs["stats"]["n_ovs"] == 3
    assert active_sha[:8] in resp.headers["location"]

    uploads = refs_uploads.recent()
    assert uploads[0]["sha256"] == active_sha
    assert uploads[0]["n_rows"] == 3
    assert uploads[0]["n_ofs"] == 3
    assert uploads[0]["n_ovs"] == 3


@pytest.mark.parametrize(
    ("kind", "filename", "writer", "rows", "stat_key", "count"),
    [
        (
            "stocksap", "StockSAP.xlsx", _write_stocksap,
            [["M26B001", 1, 2.6, 1250, "bobine"]],
            "n_lotes", 2,
        ),
        (
            "maquinas", "maquinas.xlsx", _write_maquinas,
            [
                ["M061", "ACABAMENTO MTG2", "ACABAMENTO MTG2", "A", 1, 1, "", "", "ACABAMENTO", "a"],
                ["M030", "LASER", "LASER", "C", 1, 2, "", "", "CORTE", "c"],
            ],
            "n_maquinas", 2,
        ),
        (
            "colaboradores", "ListaColaboradores.xlsx", _write_colaboradores,
            [["0000000537", "JULIO LIMA", 537], ["0000000123", "ANA SILVA", 123]],
            "n_colaboradores", 2,
        ),
    ],
)
def test_refs_upload_replaces_active_non_plan_refs_and_reload_uses_same_hash(
    tmp_path, monkeypatch, log_path, kind, filename, writer, rows, stat_key, count
):
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    upload = tmp_path / filename
    writer(upload, rows)
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    monkeypatch.setattr(main, "get_watcher", lambda: watcher)

    client = TestClient(main.app)
    with upload.open("rb") as f:
        resp = client.post(
            "/refs/upload",
            data={"kind": kind},
            files={"file": (filename, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            follow_redirects=False,
        )

    assert resp.status_code == 303
    target = getattr(watcher, main._REFS_WATCHER_ATTRS[kind])
    active_sha = ref_watcher.file_sha256(target)
    assert active_sha == ref_watcher.file_sha256(upload)
    refs = watcher.get_refs()
    assert refs[main._REFS_SHA_KEYS[kind]] == active_sha
    assert refs["stats"][stat_key] == count
    assert active_sha[:8] in resp.headers["location"]

    uploads = refs_uploads.recent()
    assert uploads[0]["kind"] == kind
    assert uploads[0]["sha256"] == active_sha
    assert uploads[0]["n_rows"] == count


def test_refs_page_shows_active_plan_hash_counts_and_path(
    tmp_path, monkeypatch, log_path
):
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    plan = doc_dir / "plan_colunas_cpis.xlsx"
    _write_plan(plan, [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
        ["B", "2603978", "263349", "CAC4E10C", 2, 0, 0, 0, 0, 0, 0, 0, 3, 600, 200, 9000],
    ])
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    monkeypatch.setattr(main, "get_watcher", lambda: watcher)
    sha = ref_watcher.file_sha256(plan)

    resp = TestClient(main.app).get("/refs")

    assert resp.status_code == 200
    assert sha[:8] in resp.text
    assert "2 linhas" in resp.text
    assert "2 OVs" in resp.text
    assert str(plan) in resp.text


def test_refs_page_shows_all_ref_cards(tmp_path, monkeypatch, log_path):
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    _write_plan(doc_dir / "plan_colunas_cpis.xlsx", [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ])
    _write_stocksap(doc_dir / "StockSAP.xlsx", [["M26B001", 1, 2.6, 1250, "bobine"]])
    _write_maquinas(doc_dir / "maquinas.xlsx", [
        ["M061", "ACABAMENTO MTG2", "ACABAMENTO MTG2", "A", 1, 1, "", "", "ACABAMENTO", "a"],
    ])
    _write_colaboradores(doc_dir / "ListaColaboradores.xlsx", [
        ["0000000537", "JULIO LIMA", 537],
    ])
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    monkeypatch.setattr(main, "get_watcher", lambda: watcher)

    resp = TestClient(main.app).get("/refs")

    assert resp.status_code == 200
    assert "plan_colunas_cpis.xlsx" in resp.text
    assert "StockSAP.xlsx" in resp.text
    assert "maquinas.xlsx" in resp.text
    assert "ListaColaboradores.xlsx" in resp.text
    assert 'value="maquinas"' in resp.text
    assert 'value="colaboradores"' in resp.text


def test_refs_page_shows_last_external_import_result(tmp_path, monkeypatch, log_path):
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    _write_plan(doc_dir / "plan_colunas_cpis.xlsx", [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ])
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    monkeypatch.setattr(main, "get_watcher", lambda: watcher)
    monkeypatch.setattr(main.ref_importer, "status", lambda: {
        "enabled": True,
        "thread_alive": True,
        "source_dir": r"F:\ocr\files",
        "interval_seconds": 900,
        "last_run_at": "2026-06-16T01:00:00+00:00",
        "last_error": None,
        "last_result": {
            "ok": True,
            "imported": [{
                "kind": "plan",
                "filename": "plan_colunas_cpis.xlsx",
                "target": str(doc_dir / "plan_colunas_cpis.xlsx"),
                "sha256": "abcdef123456",
            }],
            "skipped": [],
            "candidates": [{"kind": "plan"}],
            "refs_loaded_at": "2026-06-16T01:01:00+00:00",
        },
    })

    resp = TestClient(main.app).get("/refs")

    assert resp.status_code == 200
    assert r"F:\ocr\files" in resp.text
    assert "plan: plan_colunas_cpis.xlsx" in resp.text
    assert "hash abcdef12" in resp.text
    assert str(doc_dir / "plan_colunas_cpis.xlsx") in resp.text


def test_cross_check_storage_persists_refs_snapshot(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    snap = {
        "plan_path": "/tmp/plan_colunas_cpis.xlsx",
        "plan_sha256": "abc123",
        "plan_mtime": 123.0,
        "plan_rows": 3,
        "plan_ofs": 3,
        "plan_ovs": 3,
        "refs_loaded_at": "2026-06-03T10:00:00+00:00",
    }

    stored = storage.store_cross_check(
        sheet_id=1,
        image_path="images/a.jpg",
        operador="OPERADOR",
        date_iso="2026-06-03",
        sheet_status="extracted",
        cross_check_result={
            "checked_at": "2026-06-03T10:01:00+00:00",
            "refs_loaded_at": snap["refs_loaded_at"],
            "refs_snapshot": snap,
            "engine_version": "test",
            "summary": {"match": 0, "no_match": 0, "na": 0, "total": 0},
            "rows": [],
            "header": {},
            "footer": {},
            "to_analisar": [],
        },
    )

    payload = json.loads(Path(stored["file"]).read_text(encoding="utf-8"))
    assert payload["refs_snapshot"] == snap


def test_cross_check_index_ok_rate_ignores_na_cells(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)

    storage.store_cross_check(
        sheet_id=1,
        image_path="images/a.jpg",
        operador="OPERADOR",
        date_iso="2026-06-03",
        sheet_status="extracted",
        cross_check_result={
            "checked_at": "2026-06-03T10:01:00+00:00",
            "summary": {"match": 8, "no_match": 2, "na": 10, "total": 20},
            "rows": [],
            "header": {},
            "footer": {},
            "to_analisar": [],
        },
    )

    index = json.loads((tmp_path / "_index.json").read_text(encoding="utf-8"))
    sheet = index["sheets"]["1"]
    assert sheet["summary"]["total"] == 20
    assert sheet["ok_rate"] == 0.8


def test_cross_check_index_rewrites_stale_ok_rates(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    stale = {
        "updated_at": "2026-06-03T10:00:00+00:00",
        "sheets": {
            "1": {
                "sheet_id": 1,
                "operador": "OPERADOR",
                "date": "2026-06-03",
                "sheet_status": "extracted",
                "summary": {"match": 8, "no_match": 2, "na": 10, "total": 20},
                "engine_version": ENGINE_VERSION,
                "ok_rate": 0.0,
                "file": "2026-06-03/old.json",
            },
        },
    }
    (tmp_path / "_index.json").write_text(json.dumps(stale), encoding="utf-8")

    storage.store_cross_check(
        sheet_id=2,
        image_path="images/b.jpg",
        operador="OPERADOR",
        date_iso="2026-06-03",
        sheet_status="extracted",
        cross_check_result={
            "checked_at": "2026-06-03T10:01:00+00:00",
            "summary": {"match": 1, "no_match": 0, "na": 0, "total": 1},
            "rows": [],
            "header": {},
            "footer": {},
            "to_analisar": [],
        },
    )

    index = json.loads((tmp_path / "_index.json").read_text(encoding="utf-8"))
    assert index["sheets"]["1"]["ok_rate"] == 0.8
    assert index["sheets"]["2"]["ok_rate"] == 1.0


def test_cross_check_summary_tolerates_string_counts(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    sheet_dir = tmp_path / "2026-06-03"
    sheet_dir.mkdir()
    (sheet_dir / "old.json").write_text(json.dumps({
        "sheet_id": 1,
        "engine_version": ENGINE_VERSION,
        "summary": {"match": "8", "no_match": "2", "na": "10", "total": "20"},
        "rows": [],
        "header": {},
        "footer": {},
        "to_analisar": [],
    }), encoding="utf-8")
    stale = {
        "updated_at": "2026-06-03T10:00:00+00:00",
        "sheets": {
            "1": {
                "sheet_id": 1,
                "operador": "OPERADOR",
                "date": "2026-06-03",
                "sheet_status": "extracted",
                "summary": {"match": "8", "no_match": "2", "na": "10", "total": "20"},
                "engine_version": ENGINE_VERSION,
                "ok_rate": 0.0,
                "file": "2026-06-03/old.json",
            },
        },
    }
    (tmp_path / "_index.json").write_text(json.dumps(stale), encoding="utf-8")

    storage.store_cross_check(
        sheet_id=2,
        image_path="images/b.jpg",
        operador="OPERADOR",
        date_iso="2026-06-03",
        sheet_status="extracted",
        cross_check_result={
            "checked_at": "2026-06-03T10:01:00+00:00",
            "summary": {"match": 1, "no_match": 0, "na": 0, "total": 1},
            "rows": [],
            "header": {},
            "footer": {},
            "to_analisar": [],
        },
    )

    summary = json.loads((tmp_path / "_summary.json").read_text(encoding="utf-8"))
    assert summary["totals"] == {"match": 9, "no_match": 2, "na": 10, "total": 21}
    index = json.loads((tmp_path / "_index.json").read_text(encoding="utf-8"))
    assert index["sheets"]["1"]["summary"] == {"match": 8, "no_match": 2, "na": 10, "total": 20}


def test_cross_check_summary_excludes_stale_engine_entries(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    stale = {
        "updated_at": "2026-06-03T10:00:00+00:00",
        "sheets": {
            "1": {
                "sheet_id": 1,
                "operador": "OPERADOR_ANTIGO",
                "date": "2026-06-02",
                "sheet_status": "extracted",
                "summary": {"match": 8, "no_match": 2, "na": 10, "total": 20},
                "engine_version": "v-old",
                "ok_rate": 0.8,
                "file": "2026-06-02/old.json",
            },
        },
    }
    (tmp_path / "_index.json").write_text(json.dumps(stale), encoding="utf-8")

    storage.store_cross_check(
        sheet_id=2,
        image_path="images/b.jpg",
        operador="OPERADOR",
        date_iso="2026-06-03",
        sheet_status="extracted",
        cross_check_result={
            "checked_at": "2026-06-03T10:01:00+00:00",
            "engine_version": ENGINE_VERSION,
            "summary": {"match": 1, "no_match": 0, "na": 0, "total": 1},
            "rows": [],
            "header": {},
            "footer": {},
            "to_analisar": [],
        },
    )

    summary = storage.load_summary()

    assert summary["engine_version"] == ENGINE_VERSION
    assert summary["totals"] == {"match": 1, "no_match": 0, "na": 0, "total": 1}
    assert summary["n_sheets"] == 1
    assert summary["stale_sheets"] == 1


def test_iter_sheet_cross_checks_can_include_stale_entries(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    stale_dir = tmp_path / "2026-06-02"
    stale_dir.mkdir()
    (stale_dir / "old.json").write_text(
        json.dumps({"sheet_id": 1, "engine_version": "v-old", "rows": []}),
        encoding="utf-8",
    )
    index = {
        "updated_at": "2026-06-03T10:00:00+00:00",
        "sheets": {
            "1": {
                "sheet_id": 1,
                "operador": "OPERADOR_ANTIGO",
                "date": "2026-06-02",
                "sheet_status": "extracted",
                "summary": {"match": 1, "no_match": 0, "na": 0, "total": 1},
                "engine_version": "v-old",
                "ok_rate": 1.0,
                "file": "2026-06-02/old.json",
            },
        },
    }
    (tmp_path / "_index.json").write_text(json.dumps(index), encoding="utf-8")

    current = storage.iter_sheet_cross_checks()
    all_engines = storage.iter_sheet_cross_checks(include_stale=True)

    assert current == []
    assert [item["sheet_id"] for item in all_engines] == [1]


def test_load_sheet_cross_check_excludes_stale_by_default(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    stale_dir = tmp_path / "2026-06-02"
    stale_dir.mkdir()
    (stale_dir / "old.json").write_text(
        json.dumps({
            "sheet_id": 1,
            "engine_version": "v-old",
            "rows": [{"row_index": 0, "fields": {}}],
        }),
        encoding="utf-8",
    )
    (tmp_path / "_index.json").write_text(json.dumps({
        "updated_at": "2026-06-03T10:00:00+00:00",
        "sheets": {
            "1": {
                "sheet_id": 1,
                "operador": "OPERADOR_ANTIGO",
                "date": "2026-06-02",
                "sheet_status": "extracted",
                "summary": {"match": 1, "no_match": 0, "na": 0, "total": 1},
                "engine_version": "v-old",
                "ok_rate": 1.0,
                "file": "2026-06-02/old.json",
            },
        },
    }), encoding="utf-8")

    assert storage.load_sheet_cross_check(1) is None
    assert storage.load_sheet_cross_check(1, include_stale=True)["engine_version"] == "v-old"


def test_load_sheet_cross_check_rejects_stale_payload_even_when_index_is_current(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    sheet_dir = tmp_path / "2026-06-03"
    sheet_dir.mkdir()
    (sheet_dir / "mismatched.json").write_text(
        json.dumps({
            "sheet_id": 2,
            "engine_version": "v-old",
            "rows": [{"row_index": 0, "fields": {}}],
            "to_analisar": [{
                "section": "rows",
                "row_index": 0,
                "field": "modelo",
                "value": "OLD",
                "ref": "CURRENT",
            }],
        }),
        encoding="utf-8",
    )
    (tmp_path / "_index.json").write_text(json.dumps({
        "updated_at": "2026-06-03T10:00:00+00:00",
        "sheets": {
            "2": {
                "sheet_id": 2,
                "operador": "OPERADOR",
                "date": "2026-06-03",
                "sheet_status": "extracted",
                "summary": {"match": 1, "no_match": 0, "na": 0, "total": 1},
                "engine_version": ENGINE_VERSION,
                "ok_rate": 1.0,
                "file": "2026-06-03/mismatched.json",
            },
        },
    }), encoding="utf-8")

    assert storage.load_sheet_cross_check(2) is None
    assert storage.load_sheet_cross_check(2, include_stale=True)["engine_version"] == "v-old"
    assert storage.iter_sheet_cross_checks() == []
    assert storage.iter_sheet_cross_checks(include_stale=True)[0]["engine_version"] == "v-old"
    inbox = storage.load_to_analisar()
    assert inbox["items"] == []
    assert inbox["stale_sheets"] == 1
    summary = storage.load_summary()
    assert summary["totals"] == {"match": 0, "no_match": 0, "na": 0, "total": 0}
    assert summary["n_sheets"] == 0
    assert summary["stale_sheets"] == 1


def test_cross_check_readers_skip_index_entries_without_file(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    (tmp_path / "_index.json").write_text(json.dumps({
        "updated_at": "2026-06-03T10:00:00+00:00",
        "sheets": {
            "3": {
                "sheet_id": 3,
                "operador": "OPERADOR",
                "date": "2026-06-03",
                "sheet_status": "extracted",
                "summary": {"match": 1, "no_match": 0, "na": 0, "total": 1},
                "engine_version": ENGINE_VERSION,
                "ok_rate": 1.0,
            },
        },
    }), encoding="utf-8")

    assert storage.load_sheet_cross_check(3) is None
    assert storage.iter_sheet_cross_checks() == []
    inbox = storage.load_to_analisar()
    assert inbox["items"] == []
    assert inbox["stale_sheets"] == 1
    summary = storage.load_summary()
    assert summary["totals"] == {"match": 0, "no_match": 0, "na": 0, "total": 0}
    assert summary["n_sheets"] == 0
    assert summary["stale_sheets"] == 1


def test_cross_check_readers_reject_payload_for_different_sheet_id(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    sheet_dir = tmp_path / "2026-06-03"
    sheet_dir.mkdir()
    (sheet_dir / "wrong-sheet.json").write_text(json.dumps({
        "sheet_id": 99,
        "engine_version": ENGINE_VERSION,
        "summary": {"match": 5, "no_match": 0, "na": 0, "total": 5},
        "rows": [],
        "header": {},
        "footer": {},
        "to_analisar": [{
            "section": "rows",
            "row_index": 0,
            "field": "modelo",
            "value": "WRONG",
            "ref": "PLAN",
        }],
    }), encoding="utf-8")
    (tmp_path / "_index.json").write_text(json.dumps({
        "updated_at": "2026-06-03T10:00:00+00:00",
        "sheets": {
            "4": {
                "sheet_id": 4,
                "operador": "OPERADOR",
                "date": "2026-06-03",
                "sheet_status": "extracted",
                "summary": {"match": 5, "no_match": 0, "na": 0, "total": 5},
                "engine_version": ENGINE_VERSION,
                "ok_rate": 1.0,
                "file": "2026-06-03/wrong-sheet.json",
            },
        },
    }), encoding="utf-8")

    assert storage.load_sheet_cross_check(4) is None
    assert storage.load_sheet_cross_check(4, include_stale=True)["sheet_id"] == 99
    assert storage.iter_sheet_cross_checks() == []
    inbox = storage.load_to_analisar()
    assert inbox["items"] == []
    assert inbox["stale_sheets"] == 1
    summary = storage.load_summary()
    assert summary["totals"] == {"match": 0, "no_match": 0, "na": 0, "total": 0}
    assert summary["n_sheets"] == 0
    assert summary["stale_sheets"] == 1


def test_remove_sheet_cross_check_handles_wrapped_index(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    stored = storage.store_cross_check(
        sheet_id=1,
        image_path="images/a.jpg",
        operador="OPERADOR",
        date_iso="2026-06-03",
        sheet_status="extracted",
        cross_check_result={
            "checked_at": "2026-06-03T10:01:00+00:00",
            "summary": {"match": 1, "no_match": 0, "na": 0, "total": 1},
            "rows": [],
            "header": {},
            "footer": {},
            "to_analisar": [],
        },
    )
    assert Path(stored["file"]).exists()

    storage.remove_sheet_cross_check(1)

    assert not Path(stored["file"]).exists()
    index = json.loads((tmp_path / "_index.json").read_text(encoding="utf-8"))
    assert index["sheets"] == {}
    summary = json.loads((tmp_path / "_summary.json").read_text(encoding="utf-8"))
    assert summary["n_sheets"] == 0


def test_to_analisar_preserves_section_and_field_path(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)

    storage.store_cross_check(
        sheet_id=1,
        image_path="images/a.jpg",
        operador="OPERADOR",
        date_iso="2026-06-03",
        sheet_status="extracted",
        cross_check_result={
            "checked_at": "2026-06-03T10:01:00+00:00",
            "summary": {"match": 0, "no_match": 1, "na": 0, "total": 1},
            "rows": [],
            "header": {},
            "footer": {},
            "to_analisar": [{
                "section": "header",
                "row_index": None,
                "field": "operador",
                "field_path": "header.operador",
                "value": "OPERADOR DESCONHECIDO",
                "ref": "",
                "ref_source": "ocr_raw",
                "reason": "Motor propõe valor muito diferente do OCR",
            }],
        },
    )

    inbox = storage.load_to_analisar()
    assert inbox["total"] == 1
    assert inbox["items"][0]["section"] == "header"
    assert inbox["items"][0]["field_path"] == "header.operador"
    assert inbox["items"][0]["ref_source"] == "ocr_raw"
    assert inbox["items"][0]["ref_value"] is None
    assert inbox["items"][0]["plan_value"] is None


def test_to_analisar_excludes_stale_engine_entries(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    old_dir = tmp_path / "2026-06-02"
    old_dir.mkdir()
    (old_dir / "old.json").write_text(json.dumps({
        "engine_version": "v-old",
        "to_analisar": [{
            "section": "rows",
            "row_index": 0,
            "field": "modelo",
            "field_path": "rows[0].modelo",
            "value": "OLD",
            "ref": "PLAN_OLD",
            "ref_source": "plan",
            "reason": "old",
        }],
    }), encoding="utf-8")
    stale = {
        "updated_at": "2026-06-03T10:00:00+00:00",
        "sheets": {
            "1": {
                "sheet_id": 1,
                "operador": "OPERADOR_ANTIGO",
                "date": "2026-06-02",
                "sheet_status": "extracted",
                "summary": {"match": 0, "no_match": 1, "na": 0, "total": 1},
                "engine_version": "v-old",
                "ok_rate": 0.0,
                "file": "2026-06-02/old.json",
            },
        },
    }
    (tmp_path / "_index.json").write_text(json.dumps(stale), encoding="utf-8")

    storage.store_cross_check(
        sheet_id=2,
        image_path="images/b.jpg",
        operador="OPERADOR",
        date_iso="2026-06-03",
        sheet_status="extracted",
        cross_check_result={
            "checked_at": "2026-06-03T10:01:00+00:00",
            "engine_version": ENGINE_VERSION,
            "summary": {"match": 0, "no_match": 1, "na": 0, "total": 1},
            "rows": [],
            "header": {},
            "footer": {},
            "to_analisar": [{
                "section": "rows",
                "row_index": 0,
                "field": "modelo",
                "field_path": "rows[0].modelo",
                "value": "CURRENT",
                "ref": "PLAN_CURRENT",
                "ref_source": "plan",
                "reason": "current",
            }],
        },
    )

    inbox = storage.load_to_analisar()

    assert inbox["engine_version"] == ENGINE_VERSION
    assert inbox["total"] == 1
    assert inbox["stale_sheets"] == 1
    assert inbox["items"][0]["value"] == "CURRENT"
    assert inbox["items"][0]["ref_value"] == "PLAN_CURRENT"
    assert inbox["items"][0]["plan_value"] == "PLAN_CURRENT"


def test_to_analisar_tolerates_legacy_index_missing_metadata(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    sheet_dir = tmp_path / "2026-06-03"
    sheet_dir.mkdir()
    (sheet_dir / "current.json").write_text(json.dumps({
        "sheet_id": 9,
        "operador": "OPERADOR_PAYLOAD",
        "date": "2026-06-03",
        "engine_version": ENGINE_VERSION,
        "to_analisar": [{
            "section": "rows",
            "row_index": 0,
            "field": "modelo",
            "field_path": "rows[0].modelo",
            "value": "OCR",
            "ref": "PLAN",
            "ref_source": "plan",
            "reason": "current",
        }],
    }), encoding="utf-8")
    (tmp_path / "_index.json").write_text(json.dumps({
        "updated_at": "2026-06-03T10:00:00+00:00",
        "sheets": {
            "9": {
                "engine_version": ENGINE_VERSION,
                "file": "2026-06-03/current.json",
                "summary": {"match": 0, "no_match": 1, "na": 0, "total": 1},
            },
        },
    }), encoding="utf-8")

    inbox = storage.load_to_analisar()

    assert inbox["total"] == 1
    assert inbox["items"][0]["sheet_id"] == 9
    assert inbox["items"][0]["operador"] == "OPERADOR_PAYLOAD"
    assert inbox["items"][0]["date"] == "2026-06-03"
    assert inbox["items"][0]["ref_value"] == "PLAN"


def test_load_summary_normalizes_legacy_string_counts(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    (tmp_path / "_summary.json").write_text(json.dumps({
        "totals": {"match": "8", "no_match": "2", "na": "10", "total": "20"},
        "by_day": {
            "2026-06-03": {"match": "8", "no_match": "2", "na": "10", "total": "20"},
        },
        "by_operador": {
            "OPERADOR": {"match": "8", "no_match": "2", "na": "10", "total": "20"},
        },
        "n_sheets": "1",
    }), encoding="utf-8")

    summary = storage.load_summary()

    assert summary["totals"] == {"match": 8, "no_match": 2, "na": 10, "total": 20}
    assert summary["by_day"]["2026-06-03"] == {
        "match": 8, "no_match": 2, "na": 10, "total": 20,
    }
    assert summary["by_operador"]["OPERADOR"] == {
        "match": 8, "no_match": 2, "na": 10, "total": 20,
    }
    assert summary["n_sheets"] == 1


def test_load_summary_ignores_stale_cache_when_index_is_missing(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    (tmp_path / "_summary.json").write_text(json.dumps({
        "engine_version": "v-old",
        "totals": {"match": "8", "no_match": "2", "na": "10", "total": "20"},
        "by_day": {
            "2026-06-03": {"match": "8", "no_match": "2", "na": "10", "total": "20"},
        },
        "by_operador": {
            "OPERADOR": {"match": "8", "no_match": "2", "na": "10", "total": "20"},
        },
        "n_sheets": "1",
    }), encoding="utf-8")

    summary = storage.load_summary()

    assert summary["engine_version"] == ENGINE_VERSION
    assert summary["totals"] == {"match": 0, "no_match": 0, "na": 0, "total": 0}
    assert summary["by_day"] == {}
    assert summary["by_operador"] == {}
    assert summary["n_sheets"] == 0
    assert summary["stale_sheets"] == 1


def test_load_to_analisar_normalizes_legacy_items(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    (tmp_path / "_to_analisar.json").write_text(json.dumps({
        "total": 1,
        "items": [{
            "sheet_id": 1,
            "operador": "OPERADOR",
            "date": "2026-06-03",
            "row_index": 2,
            "field": "modelo",
            "value": "OCR",
            "plan_value": "PLAN",
            "reason": "legacy",
        }],
    }), encoding="utf-8")

    inbox = storage.load_to_analisar()

    assert inbox["total"] == 1
    assert inbox["items"][0]["section"] == "rows"
    assert inbox["items"][0]["field_path"] == "rows[2].modelo"
    assert inbox["items"][0]["ref_value"] == "PLAN"
    assert inbox["items"][0]["plan_value"] == "PLAN"


def test_load_to_analisar_ignores_stale_cache_when_index_is_missing(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    (tmp_path / "_to_analisar.json").write_text(json.dumps({
        "engine_version": "v-old",
        "total": 2,
        "items": [{
            "sheet_id": 1,
            "operador": "OPERADOR",
            "date": "2026-06-03",
            "row_index": 2,
            "field": "modelo",
            "value": "OCR",
            "plan_value": "PLAN",
            "reason": "legacy",
        }],
    }), encoding="utf-8")

    inbox = storage.load_to_analisar()

    assert inbox["engine_version"] == ENGINE_VERSION
    assert inbox["total"] == 0
    assert inbox["items"] == []
    assert inbox["stale_sheets"] == 2


def test_load_to_analisar_keeps_legacy_plan_value_when_ref_empty(tmp_path, monkeypatch):
    monkeypatch.setattr(storage, "_base_dir", lambda: tmp_path)
    (tmp_path / "_to_analisar.json").write_text(json.dumps({
        "total": 1,
        "items": [{
            "sheet_id": 1,
            "operador": "OPERADOR",
            "date": "2026-06-03",
            "section": "rows",
            "row_index": 0,
            "field": "modelo",
            "field_path": "rows[0].modelo",
            "value": "OCR",
            "ref": "",
            "plan_value": "PLAN",
            "source": "plan",
            "reason": "legacy with empty ref",
        }],
    }), encoding="utf-8")

    inbox = storage.load_to_analisar()

    assert inbox["items"][0]["plan_value"] == "PLAN"
    assert inbox["items"][0]["ref_value"] == "PLAN"
    assert inbox["items"][0]["ref_source"] == "plan"


def test_update_script_protects_all_uploaded_refs():
    script = Path("scripts/ops/update.ps1").read_text(encoding="utf-8")

    assert "kanban_refs/04_Documentacao/plan_colunas_cpis.xlsx" in script
    assert "kanban_refs/04_Documentacao/StockSAP.xlsx" in script
    assert "kanban_refs/04_Documentacao/maquinas.xlsx" in script
    assert "kanban_refs/04_Documentacao/ListaColaboradores.xlsx" in script


# --- cross declarado (fase A) — plan_headers + entry["extra"] ---------------

@pytest.fixture()
def _clean_registry():
    """Deixa o registry byte-idêntico aos builtins no fim."""
    from app import templates_registry as reg
    yield reg
    reg.set_runtime_templates([])


def _declared_spec(column="pbase"):
    from app.web import template_store
    return template_store.spec_from_dict({
        "name": "u2_teste_declared",
        "setor_aliases": ["TESTE DECLARED"],
        "row_fields": ["of", "pbase"],
        "declared_cross": {"pbase": {"ref": "plan", "column": column,
                                     "cmp": "num", "tol": 2.0}},
    }, unidade_id=2, db_id=9)


_PLAN_ROW_PBASE = ["A", "2603977", "263348", "CAC4E10B", 1,
                   0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050, 250.0]


def test_plan_headers_always_published(tmp_path):
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    _write_plan(doc_dir / "plan_colunas_cpis.xlsx", [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ])
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    refs = watcher.force_reload()
    assert "of" in refs["plan_headers"]
    assert "comp" in refs["plan_headers"]
    assert refs["plan_headers"] == sorted(refs["plan_headers"])


def test_entry_extra_only_with_declared_templates(tmp_path, _clean_registry):
    reg = _clean_registry
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    plan = doc_dir / "plan_colunas_cpis.xlsx"
    _write_plan(plan, [_PLAN_ROW_PBASE], extra_cols=("pbase",))

    # SEM templates declarados: entry byte-idêntica (sem chave "extra")
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    refs = watcher.force_reload()
    entry = refs["of_to_entries"]["263348"][0]
    assert "extra" not in entry

    # COM template declarado instalado: re-mine apanha a coluna
    reg.set_runtime_templates([_declared_spec()])
    refs = watcher.force_reload()
    entry = refs["of_to_entries"]["263348"][0]
    assert entry["extra"] == {"pbase": 250.0}
    # só as colunas declaradas entram (disciplina de memória R225)
    assert set(entry["extra"]) == {"pbase"}


def test_entry_extra_propagates_to_inverted_indexes(tmp_path, _clean_registry):
    """Os índices invertidos (plan_by_cliente/ov) fazem cópia shallow da
    entry — o `extra` tem de viajar com eles (caminho holístico)."""
    reg = _clean_registry
    reg.set_runtime_templates([_declared_spec()])
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    _write_plan(doc_dir / "plan_colunas_cpis.xlsx",
                [_PLAN_ROW_PBASE], extra_cols=("pbase",))
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    refs = watcher.force_reload()

    by_cliente = refs["plan_by_cliente"].get("A") or []
    assert by_cliente and by_cliente[0].get("extra") == {"pbase": 250.0}
    by_ov = refs["plan_by_ov"].get("2603977") or []
    assert by_ov and by_ov[0].get("extra") == {"pbase": 250.0}


def test_declared_column_absent_from_plan_no_extra(tmp_path, _clean_registry):
    """Coluna declarada que não existe no plano → sem chave extra (NA em
    runtime), nada parte."""
    reg = _clean_registry
    reg.set_runtime_templates([_declared_spec(column="inexistente")])
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    _write_plan(doc_dir / "plan_colunas_cpis.xlsx", [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ])
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    refs = watcher.force_reload()
    assert "extra" not in refs["of_to_entries"]["263348"][0]


# --- R134 — robustez do upload do plano -----------------------------------

def test_miner_skips_blank_rows_without_truncating(tmp_path):
    """O minerador não deve terminar no primeiro `cliente` (col 0) vazio:
    antes (`if r[0] is None: break`) truncava OFs silenciosamente."""
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    plan = doc_dir / "plan_colunas_cpis.xlsx"
    _write_plan(plan, [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
        # cliente (col 0) em branco mas OF presente — antes parava aqui (break)
        [None, "2603978", "263349", "CAC4E10C", 2, 0, 0, 0, 0, 0, 0, 0, 3, 600, 200, 9000],
        # linha totalmente vazia no meio — deve ser saltada, não terminar
        [None] * 16,
        ["C", "2603980", "263350", "CAC4E10D", 3, 0, 0, 0, 0, 0, 0, 0, 3, 610, 210, 9100],
    ])
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    refs = watcher.force_reload()

    assert refs["stats"]["n_ofs"] == 3
    assert {"263348", "263349", "263350"} <= set(refs["of_to_entries"].keys())


def test_invalidate_index_cache_clears_all():
    from app.pipeline import scoring_engine
    scoring_engine._INDEX_CACHE[1] = {"loaded_at": "x"}
    scoring_engine._INDEX_CACHE[2] = {"loaded_at": "y"}
    scoring_engine.invalidate_index_cache()
    assert scoring_engine._INDEX_CACHE == {}


def test_force_reload_clears_scoring_index_cache(tmp_path):
    from app.pipeline import scoring_engine
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    _write_plan(doc_dir / "plan_colunas_cpis.xlsx", [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ])
    scoring_engine._INDEX_CACHE[424242] = {"loaded_at": "stale"}
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    watcher.force_reload()
    assert 424242 not in scoring_engine._INDEX_CACHE


def test_ref_watcher_loads_cliente_aliases_from_repo_root(tmp_path):
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    _write_plan(doc_dir / "plan_colunas_cpis.xlsx", [
        ["MTG", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ])
    lexicons = tmp_path / "lexicons"
    lexicons.mkdir()
    (lexicons / "cliente_aliases.json").write_text(
        json.dumps({"HTG": "MTG", "_comment": "ignored"}),
        encoding="utf-8",
    )

    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    refs = watcher.force_reload()

    assert refs["cliente_aliases"] == {"HTG": "MTG"}
    assert refs["cliente_aliases_mtime"] > 0


def test_get_refs_degrades_on_corrupt_file_without_crashing(tmp_path):
    """Um ficheiro on-disk corrupto não deve rebentar scans: get_refs mantém
    as refs anteriores e marca o mtime falhado para não martelar o miner."""
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    plan = doc_dir / "plan_colunas_cpis.xlsx"
    _write_plan(plan, [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ])
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    good_sha = watcher.force_reload()["plan_sha256"]
    assert good_sha

    plan.write_bytes(b"not a real xlsx")        # corromper
    watcher._refs["plan_mtime"] = 0.0            # forçar _needs_reload
    watcher._last_check_ts = 0.0                 # ultrapassar o debounce

    refs = watcher.get_refs()                    # não deve levantar
    assert refs["plan_sha256"] == good_sha       # manteve refs anteriores
    assert watcher._failed_mtimes is not None


def test_refs_upload_rolls_back_when_reload_fails(tmp_path, monkeypatch, log_path):
    """Se o reload falhar depois do os.replace, o ficheiro vivo é restaurado
    ao plano anterior e as refs ficam consistentes (não fica o ficheiro novo
    ativo com refs antigas)."""
    doc_dir = tmp_path / "docs"
    doc_dir.mkdir()
    _write_plan(doc_dir / "plan_colunas_cpis.xlsx", [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
    ])
    watcher = ref_watcher.RefWatcher(doc_dir=doc_dir, repo_root=tmp_path)
    old_sha = watcher.force_reload()["plan_sha256"]
    assert old_sha

    monkeypatch.setattr(main, "get_watcher", lambda: watcher)

    def _boom(*_a, **_k):
        raise RuntimeError("mine boom")

    monkeypatch.setattr(ref_watcher, "_mine_from_excel", _boom)

    upload_plan = tmp_path / "upload_plan.xlsx"
    _write_plan(upload_plan, [
        ["A", "2603977", "263348", "CAC4E10B", 1, 0, 0, 0, 0, 0, 0, 0, 4, 659, 242, 11050],
        ["B", "2603978", "263349", "CAC4E10C", 2, 0, 0, 0, 0, 0, 0, 0, 3, 600, 200, 9000],
    ])
    client = TestClient(main.app)
    with upload_plan.open("rb") as f:
        resp = client.post(
            "/refs/upload",
            data={"kind": "plan"},
            files={"file": ("plan_colunas_cpis.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            follow_redirects=False,
        )

    assert resp.status_code == 303
    assert "err" in resp.headers["location"]
    # ficheiro vivo restaurado ao plano antigo
    assert ref_watcher.file_sha256(watcher.plan_path) == old_sha
    # backup transitório consumido pelo restore (os.replace)
    assert not (doc_dir / "plan_colunas_cpis.prevbak.xlsx").exists()
